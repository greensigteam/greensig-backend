from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter, SearchFilter
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import timedelta

from .models import TypeReclamation, Urgence, Reclamation, HistoriqueReclamation, SatisfactionClient
from api_users.models import Equipe
from django.db import transaction
from .serializers import (
    TypeReclamationSerializer,
    UrgenceSerializer,
    ReclamationListSerializer,
    ReclamationDetailSerializer,
    ReclamationCreateSerializer,
    HistoriqueReclamationSerializer,
    SatisfactionClientSerializer
)
from .permissions import IsReclamationCreatorOrTeamReader
from .filters import ReclamationFilter

class TypeReclamationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet pour les types de réclamation.
    Lecture seule. Retourne uniquement les actifs.
    """
    queryset = TypeReclamation.objects.filter(actif=True)
    serializer_class = TypeReclamationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [OrderingFilter]
    ordering_fields = ['nom_reclamation', 'categorie']


class UrgenceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet pour les niveaux d'urgence.
    Lecture seule. Ordonné par priorité.
    """
    queryset = Urgence.objects.all().order_by('ordre')
    serializer_class = UrgenceSerializer
    permission_classes = [permissions.IsAuthenticated]


class ReclamationViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des réclamations.

    Permissions:
    - Admin/Staff : accès à TOUTES les réclamations
    - Chef d'équipe : accès uniquement à SES réclamations créées
    - Client : accès uniquement à SES réclamations (créées par lui ou liées à lui)
    - Tout utilisateur authentifié peut créer une réclamation
    """
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_class = ReclamationFilter
    ordering_fields = ['date_creation']
    search_fields = ['numero_reclamation', 'description']

    def get_queryset(self):
        user = self.request.user
        # Optimisation: select_related pour éviter N+1 queries
        queryset = Reclamation.objects.filter(actif=True).select_related(
            'createur',
            'client__utilisateur',
            'site',
            'zone',
            'urgence',
            'type_reclamation',
            'equipe_affectee',
            'equipe_affectee__site__superviseur__utilisateur',
            'intervention_refusee_par',  # Pour afficher le nom de celui qui a refusé l'intervention (client)
            'rejetee_par',  # Pour afficher le nom de l'admin qui a rejeté
            'cloture_refusee_par'  # Pour afficher le nom du client qui a refusé la clôture
        )

        # Prefetch pour le détail (historique, photos, taches, satisfaction)
        if self.action in ['retrieve', 'suivi']:
            from django.db.models import Prefetch
            from api_planification.models import Tache
            queryset = queryset.prefetch_related(
                'historique__auteur',
                'photos',
                'satisfaction',
                Prefetch(
                    'taches_correctives',
                    queryset=Tache.objects.select_related(
                        'id_type_tache', 'id_equipe'
                    ).prefetch_related('equipes')
                )
            )

        # Admin / Staff : accès à TOUTES les réclamations
        if user.is_staff or user.is_superuser:
            return queryset

        # Superviseur : accès aux réclamations de ses sites
        if hasattr(user, 'superviseur_profile'):
            try:
                superviseur = user.superviseur_profile

                # Le superviseur voit:
                # 1. Les réclamations sur les sites qu'il supervise
                # 2. Les réclamations affectées à ses équipes
                # 3. Les réclamations qu'il a créées lui-même
                return queryset.filter(
                    Q(site__superviseur=superviseur) |  # Sites supervisés
                    Q(equipe_affectee__site__superviseur=superviseur) |  # Équipes de ses sites
                    Q(createur=user)  # Ses propres réclamations
                ).distinct()
            except AttributeError:
                return queryset.filter(createur=user)

        # Client : accès à ses réclamations uniquement (créées par lui ou liées à sa structure si visible_client=True)
        if hasattr(user, 'client_profile'):
            # Le client voit :
            # 1. Les réclamations qu'il a créées lui-même (toujours visibles)
            # 2. Les réclamations de sa structure SEULEMENT si visible_client=True
            return queryset.filter(
                Q(createur=user) |
                Q(structure_client=user.client_profile.structure, visible_client=True)
            )

        # Tout autre utilisateur : réclamations qu'il a créées
        return queryset.filter(createur=user)

    def perform_destroy(self, instance):
        """Suppression réelle de la réclamation.

        Les tâches correctives associées sont supprimées automatiquement (CASCADE).
        """
        instance.delete()

    def get_serializer_class(self):
        if self.action == 'list':
            return ReclamationListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return ReclamationCreateSerializer
        return ReclamationDetailSerializer

    def perform_create(self, serializer):
        """
        Assigner automatiquement le créateur (utilisateur connecté).
        Si c'est un client, on associe aussi le client_profile.
        Remplir automatiquement date_constatation si non fournie.
        Déduire structure_client du site si non fourni.
        """
        user = self.request.user
        extra_kwargs = {'createur': user}

        # Si l'utilisateur est un client, on associe la structure cliente
        if hasattr(user, 'client_profile'):
            extra_kwargs['client'] = user.client_profile  # Legacy
            if user.client_profile.structure:
                extra_kwargs['structure_client'] = user.client_profile.structure

        # Remplir automatiquement date_constatation avec la date actuelle si non fournie
        if 'date_constatation' not in serializer.validated_data or not serializer.validated_data['date_constatation']:
            extra_kwargs['date_constatation'] = timezone.now()

        reclamation = serializer.save(**extra_kwargs, _current_user=user)

        # Déduire structure_client du site si pas déjà défini
        if not reclamation.structure_client and reclamation.site and reclamation.site.structure_client:
            reclamation.structure_client = reclamation.site.structure_client
            reclamation.save(update_fields=['structure_client'])

        # Création de l'historique initial
        HistoriqueReclamation.objects.create(
            reclamation=reclamation,
            statut_precedent=None,
            statut_nouveau=reclamation.statut,
            auteur=user if user.is_authenticated else None,
            commentaire="Création de la réclamation"
        )

    def perform_update(self, serializer):
        # Récupération de l'instance avant modification pour comparer le statut
        instance = self.get_object()
        old_statut = instance.statut
        
        updated_instance = serializer.save(_current_user=self.request.user)
        
        # Si le statut a changé, on ajoute une entrée dans l'historique
        if old_statut != updated_instance.statut:
            HistoriqueReclamation.objects.create(
                reclamation=updated_instance,
                statut_precedent=old_statut,
                statut_nouveau=updated_instance.statut,
                auteur=self.request.user if self.request.user.is_authenticated else None,
                commentaire=f"Changement de statut : {old_statut} -> {updated_instance.statut}"
            )


    @action(detail=True, methods=['get'])
    def suivi(self, request, pk=None):
        """
        Endpoint spécifique pour le suivi temps réel.
        Retourne la timeline et le statut (à enrichir selon besoin).
        """
        reclamation = self.get_object()
        serializer = ReclamationDetailSerializer(reclamation)
        return Response(serializer.data)

    @action(detail=True, methods=['put'])
    def assignation(self, request, pk=None):
        """
        Endpoint pour assigner une réclamation à une équipe.
        Met à jour la réclamation, les tâches associées, et l'historique.
        """
        reclamation = self.get_object()
        equipe_id = request.data.get('equipe_id')
        
        if not equipe_id:
             return Response({"error": "L'ID de l'équipe est requis."}, status=status.HTTP_400_BAD_REQUEST)
             
        try:
             equipe = Equipe.objects.get(pk=equipe_id)
        except Equipe.DoesNotExist:
             return Response({"error": "Équipe non trouvée."}, status=status.HTTP_404_NOT_FOUND)
             
        # Transaction atomique pour garantir la cohérence
        with transaction.atomic():
             old_statut = reclamation.statut
             old_equipe = reclamation.equipe_affectee
             
             # 1. Mise à jour de la Réclamation
             reclamation.equipe_affectee = equipe
             # Si nouvelle -> Prise en compte
             if reclamation.statut == 'NOUVELLE':
                 reclamation.statut = 'PRISE_EN_COMPTE'
             
             reclamation._current_user = request.user
             reclamation.save()
             
             # 2. Propagation aux Tâches (via le related_name 'taches_correctives')
             # On met à jour toutes les tâches correctives liées
             updated_count = reclamation.taches_correctives.update(id_equipe=equipe)
             
             # 3. Création de l'historique
             commentaire = f"Assignation à l'équipe: {equipe.nom_equipe}"
             if old_equipe:
                 commentaire += f" (anciennement: {old_equipe.nom_equipe})"
             
             if old_statut != reclamation.statut:
                 commentaire += f" - Statut changé de {old_statut} à {reclamation.statut}"

             HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau=reclamation.statut,
                auteur=request.user if request.user.is_authenticated else None,
                commentaire=commentaire
            )
            
        serializer = ReclamationDetailSerializer(reclamation)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cloturer(self, request, pk=None):
        """
        Endpoint pour proposer la clôture d'une réclamation (ADMIN/SUPERVISEUR).

        Nouveau workflow (validation obligatoire par le créateur):
        1. Admin/Superviseur propose la clôture → statut EN_ATTENTE_VALIDATION_CLOTURE
        2. Le créateur valide via /valider_cloture/ → statut CLOTUREE

        Garde-fous:
        - La réclamation doit être en statut EN_COURS ou RESOLUE
        - Il doit y avoir au moins une tâche corrective
        - Toutes les tâches doivent être terminées (statut TERMINEE)
        - Toutes les tâches doivent être validées (etat_validation VALIDEE)
        """
        reclamation = self.get_object()

        # Vérification 1: seuls ADMIN/SUPERVISEUR peuvent proposer la clôture
        user = request.user
        is_admin = user.is_staff or user.is_superuser
        is_superviseur = hasattr(user, 'superviseur_profile')

        if not (is_admin or is_superviseur):
            return Response(
                {"error": "Seuls les administrateurs et superviseurs peuvent proposer la clôture."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification 2: Le statut doit être approprié
        statuts_valides = ['EN_COURS', 'RESOLUE']
        if reclamation.statut not in statuts_valides:
            return Response(
                {"error": f"La réclamation doit être en cours ou résolue pour proposer la clôture. Statut actuel: {reclamation.get_statut_display()}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérification 3: Il doit y avoir au moins une tâche corrective
        taches = reclamation.taches_correctives.all()
        nombre_taches = taches.count()

        if nombre_taches == 0:
            return Response(
                {"error": "Impossible de clôturer une réclamation sans tâche corrective. Veuillez d'abord créer et réaliser au moins une intervention."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérification 4: Toutes les tâches doivent être terminées
        taches_non_terminees = taches.exclude(statut='TERMINEE')
        if taches_non_terminees.exists():
            noms_taches = [f"#{t.id}" for t in taches_non_terminees[:5]]
            return Response({
                "error": f"Toutes les tâches doivent être terminées avant de proposer la clôture.",
                "taches_non_terminees": list(taches_non_terminees.values('id', 'statut', 'id_type_tache__nom_tache')[:5]),
                "nombre_non_terminees": taches_non_terminees.count()
            }, status=status.HTTP_400_BAD_REQUEST)

        # Vérification 5: Toutes les tâches doivent être validées
        taches_non_validees = taches.exclude(etat_validation='VALIDEE')
        if taches_non_validees.exists():
            return Response({
                "error": f"Toutes les tâches doivent être validées avant de proposer la clôture.",
                "taches_non_validees": list(taches_non_validees.values('id', 'etat_validation', 'id_type_tache__nom_tache')[:5]),
                "nombre_non_validees": taches_non_validees.count()
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            old_statut = reclamation.statut
            reclamation.statut = 'EN_ATTENTE_VALIDATION_CLOTURE'
            reclamation.cloture_proposee_par = user
            reclamation.date_proposition_cloture = timezone.now()
            
            reclamation._current_user = user
            reclamation.save()

            # Historique
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='EN_ATTENTE_VALIDATION_CLOTURE',
                auteur=request.user,
                commentaire=f"Proposition de clôture par {user.get_full_name() or user.email}"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Clôture proposée avec succès. En attente de validation par le créateur.',
            'reclamation': serializer.data
        })

    @action(detail=True, methods=['post'])
    def valider_cloture(self, request, pk=None):
        """
        Endpoint pour valider la clôture d'une réclamation (CREATEUR uniquement).

        Workflow:
        - Seul le créateur de la réclamation peut valider la clôture
        - La réclamation doit être en statut EN_ATTENTE_VALIDATION_CLOTURE
        - Passage au statut CLOTUREE définitif
        """
        reclamation = self.get_object()
        user = request.user

        # Vérification 1: L'utilisateur doit être le créateur de la réclamation
        if reclamation.createur != user:
            return Response(
                {"error": "Seul le créateur de la réclamation peut valider sa clôture."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification 2: La réclamation doit être en attente de validation
        if reclamation.statut != 'EN_ATTENTE_VALIDATION_CLOTURE':
            return Response(
                {"error": f"La réclamation doit être en attente de validation de clôture. Statut actuel: {reclamation.get_statut_display()}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            old_statut = reclamation.statut
            reclamation.statut = 'CLOTUREE'
            
            reclamation._current_user = user
            reclamation.save()

            # Historique
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='CLOTUREE',
                auteur=user,
                commentaire=f"Clôture validée par le créateur {user.get_full_name() or user.email}"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Clôture validée avec succès. La réclamation est définitivement clôturée.',
            'reclamation': serializer.data
        })

    @action(detail=True, methods=['post'])
    def refuser_cloture(self, request, pk=None):
        """
        Endpoint pour refuser la clôture d'une réclamation (CREATEUR uniquement).

        Workflow:
        - Seul le créateur de la réclamation peut refuser la clôture
        - La réclamation doit être en statut EN_ATTENTE_VALIDATION_CLOTURE
        - Un commentaire de refus est OBLIGATOIRE
        - Retour au statut RESOLUE pour permettre de nouvelles interventions
        """
        reclamation = self.get_object()
        user = request.user

        # Vérification 1: L'utilisateur doit être le créateur de la réclamation
        if reclamation.createur != user:
            return Response(
                {"error": "Seul le créateur de la réclamation peut refuser sa clôture."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification 2: La réclamation doit être en attente de validation
        if reclamation.statut != 'EN_ATTENTE_VALIDATION_CLOTURE':
            return Response(
                {"error": f"La réclamation doit être en attente de validation de clôture. Statut actuel: {reclamation.get_statut_display()}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérification 3: Le commentaire de refus est OBLIGATOIRE
        commentaire_refus = request.data.get('commentaire_refus', '').strip()
        if not commentaire_refus:
            return Response(
                {"error": "Un commentaire expliquant le refus est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            old_statut = reclamation.statut

            # Retour au statut RESOLUE (état avant proposition de clôture)
            reclamation.statut = 'RESOLUE'

            # Sauvegarde du refus de clôture (champs dédiés)
            reclamation.cloture_refusee_par = user
            reclamation.date_refus_cloture = timezone.now()
            reclamation.commentaire_refus_cloture = commentaire_refus

            # Réinitialisation des champs de proposition de clôture
            reclamation.cloture_proposee_par = None
            reclamation.date_proposition_cloture = None

            reclamation._current_user = user
            reclamation.save()

            # Historique avec le commentaire de refus
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='RESOLUE',
                auteur=user,
                commentaire=f"Clôture refusée par le créateur {user.get_full_name() or user.email}. Motif: {commentaire_refus}"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Clôture refusée. La réclamation retourne au statut "Résolue" pour permettre de nouvelles interventions.',
            'reclamation': serializer.data
        })

    @action(detail=True, methods=['post'])
    def refuser_intervention(self, request, pk=None):
        """
        Endpoint pour qu'un CLIENT refuse une intervention effectuée.

        Workflow:
        - Le client (créateur ou lié à la structure) peut refuser une intervention
        - La réclamation doit être en statut RESOLUE ou EN_ATTENTE_VALIDATION_CLOTURE
        - Un commentaire expliquant le refus est OBLIGATOIRE
        - Passage au statut INTERVENTION_REFUSEE
        - La réclamation peut ensuite être reprise pour une nouvelle intervention
        """
        reclamation = self.get_object()
        user = request.user

        # Vérification 1: L'utilisateur doit être le créateur ou un client de la structure
        is_creator = reclamation.createur == user
        is_structure_client = (
            hasattr(user, 'client_profile') and
            reclamation.structure_client and
            user.client_profile.structure == reclamation.structure_client
        )

        if not (is_creator or is_structure_client):
            return Response(
                {"error": "Seul le créateur de la réclamation ou un client de la structure peut refuser l'intervention."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification 2: La réclamation doit être en statut approprié (intervention terminée)
        statuts_valides = ['RESOLUE', 'EN_ATTENTE_VALIDATION_CLOTURE']
        if reclamation.statut not in statuts_valides:
            return Response(
                {"error": f"L'intervention ne peut être refusée que lorsque la réclamation est en statut 'Résolue' ou 'En attente de validation'. Statut actuel: {reclamation.get_statut_display()}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérification 3: Le commentaire de refus est OBLIGATOIRE
        motif_refus = request.data.get('motif_refus', '').strip()
        if not motif_refus:
            return Response(
                {"error": "Un commentaire expliquant le motif du refus est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            old_statut = reclamation.statut

            # Mise à jour du statut
            reclamation.statut = 'INTERVENTION_REFUSEE'

            # Enregistrement des détails du refus
            reclamation.intervention_refusee_par = user
            reclamation.date_refus_intervention = timezone.now()
            reclamation.motif_refus_intervention = motif_refus
            reclamation.nombre_refus = (reclamation.nombre_refus or 0) + 1

            # Réinitialisation des champs de proposition de clôture
            reclamation.cloture_proposee_par = None
            reclamation.date_proposition_cloture = None

            reclamation._current_user = user
            reclamation.save()

            # Historique avec le motif du refus
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='INTERVENTION_REFUSEE',
                auteur=user,
                commentaire=f"Intervention refusée par {user.get_full_name() or user.email}. Motif: {motif_refus}"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Intervention refusée. Une nouvelle intervention peut être planifiée.',
            'reclamation': serializer.data
        })

    @action(detail=True, methods=['post'])
    def reprendre_intervention(self, request, pk=None):
        """
        Endpoint pour reprendre une réclamation après refus d'intervention.

        Workflow:
        - ADMIN/SUPERVISEUR uniquement
        - La réclamation doit être en statut INTERVENTION_REFUSEE
        - Passage au statut EN_COURS pour permettre une nouvelle intervention
        """
        reclamation = self.get_object()
        user = request.user

        # Vérification: seuls ADMIN/SUPERVISEUR peuvent reprendre
        is_admin = user.is_staff or user.is_superuser
        is_superviseur = hasattr(user, 'superviseur_profile')

        if not (is_admin or is_superviseur):
            return Response(
                {"error": "Seuls les administrateurs et superviseurs peuvent reprendre une intervention."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification: la réclamation doit être en statut INTERVENTION_REFUSEE
        if reclamation.statut != 'INTERVENTION_REFUSEE':
            return Response(
                {"error": f"Seules les réclamations avec intervention refusée peuvent être reprises. Statut actuel: {reclamation.get_statut_display()}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            old_statut = reclamation.statut
            reclamation.statut = 'EN_COURS'

            reclamation._current_user = user
            reclamation.save()

            # Historique
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='EN_COURS',
                auteur=user,
                commentaire=f"Intervention reprise par {user.get_full_name() or user.email} suite au refus client"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Réclamation reprise. Une nouvelle intervention peut être effectuée.',
            'reclamation': serializer.data
        })

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """
        Endpoint pour rejeter une réclamation (ADMIN uniquement).

        Workflow:
        - Seul l'admin peut rejeter une réclamation
        - La réclamation ne doit pas être CLOTUREE
        - Une justification est OBLIGATOIRE
        - Passage au statut REJETEE définitif
        """
        reclamation = self.get_object()
        user = request.user

        # Vérification 1: Seul l'admin peut rejeter
        if not (user.is_staff or user.is_superuser):
            return Response(
                {"error": "Seuls les administrateurs peuvent rejeter une réclamation."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Vérification 2: La réclamation ne doit pas être déjà clôturée
        if reclamation.statut == 'CLOTUREE':
            return Response(
                {"error": "Une réclamation clôturée ne peut pas être rejetée."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérification 3: La justification est OBLIGATOIRE
        justification = request.data.get('justification', '').strip()
        if not justification:
            return Response(
                {"error": "Une justification est obligatoire pour rejeter une réclamation."},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            old_statut = reclamation.statut
            reclamation.statut = 'REJETEE'
            reclamation.justification_rejet = justification
            reclamation.rejetee_par = user
            reclamation.date_rejet = timezone.now()

            reclamation._current_user = user
            reclamation.save()

            # Historique
            HistoriqueReclamation.objects.create(
                reclamation=reclamation,
                statut_precedent=old_statut,
                statut_nouveau='REJETEE',
                auteur=user,
                commentaire=f"Réclamation rejetée par {user.get_full_name() or user.email}. Motif: {justification}"
            )

        serializer = ReclamationDetailSerializer(reclamation)
        return Response({
            'message': 'Réclamation rejetée avec succès.',
            'reclamation': serializer.data
        })

    @action(detail=False, methods=['post'], url_path='detect-site')
    def detect_site(self, request):
        """
        Endpoint pour détecter le site à partir d'une géométrie.
        Utilisé par le frontend pour afficher le site avant création.

        Body: { "geometry": { "type": "Point", "coordinates": [lng, lat] } }
        Returns: { "site_id": 1, "site_nom": "Nom du site" } ou { "site_id": null }
        """
        from django.contrib.gis.geos import GEOSGeometry
        from api.models import Site, SousSite
        import json

        geometry_data = request.data.get('geometry')
        if not geometry_data:
            return Response({"error": "Geometry is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Convertir le GeoJSON en objet GEOS
            geom = GEOSGeometry(json.dumps(geometry_data), srid=4326)
        except Exception as e:
            return Response({"error": f"Invalid geometry: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        # 1. D'abord chercher un SousSite qui contient la géométrie
        found_zone = SousSite.objects.filter(geometrie__intersects=geom).select_related('site').first()
        if found_zone and found_zone.site:
            return Response({
                "site_id": found_zone.site.id,
                "site_nom": found_zone.site.nom_site,
                "zone_id": found_zone.id,
                "zone_nom": found_zone.nom
            })

        # 2. Sinon chercher un Site dont l'emprise contient la géométrie
        found_site = Site.objects.filter(geometrie_emprise__intersects=geom).first()
        if found_site:
            return Response({
                "site_id": found_site.id,
                "site_nom": found_site.nom_site,
                "zone_id": None,
                "zone_nom": None
            })

        # 3. Aucun site trouvé
        return Response({
            "site_id": None,
            "site_nom": None,
            "zone_id": None,
            "zone_nom": None
        })

    @action(detail=False, methods=['get'])
    def map(self, request):
        """
        Endpoint pour afficher les réclamations sur la carte principale.

        Retourne un GeoJSON FeatureCollection avec les réclamations:
        - Qui ont une localisation (geometry non null)
        - Qui ne sont pas CLOTUREE ni REJETEE (réclamations archivées masquées)

        Note: Les mini-cartes dans les fiches détails affichent toutes les
        réclamations (y compris CLOTUREE/REJETEE) pour la traçabilité.

        Query params optionnels:
        - bbox: Bounding box au format "west,south,east,north"
        - statut: Filtrer par statut spécifique

        Chaque feature inclut:
        - La géométrie de la réclamation
        - Les propriétés: id, numero, statut, urgence, couleur_statut
        """
        from django.contrib.gis.geos import Polygon
        import json

        # Couleurs par statut (du plus urgent au moins urgent)
        STATUT_COLORS = {
            'NOUVELLE': '#ef4444',        # Rouge vif - nouvelle réclamation
            'PRISE_EN_COMPTE': '#f97316', # Orange - en cours de prise en compte
            'EN_COURS': '#eab308',         # Jaune - en cours de traitement
            'RESOLUE': '#22c55e',          # Vert - résolue, en attente de clôture
            'EN_ATTENTE_VALIDATION_CLOTURE': '#10b981', # Vert clair - en attente validation
            'INTERVENTION_REFUSEE': '#dc2626', # Rouge foncé - intervention refusée par client
            'REJETEE': '#6b7280',          # Gris - rejetée
            # CLOTUREE ne sont pas affichées sur la carte principale
            # (mais visibles dans les mini-cartes des fiches détails)
        }

        # Base queryset - exclure les réclamations clôturées/rejetées et celles sans localisation
        # Note: Les mini-cartes dans les fiches détails affichent toutes les réclamations
        queryset = self.get_queryset().exclude(
            statut__in=['CLOTUREE', 'REJETEE']  # Exclure les réclamations archivées
        ).exclude(
            localisation__isnull=True
        ).select_related(
            'urgence', 'type_reclamation', 'site', 'zone'
        )

        # Filtre par statut si spécifié
        statut_filter = request.query_params.get('statut')
        if statut_filter:
            queryset = queryset.filter(statut=statut_filter)

        # Filtre par bbox si fourni
        bbox_str = request.query_params.get('bbox')
        if bbox_str:
            try:
                west, south, east, north = map(float, bbox_str.split(','))
                bbox_polygon = Polygon.from_bbox((west, south, east, north))
                queryset = queryset.filter(localisation__intersects=bbox_polygon)
            except (ValueError, AttributeError):
                pass  # Ignorer bbox invalide

        # Construire le GeoJSON
        features = []
        for rec in queryset[:200]:  # Limiter à 200 réclamations
            # Convertir la géométrie en GeoJSON
            geom_json = json.loads(rec.localisation.geojson)

            feature = {
                'type': 'Feature',
                'id': f'reclamation-{rec.id}',
                'geometry': geom_json,
                'properties': {
                    'id': rec.id,
                    'object_type': 'Reclamation',
                    'numero_reclamation': rec.numero_reclamation,
                    'statut': rec.statut,
                    'statut_display': dict(rec.STATUT_CHOICES).get(rec.statut, rec.statut),
                    'couleur_statut': STATUT_COLORS.get(rec.statut, '#6b7280'),
                    'urgence': rec.urgence.niveau_urgence if rec.urgence else None,
                    'urgence_couleur': rec.urgence.couleur if rec.urgence else None,
                    'type_reclamation': rec.type_reclamation.nom_reclamation if rec.type_reclamation else None,
                    'type_reclamation_symbole': rec.type_reclamation.symbole if rec.type_reclamation else None,
                    'type_reclamation_categorie': rec.type_reclamation.categorie if rec.type_reclamation else None,
                    'description': rec.description[:100] + '...' if rec.description and len(rec.description) > 100 else rec.description,
                    'site': rec.site.id if rec.site else None,  # ✅ ID du site (pour filtrage)
                    'site_nom': rec.site.nom_site if rec.site else None,
                    'zone_nom': rec.zone.nom if rec.zone else None,
                    'date_creation': rec.date_creation.isoformat() if rec.date_creation else None,
                }
            }
            features.append(feature)

        return Response({
            'type': 'FeatureCollection',
            'features': features,
            'count': len(features),
            'statut_colors': STATUT_COLORS
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Endpoint pour les statistiques des réclamations (User 6.6.14).
        Filtres: date_debut, date_fin, site, zone, type_reclamation
        """
        queryset = self.get_queryset()
        
        # Filtres temporels
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        if date_debut:
            queryset = queryset.filter(date_creation__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date_creation__lte=date_fin)
        
        # Filtres géographiques
        site_id = request.query_params.get('site')
        zone_id = request.query_params.get('zone')
        if site_id:
            queryset = queryset.filter(site_id=site_id)
        if zone_id:
            queryset = queryset.filter(zone_id=zone_id)
        
        # Filtre type
        type_id = request.query_params.get('type_reclamation')
        if type_id:
            queryset = queryset.filter(type_reclamation_id=type_id)
        
        # Calcul des KPIs
        stats_data = {
            'total': queryset.count(),
            'par_statut': dict(queryset.values('statut').annotate(count=Count('id')).values_list('statut', 'count')),
            'par_type': list(queryset.values('type_reclamation__nom_reclamation').annotate(count=Count('id'))),
            'par_urgence': list(queryset.values('urgence__niveau_urgence').annotate(count=Count('id'))),
            'par_zone': list(queryset.filter(zone__isnull=False).values('zone__nom').annotate(count=Count('id'))),
        }
        
        # Délai moyen de traitement (pour les clôturées)
        cloturees = queryset.filter(statut='CLOTUREE', date_cloture_reelle__isnull=False)
        if cloturees.exists():
            delais = []
            for rec in cloturees:
                if rec.date_creation and rec.date_cloture_reelle:
                    delta = rec.date_cloture_reelle - rec.date_creation
                    delais.append(delta.total_seconds() / 3600)  # en heures
            if delais:
                stats_data['delai_moyen_heures'] = sum(delais) / len(delais)
        
        # Taux de satisfaction
        satisfactions = SatisfactionClient.objects.filter(reclamation__in=queryset)
        if satisfactions.exists():
            stats_data['satisfaction_moyenne'] = satisfactions.aggregate(Avg('note'))['note__avg']
            stats_data['nombre_evaluations'] = satisfactions.count()
        
        return Response(stats_data)


# ==============================================================================
# VIEWSET - SATISFACTION CLIENT (User 6.6.13)
# ==============================================================================

class SatisfactionClientViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des évaluations de satisfaction client.

    Permissions:
    - **Création/Modification/Suppression**: SEUL le créateur de la réclamation
    - **Lecture**: Le créateur + superviseur de l'équipe + ADMIN (retour d'expérience)
    """
    queryset = SatisfactionClient.objects.all()
    serializer_class = SatisfactionClientSerializer
    permission_classes = [IsReclamationCreatorOrTeamReader]

    def get_queryset(self):
        """
        Retourne les évaluations accessibles par l'utilisateur:
        - Ses propres évaluations (créateur)
        - Les évaluations des réclamations traitées par ses équipes (superviseur)
        - Toutes les évaluations (ADMIN/Staff)
        """
        user = self.request.user

        # ADMIN/Staff: accès à toutes les évaluations
        if user.is_staff or user.is_superuser:
            queryset = SatisfactionClient.objects.all()
        else:
            # Filtres cumulatifs (OR)
            from django.db.models import Q

            filters = Q(reclamation__createur=user)  # Évaluations créées par l'utilisateur

            # Si superviseur: ajouter les évaluations des réclamations de ses équipes
            if hasattr(user, 'superviseur_profile'):
                superviseur = user.superviseur_profile
                # Réclamations traitées par les équipes gérées par ce superviseur
                filters |= Q(reclamation__equipe_affectee__site__superviseur=superviseur)

            queryset = SatisfactionClient.objects.filter(filters)

        # Optimisation: select_related pour éviter N+1
        queryset = queryset.select_related(
            'reclamation',
            'reclamation__createur',
            'reclamation__equipe_affectee',
            'reclamation__equipe_affectee__site__superviseur'
        )

        # Filtre optionnel par reclamation_id
        reclamation_id = self.request.query_params.get('reclamation')
        if reclamation_id:
            queryset = queryset.filter(reclamation_id=reclamation_id)

        return queryset

    def create(self, request, *args, **kwargs):
        """
        Surcharge du create pour gérer l'unicité (OneToOne) de manière gracieuse.
        Si une évaluation existe déjà, on la met à jour.

        Validation: SEUL le créateur de la réclamation peut créer/modifier son évaluation.
        """
        reclamation_id = request.data.get('reclamation')
        if not reclamation_id:
            return Response({"detail": "Le champ réclamation est requis."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Vérifier que la réclamation existe
            reclamation = Reclamation.objects.get(pk=reclamation_id)
        except Reclamation.DoesNotExist:
            return Response({"detail": "Réclamation non trouvée."}, status=status.HTTP_404_NOT_FOUND)

        # VALIDATION CRITIQUE: Vérifier que l'utilisateur est bien le créateur de la réclamation
        if reclamation.createur != request.user:
            return Response(
                {"detail": "Vous ne pouvez évaluer que vos propres réclamations."},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            # On utilise update_or_create pour éviter les erreurs 400 de doublons
            satisfaction, created = SatisfactionClient.objects.update_or_create(
                reclamation_id=reclamation_id,
                defaults={
                    'note': request.data.get('note'),
                    'commentaire': request.data.get('commentaire', '')
                }
            )

            # Envoyer une notification pour informer de l'évaluation
            from api.services.notifications import NotificationService
            NotificationService.notify_satisfaction_evaluee(satisfaction, acteur=request.user)

            serializer = self.get_serializer(satisfaction)
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==============================================================================
# VUE EXPORT EXCEL DES RÉCLAMATIONS
# ==============================================================================

from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO


class ReclamationExportExcelView(APIView):
    """
    Vue pour l'export Excel des réclamations avec horodatage de toutes les étapes.

    Paramètres de requête:
    - statut: filtrer par statut
    - urgence: filtrer par urgence
    - type_reclamation: filtrer par type
    - site: filtrer par site
    - date_debut: date de création minimum
    - date_fin: date de création maximum
    """
    permission_classes = [permissions.IsAuthenticated]

    # Labels français pour les statuts
    STATUT_LABELS = {
        'NOUVELLE': 'En attente de lecture',
        'PRISE_EN_COMPTE': 'Prise en compte',
        'EN_COURS': 'En attente de réalisation',
        'RESOLUE': 'Tâche terminée côté administrateur',
        'EN_ATTENTE_VALIDATION_CLOTURE': 'En attente de validation de clôture',
        'INTERVENTION_REFUSEE': 'Intervention refusée',
        'CLOTUREE': 'Validée côté client',
        'REJETEE': 'Rejetée',
    }

    def get(self, request, *args, **kwargs):
        user = request.user

        # Construire le queryset de base selon le rôle
        roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

        if 'ADMIN' in roles:
            queryset = Reclamation.objects.filter(actif=True)
        elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
            structure = user.client_profile.structure
            if structure:
                # Le client voit ses réclamations + celles de sa structure si visible_client=True
                queryset = Reclamation.objects.filter(
                    Q(createur=user) |
                    Q(structure_client=structure, visible_client=True),
                    actif=True
                )
            else:
                queryset = Reclamation.objects.filter(createur=user, actif=True)
        elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
            superviseur = user.superviseur_profile
            queryset = Reclamation.objects.filter(site__superviseur=superviseur, actif=True)
        else:
            queryset = Reclamation.objects.filter(createur=user, actif=True)

        # Appliquer les filtres
        statut = request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)

        urgence = request.query_params.get('urgence')
        if urgence:
            queryset = queryset.filter(urgence_id=urgence)

        type_rec = request.query_params.get('type_reclamation')
        if type_rec:
            queryset = queryset.filter(type_reclamation_id=type_rec)

        site = request.query_params.get('site')
        if site:
            queryset = queryset.filter(site_id=site)

        date_debut = request.query_params.get('date_debut')
        if date_debut:
            queryset = queryset.filter(date_creation__gte=date_debut)

        date_fin = request.query_params.get('date_fin')
        if date_fin:
            queryset = queryset.filter(date_creation__lte=date_fin)

        # Optimiser avec select_related et prefetch_related
        queryset = queryset.select_related(
            'type_reclamation',
            'urgence',
            'createur',
            'site',
            'zone',
            'equipe_affectee',
            'rejetee_par',
            'cloture_proposee_par',
            'cloture_refusee_par',
            'intervention_refusee_par'
        ).prefetch_related('historique__auteur').order_by('-date_creation')

        if not queryset.exists():
            return Response({'error': 'Aucune réclamation à exporter'}, status=status.HTTP_404_NOT_FOUND)

        # Créer le workbook
        wb = Workbook()

        # ========== ONGLET 1: LISTE DES RÉCLAMATIONS ==========
        ws_list = wb.active
        ws_list.title = "Réclamations"

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes pour la liste
        headers = [
            'N° Réclamation',
            'Type',
            'Urgence',
            'Statut',
            'Site',
            'Zone',
            'Créateur',
            'Équipe affectée',
            'Description',
            'Date constatation',
            'Date création',
            'Date prise en compte',
            'Date début traitement',
            'Date résolution',
            'Date clôture réelle',
            'Délai traitement (h)',
        ]
        ws_list.append(headers)

        # Appliquer le style d'en-tête
        for cell in ws_list[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Couleurs par statut (synchronisées avec le frontend constants.ts)
        # Versions claires des couleurs UI pour lisibilité Excel
        statut_colors = {
            'NOUVELLE': 'FEE2E2',        # Rouge clair (UI: #ef4444)
            'PRISE_EN_COMPTE': 'FFEDD5', # Orange clair (UI: #f97316)
            'EN_COURS': 'FFEDD5',        # Orange clair (UI: #f97316)
            'RESOLUE': 'DCFCE7',         # Vert clair (UI: #22c55e)
            'EN_ATTENTE_VALIDATION_CLOTURE': 'D1FAE5',  # Émeraude clair (UI: #10b981)
            'INTERVENTION_REFUSEE': 'FECACA',  # Rouge foncé clair (UI: #dc2626)
            'CLOTUREE': 'DCFCE7',        # Vert clair (UI: #22c55e)
            'REJETEE': 'E5E7EB',         # Gris clair (UI: #6b7280)
        }

        # Données
        for rec in queryset:
            # Calcul du délai de traitement en heures
            delai = None
            if rec.date_cloture_reelle and rec.date_creation:
                delta = rec.date_cloture_reelle - rec.date_creation
                delai = round(delta.total_seconds() / 3600, 1)

            row = [
                rec.numero_reclamation,
                rec.type_reclamation.nom_reclamation if rec.type_reclamation else '-',
                rec.urgence.niveau_urgence if rec.urgence else '-',
                self.STATUT_LABELS.get(rec.statut, rec.statut),
                rec.site.nom_site if rec.site else '-',
                rec.zone.nom if rec.zone else '-',
                f"{rec.createur.prenom} {rec.createur.nom}".strip() if rec.createur else '-',
                rec.equipe_affectee.nom_equipe if rec.equipe_affectee else '-',
                rec.description[:100] + '...' if len(rec.description) > 100 else rec.description,
                rec.date_constatation.strftime('%d/%m/%Y %H:%M') if rec.date_constatation else '-',
                rec.date_creation.strftime('%d/%m/%Y %H:%M') if rec.date_creation else '-',
                rec.date_prise_en_compte.strftime('%d/%m/%Y %H:%M') if rec.date_prise_en_compte else '-',
                rec.date_debut_traitement.strftime('%d/%m/%Y %H:%M') if rec.date_debut_traitement else '-',
                rec.date_resolution.strftime('%d/%m/%Y %H:%M') if rec.date_resolution else '-',
                rec.date_cloture_reelle.strftime('%d/%m/%Y %H:%M') if rec.date_cloture_reelle else '-',
                delai if delai else '-',
            ]
            ws_list.append(row)

            # Couleur selon statut
            if rec.statut in statut_colors:
                fill = PatternFill(start_color=statut_colors[rec.statut], end_color=statut_colors[rec.statut], fill_type='solid')
                for cell in ws_list[ws_list.max_row]:
                    cell.fill = fill

        # Ajuster les largeurs de colonnes
        column_widths = [18, 20, 12, 25, 20, 15, 20, 20, 40, 18, 18, 18, 18, 18, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws_list.column_dimensions[ws_list.cell(1, i).column_letter].width = width

        # Activer les filtres
        ws_list.auto_filter.ref = ws_list.dimensions

        # ========== ONGLET 2: HISTORIQUE DÉTAILLÉ ==========
        ws_hist = wb.create_sheet(title="Historique")

        # En-têtes historique
        hist_headers = [
            'N° Réclamation',
            'Type',
            'Date changement',
            'Ancien statut',
            'Nouveau statut',
            'Modifié par',
            'Commentaire'
        ]
        ws_hist.append(hist_headers)

        # Style en-tête
        for cell in ws_hist[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = header_border

        # Données historique
        for rec in queryset:
            for hist in rec.historique.all().order_by('date_changement'):
                hist_row = [
                    rec.numero_reclamation,
                    rec.type_reclamation.nom_reclamation if rec.type_reclamation else '-',
                    hist.date_changement.strftime('%d/%m/%Y %H:%M:%S') if hist.date_changement else '-',
                    self.STATUT_LABELS.get(hist.statut_precedent, hist.statut_precedent) if hist.statut_precedent else '(Création)',
                    self.STATUT_LABELS.get(hist.statut_nouveau, hist.statut_nouveau),
                    f"{hist.auteur.prenom} {hist.auteur.nom}".strip() if hist.auteur else '-',
                    hist.commentaire or '-'
                ]
                ws_hist.append(hist_row)

        # Largeurs colonnes historique
        hist_widths = [18, 20, 20, 25, 25, 20, 50]
        for i, width in enumerate(hist_widths, 1):
            ws_hist.column_dimensions[ws_hist.cell(1, i).column_letter].width = width

        ws_hist.auto_filter.ref = ws_hist.dimensions

        # ========== ONGLET 3: STATISTIQUES ==========
        ws_stats = wb.create_sheet(title="Statistiques")

        stats_title_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        stats_title_font = Font(bold=True, color='FFFFFF', size=12)

        # Titre
        ws_stats.append(['STATISTIQUES DES RÉCLAMATIONS'])
        ws_stats.merge_cells('A1:B1')
        ws_stats['A1'].font = stats_title_font
        ws_stats['A1'].fill = stats_title_fill
        ws_stats['A1'].alignment = Alignment(horizontal='center')

        ws_stats.append([])
        ws_stats.append(['Date d\'export:', datetime.now().strftime('%d/%m/%Y à %H:%M')])
        ws_stats.append(['Total réclamations:', queryset.count()])
        ws_stats.append([])

        # Stats par statut
        ws_stats.append(['RÉPARTITION PAR STATUT'])
        ws_stats[ws_stats.max_row][0].font = Font(bold=True)
        statut_counts = queryset.values('statut').annotate(count=Count('id'))
        for item in statut_counts:
            ws_stats.append([self.STATUT_LABELS.get(item['statut'], item['statut']), item['count']])

        ws_stats.append([])

        # Stats par urgence
        ws_stats.append(['RÉPARTITION PAR URGENCE'])
        ws_stats[ws_stats.max_row][0].font = Font(bold=True)
        urgence_counts = queryset.values('urgence__niveau_urgence').annotate(count=Count('id'))
        for item in urgence_counts:
            ws_stats.append([item['urgence__niveau_urgence'] or '-', item['count']])

        ws_stats.append([])

        # Stats par type
        ws_stats.append(['RÉPARTITION PAR TYPE'])
        ws_stats[ws_stats.max_row][0].font = Font(bold=True)
        type_counts = queryset.values('type_reclamation__nom_reclamation').annotate(count=Count('id'))
        for item in type_counts:
            ws_stats.append([item['type_reclamation__nom_reclamation'] or '-', item['count']])

        # Largeurs
        ws_stats.column_dimensions['A'].width = 35
        ws_stats.column_dimensions['B'].width = 15

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Réponse HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reclamations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
