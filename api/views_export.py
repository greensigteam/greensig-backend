# api/views_export.py
from rest_framework import status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q, Count
from django.contrib.gis.geos import GEOSGeometry
from celery.result import AsyncResult
import json

from api_users.permissions import IsAdmin, IsAdminOrSuperviseur, CanExportData

from .models import (
    Site, SousSite, Objet, Arbre, Gazon, Palmier, Arbuste, Vivace, Cactus, Graminee,
    Puit, Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte, Ballon
)
from .serializers import (
    SiteSerializer, SousSiteSerializer, ArbreSerializer, GazonSerializer,
    PalmierSerializer, ArbusteSerializer, VivaceSerializer, CactusSerializer,
    GramineeSerializer, PuitSerializer, PompeSerializer, VanneSerializer,
    ClapetSerializer, CanalisationSerializer, AspersionSerializer,
    GoutteSerializer, BallonSerializer
)
from .filters import (
    SiteFilter, SousSiteFilter, ArbreFilter, GazonFilter, PalmierFilter,
    ArbusteFilter, VivaceFilter, CactusFilter, GramineeFilter, PuitFilter,
    PompeFilter, VanneFilter, ClapetFilter, CanalisationFilter,
    AspersionFilter, GoutteFilter, BallonFilter
)


# ==============================================================================
# VUE POUR L'EXPORT PDF
# ==============================================================================

class ExportPDFView(APIView):
    """
    Vue pour exporter la carte en PDF.
    Accepte les paramètres: title, mapImageBase64, visibleLayers, center, zoom

    Mode async (recommandé pour les gros exports):
    - Ajouter ?async=true ou "async": true dans le body
    - Retourne un task_id pour suivre la progression
    - Utiliser GET /api/tasks/<task_id>/status/ pour vérifier le statut
    """
    def post(self, request, *args, **kwargs):
        # Récupérer les données du POST
        title = request.data.get('title', 'Export Carte GreenSIG')
        map_image_base64 = request.data.get('mapImageBase64', '')
        visible_layers = request.data.get('visibleLayers', {})
        center = request.data.get('center', [0, 0])
        zoom = request.data.get('zoom', 15)
        site_names = request.data.get('siteNames', [])

        # Mode asynchrone ?
        async_mode = request.data.get('async', request.query_params.get('async', 'false'))
        is_async = str(async_mode).lower() in ('true', '1', 'yes')

        if is_async:
            # Exécuter en arrière-plan via Celery
            from .tasks import export_pdf_async
            task = export_pdf_async.delay(
                user_id=request.user.id,
                title=title,
                map_image_base64=map_image_base64,
                visible_layers=visible_layers,
                center=center,
                zoom=zoom,
                site_names=site_names
            )
            return Response({
                'task_id': task.id,
                'status': 'PENDING',
                'message': 'Export PDF démarré en arrière-plan. Utilisez /api/tasks/{task_id}/status/ pour suivre la progression.'
            }, status=status.HTTP_202_ACCEPTED)

        # Mode synchrone (comportement original)
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib.utils import ImageReader
        from django.http import HttpResponse
        from datetime import datetime
        import base64
        import io

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        page_width, page_height = landscape(A4)
        pdf = canvas.Canvas(buffer, pagesize=landscape(A4))

        # Utilisateur
        user = request.user
        user_name = user.get_full_name() if hasattr(user, 'get_full_name') else f"{getattr(user, 'prenom', '')} {getattr(user, 'nom', user.username)}".strip()
        user_info = f"Exporté par: {user_name} ({user.email})" if user.email else f"Exporté par: {user_name}"
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(2*cm, page_height - 2*cm, user_info)

        # Site(s) - Afficher tous les sites
        if site_names:
            sites_text = f"Site(s): {', '.join(site_names)}"
            pdf.setFont("Helvetica", 10)

            # Si le texte est trop long, utiliser une police plus petite ou couper en lignes
            max_width = page_width - 10*cm  # Largeur disponible (en laissant de la marge pour la légende)
            text_width = pdf.stringWidth(sites_text, "Helvetica", 10)

            if text_width > max_width:
                # Réduire la taille de la police si nécessaire
                pdf.setFont("Helvetica", 8)
                text_width = pdf.stringWidth(sites_text, "Helvetica", 8)

                if text_width > max_width:
                    # Si toujours trop long, couper en plusieurs lignes
                    pdf.setFont("Helvetica", 8)
                    words = sites_text.split(', ')
                    lines = []
                    current_line = words[0]

                    for word in words[1:]:
                        test_line = current_line + ', ' + word
                        if pdf.stringWidth(test_line, "Helvetica", 8) <= max_width:
                            current_line = test_line
                        else:
                            lines.append(current_line)
                            current_line = word
                    lines.append(current_line)

                    y_pos = page_height - 2.6*cm
                    for line in lines:
                        pdf.drawString(2*cm, y_pos, line)
                        y_pos -= 0.4*cm
                    date_y = y_pos - 0.2*cm
                else:
                    pdf.drawString(2*cm, page_height - 2.6*cm, sites_text)
                    date_y = page_height - 3.2*cm
            else:
                pdf.drawString(2*cm, page_height - 2.6*cm, sites_text)
                date_y = page_height - 3.2*cm
        else:
            date_y = page_height - 2.6*cm

        # Date
        pdf.setFont("Helvetica", 10)
        date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        pdf.drawString(2*cm, date_y, f"Date d'export: {date_str}")

        # Logo GreenSIG (en haut à droite)
        try:
            import os
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo_width = 3*cm
                logo_height = 1.5*cm
                logo_x = page_width - logo_width - 2*cm
                logo_y = page_height - 2.5*cm
                pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            pass  # Si le logo n'est pas disponible, on continue sans

        # Image de la carte (si fournie)
        if map_image_base64:
            try:
                # Décoder l'image base64
                image_data = base64.b64decode(map_image_base64.split(',')[1] if ',' in map_image_base64 else map_image_base64)
                image = ImageReader(io.BytesIO(image_data))

                # Dessiner l'image (centré, 70% de la largeur)
                img_width = page_width * 0.7
                img_height = (page_height - 6*cm) * 0.7
                img_x = 2*cm
                img_y = page_height - 5*cm - img_height

                pdf.drawImage(image, img_x, img_y, width=img_width, height=img_height, preserveAspectRatio=True)
            except Exception as e:
                pdf.setFont("Helvetica", 10)
                pdf.drawString(2*cm, page_height - 5*cm, f"Erreur lors du chargement de l'image: {str(e)}")

        # Légende (à droite de l'image)
        legend_x = page_width - 6*cm
        legend_y = page_height - 3*cm

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(legend_x, legend_y, "Légende")
        legend_y -= 0.6*cm

        # Définition des catégories et éléments (Synchronisé avec constants.ts)
        legend_data = [
            {
                "title": "SITES",
                "items": [
                    {"name": "Sites", "color": (59, 130, 246)},      # #3b82f6
                ]
            },
            {
                "title": "VÉGÉTATION",
                "items": [
                    {"name": "Arbres", "color": (5, 150, 105)},      # #059669
                    {"name": "Gazons", "color": (132, 204, 22)},     # #84cc16
                    {"name": "Palmiers", "color": (249, 115, 22)},    # #f97316
                    {"name": "Arbustes", "color": (16, 185, 129)},    # #10b981
                    {"name": "Vivaces", "color": (236, 72, 153)},     # #ec4899
                    {"name": "Cactus", "color": (6, 182, 212)},      # #06b6d4
                    {"name": "Graminées", "color": (234, 179, 8)},     # #eab308
                ]
            },
            {
                "title": "HYDRAULIQUE",
                "items": [
                    {"name": "Puits", "color": (14, 165, 233)},      # #0ea5e9
                    {"name": "Pompes", "color": (6, 182, 212)},      # #06b6d4
                    {"name": "Vannes", "color": (20, 184, 166)},     # #14b8a6
                    {"name": "Clapets", "color": (8, 145, 178)},     # #0891b2
                    {"name": "Canalisations", "color": (2, 132, 199)}, # #0284c7
                    {"name": "Aspersions", "color": (56, 189, 248)}, # #38bdf8
                    {"name": "Gouttes", "color": (125, 211, 252)},   # #7dd3fc
                    {"name": "Ballons", "color": (3, 105, 161)},     # #0369a1
                ]
            },
            {
                "title": "RÉCLAMATIONS",
                "items": [
                    {"name": "En attente de lecture", "color": (239, 68, 68)},      # #ef4444
                    {"name": "Prise en compte", "color": (249, 115, 22)}, # #f97316
                    {"name": "En attente de réalisation", "color": (234, 179, 8)},       # #eab308
                    {"name": "Tâche terminée (côté admin.)", "color": (34, 197, 94)},        # #22c55e
                    {"name": "Clôturée (par le client)", "color": (16, 185, 129)},    # #10b981
                ]
            }
        ]

        # Rendre la légende complète
        for category in legend_data:
            # Titre de catégorie
            pdf.setFont("Helvetica-Bold", 9)
            pdf.setFillColorRGB(0.4, 0.4, 0.4)
            pdf.drawString(legend_x, legend_y, category["title"])
            legend_y -= 0.45*cm

            # Cas spécial pour les réclamations : afficher les symbologies
            if category["title"] == "RÉCLAMATIONS":
                for item in category["items"]:
                    color = item["color"]

                    # Symbole Point: Cercle avec "!"
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.circle(legend_x + 0.25*cm, legend_y + 0.1*cm, 0.12*cm, fill=1)
                    pdf.setFillColorRGB(1, 1, 1)  # Blanc pour le "!"
                    pdf.setFont("Helvetica-Bold", 7)
                    pdf.drawString(legend_x + 0.22*cm, legend_y + 0.05*cm, "!")

                    # Symbole Polygon: Rectangle avec bordure pointillée
                    pdf.setStrokeColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255, alpha=0.25)
                    pdf.setLineWidth(1.5)
                    pdf.setDash([2, 1])  # Ligne pointillée
                    pdf.rect(legend_x + 0.55*cm, legend_y + 0.02*cm, 0.2*cm, 0.16*cm, fill=1, stroke=1)
                    pdf.setDash([])  # Reset dash

                    # Texte
                    pdf.setFillColorRGB(0, 0, 0)
                    pdf.setFont("Helvetica", 8)
                    pdf.drawString(legend_x + 0.9*cm, legend_y, item["name"])
                    legend_y -= 0.4*cm
            else:
                # Autres catégories: symbologie normale
                for item in category["items"]:
                    # Puce de couleur
                    color = item["color"]
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.circle(legend_x + 0.2*cm, legend_y + 0.1*cm, 0.15*cm, fill=1)

                    # Texte
                    pdf.setFillColorRGB(0, 0, 0)
                    pdf.setFont("Helvetica", 8)
                    pdf.drawString(legend_x + 0.6*cm, legend_y, item["name"])
                    legend_y -= 0.4*cm

            legend_y -= 0.2*cm

        # Informations de la vue
        info_y = 2*cm
        pdf.setFont("Helvetica", 8)
        pdf.drawString(2*cm, info_y, f"Centre: [{center[0]:.6f}, {center[1]:.6f}] | Zoom: {zoom}")

        # Finaliser le PDF
        pdf.showPage()
        pdf.save()

        # Préparer la réponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"carte_greensig_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response




# ==============================================================================
# VUES POUR L'EXPORT DE DONNÉES
# ==============================================================================

class ExportDataView(APIView):
    """
    Vue générique pour exporter les données en Excel, GeoJSON, KML ou Shapefile.

    Permission: ADMIN, SUPERVISEUR et CLIENT (données filtrées par structure pour CLIENT)

    Paramètres de requête:
    - model: nom du modèle (arbres, gazons, palmiers, etc.)
    - format: xlsx, geojson, kml, shp (défaut: xlsx)
    - ids: optionnel, liste d'IDs séparés par virgules pour export sélectif
    - filtres: optionnels (même syntaxe que les endpoints de liste)

    Mode async (recommandé pour les gros exports):
    - Ajouter ?async=true à l'URL
    - Retourne un task_id pour suivre la progression
    - Utiliser GET /api/tasks/<task_id>/status/ pour vérifier le statut
    """
    permission_classes = [permissions.IsAuthenticated, CanExportData]

    MODEL_MAPPING = {
        'sites': Site,
        'sous-sites': SousSite,
        'arbres': Arbre,
        'gazons': Gazon,
        'palmiers': Palmier,
        'arbustes': Arbuste,
        'vivaces': Vivace,
        'cactus': Cactus,
        'graminees': Graminee,
        'puits': Puit,
        'pompes': Pompe,
        'vannes': Vanne,
        'clapets': Clapet,
        'canalisations': Canalisation,
        'aspersions': Aspersion,
        'gouttes': Goutte,
        'ballons': Ballon,
    }

    def get(self, request, model_name, *args, **kwargs):
        # Vérifier que le modèle existe
        if model_name not in self.MODEL_MAPPING:
            return Response({'error': f'Modèle invalide: {model_name}'}, status=400)

        # Mode asynchrone ?
        async_mode = request.query_params.get('async', 'false')
        is_async = str(async_mode).lower() in ('true', '1', 'yes')

        if is_async:
            # Récupérer les paramètres pour Celery
            export_format = request.query_params.get('format', 'xlsx').lower()
            ids_param = request.query_params.get('ids', '')
            ids = None
            if ids_param:
                try:
                    ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
                except ValueError:
                    return Response({'error': 'ids parameter must be comma-separated integers'}, status=400)

            filters = {}
            site_id = request.query_params.get('site')
            if site_id:
                filters['site'] = site_id

            # Exécuter en arrière-plan via Celery
            from .tasks import export_data_async
            task = export_data_async.delay(
                user_id=request.user.id,
                model_name=model_name,
                export_format=export_format,
                filters=filters if filters else None,
                ids=ids
            )
            return Response({
                'task_id': task.id,
                'status': 'PENDING',
                'message': f'Export {export_format.upper()} démarré en arrière-plan. Utilisez /api/tasks/{{task_id}}/status/ pour suivre la progression.'
            }, status=status.HTTP_202_ACCEPTED)

        # Mode synchrone (comportement original)
        from openpyxl import Workbook
        from django.http import HttpResponse
        from datetime import datetime

        model_class = self.MODEL_MAPPING[model_name]

        # Récupérer le format d'export
        export_format = request.query_params.get('format', 'xlsx').lower()
        valid_formats = ['xlsx', 'geojson', 'kml', 'shp', 'shapefile']
        if export_format not in valid_formats:
            return Response({'error': f'Format invalide. Utilisez: {", ".join(valid_formats)}'}, status=400)

        # Normaliser shapefile
        if export_format == 'shapefile':
            export_format = 'shp'

        # Appliquer les filtres
        queryset = model_class.objects.all()

        # Filtrer par structure pour les utilisateurs CLIENT
        user = request.user
        if user.roles_utilisateur.filter(role__nom_role='CLIENT').exists():
            if hasattr(user, 'client_profile') and user.client_profile.structure:
                structure = user.client_profile.structure
                # Les modèles Objet ont une relation site -> structure_client
                if hasattr(model_class, 'site'):
                    queryset = queryset.filter(site__structure_client=structure)
                # Sites ont directement structure_client
                elif hasattr(model_class, 'structure_client'):
                    queryset = queryset.filter(structure_client=structure)
                # SousSite via site
                elif model_class == SousSite:
                    queryset = queryset.filter(site__structure_client=structure)

        # Filtre par IDs si fourni
        ids_param = request.query_params.get('ids', '')
        if ids_param:
            try:
                ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
                queryset = queryset.filter(pk__in=ids)
            except ValueError:
                return Response({'error': 'ids parameter must be comma-separated integers'}, status=400)

        # Filtre par site si fourni
        site_id = request.query_params.get('site')
        if site_id:
            queryset = queryset.filter(site_id=site_id)

        # Vérifier qu'il y a des données
        if not queryset.exists():
            return Response({'error': 'Aucune donnée à exporter'}, status=404)

        # ==============================================================================
        # EXPORT GeoJSON
        # ==============================================================================
        if export_format == 'geojson':
            from .services.geo_io import export_to_geojson

            geojson_data = export_to_geojson(queryset)

            response = HttpResponse(
                json.dumps(geojson_data, ensure_ascii=False, indent=2),
                content_type='application/geo+json; charset=utf-8'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.geojson"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # ==============================================================================
        # EXPORT KML
        # ==============================================================================
        if export_format == 'kml':
            from .services.geo_io import export_to_kml

            kml_content = export_to_kml(queryset)

            response = HttpResponse(
                kml_content,
                content_type='application/vnd.google-earth.kml+xml; charset=utf-8'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.kml"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # ==============================================================================
        # EXPORT Shapefile (ZIP)
        # ==============================================================================
        if export_format == 'shp':
            from .services.geo_io import export_to_shapefile

            try:
                zip_content = export_to_shapefile(queryset, model_name)

                response = HttpResponse(
                    zip_content,
                    content_type='application/zip'
                )
                filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError as e:
                return Response({'error': str(e)}, status=500)
            except Exception as e:
                return Response({'error': f'Shapefile export error: {str(e)}'}, status=500)

        # ==============================================================================
        # EXPORT XLSX
        # ==============================================================================
        # Récupérer tous les objets (pas de pagination pour l'export)
        objects = list(queryset.values())

        # Obtenir les noms de colonnes
        field_names = list(objects[0].keys())

        # Export Excel
        if export_format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = model_name[:31]  # Excel limite à 31 caractères

            # Écrire l'en-tête
            ws.append(field_names)

            # Écrire les données
            for obj in objects:
                ws.append([obj[field] for field in field_names])

            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Sauvegarder dans un buffer
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response


# ==============================================================================
# VUE EXPORT EXCEL AMÉLIORÉ POUR INVENTAIRE
# ==============================================================================

class InventoryExportExcelView(APIView):
    """
    Vue spécialisée pour l'export Excel professionnel de l'inventaire.

    Permission: ADMIN, SUPERVISEUR et CLIENT (données filtrées par structure pour CLIENT)

    Supporte:
    - Formatage professionnel (styles, couleurs, filtres Excel)
    - Export multi-types avec onglets séparés
    - Filtres passés depuis le frontend
    - Statistiques par type

    Paramètres de requête:
    - types: liste de types séparés par virgules (ex: 'Arbre,Palmier,Gazon')
    - site: ID du site (optionnel)
    - etat: état des objets (bon, moyen, mauvais, critique)
    - famille: famille botanique (optionnel)
    - search: recherche textuelle (optionnel)
    """
    permission_classes = [permissions.IsAuthenticated, CanExportData]

    MODEL_MAPPING = {
        'Arbre': Arbre,
        'Palmier': Palmier,
        'Gazon': Gazon,
        'Arbuste': Arbuste,
        'Vivace': Vivace,
        'Cactus': Cactus,
        'Graminee': Graminee,
        'Puit': Puit,
        'Pompe': Pompe,
        'Vanne': Vanne,
        'Clapet': Clapet,
        'Canalisation': Canalisation,
        'Aspersion': Aspersion,
        'Goutte': Goutte,
        'Ballon': Ballon,
    }

    # Mapping des champs à exporter par type (colonnes personnalisées)
    # Note: 'superficie_calculee' est une annotation calculée dynamiquement
    # Note: 'derniere_intervention' est calculée depuis les tâches si last_intervention_date est null
    FIELD_MAPPINGS = {
        'Arbre': ['nom', 'famille', 'taille', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Palmier': ['nom', 'famille', 'taille', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Gazon': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Arbuste': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Vivace': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Cactus': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Graminee': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Puit': ['nom', 'profondeur', 'diametre', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Pompe': ['nom', 'type', 'diametre', 'puissance', 'debit', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Vanne': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Clapet': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Canalisation': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Aspersion': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Goutte': ['type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Ballon': ['marque', 'pression', 'volume', 'materiau', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
    }

    # Types qui ont une géométrie polygone (pour calcul de surface)
    POLYGON_TYPES = ['Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee']

    # Types qui ont le champ last_intervention_date
    TYPES_WITH_INTERVENTION_DATE = [
        'Arbre', 'Palmier', 'Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee',
        'Puit', 'Pompe'
    ]

    # Labels français pour les colonnes
    FIELD_LABELS = {
        'nom': 'Nom',
        'marque': 'Marque',
        'famille': 'Famille',
        'taille': 'Taille',
        'densite': 'Densité',
        'area_sqm': 'Surface (m²)',
        'superficie_calculee': 'Surface (m²)',
        'profondeur': 'Profondeur (m)',
        'diametre': 'Diamètre (cm)',
        'type': 'Type',
        'puissance': 'Puissance (kW)',
        'debit': 'Débit (m³/h)',
        'materiau': 'Matériau',
        'pression': 'Pression (bar)',
        'volume': 'Volume (L)',
        'site__nom_site': 'Site',
        'sous_site__nom': 'Sous-site',
        'etat': 'État',
        'last_intervention_date': 'Dernière intervention',
        'derniere_intervention': 'Dernière intervention',
    }

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as ExcelImage
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        import os
        from django.conf import settings

        # Récupérer les types à exporter
        types_param = request.query_params.get('types', '')
        if types_param:
            types_list = [t.strip() for t in types_param.split(',') if t.strip()]
        else:
            # Si aucun type spécifié, exporter tous
            types_list = list(self.MODEL_MAPPING.keys())

        # Valider les types
        invalid_types = [t for t in types_list if t not in self.MODEL_MAPPING]
        if invalid_types:
            return Response({'error': f'Types invalides: {", ".join(invalid_types)}'}, status=400)

        # Créer le workbook
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        # Styles réutilisables
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        stats_font = Font(bold=True, size=10)
        stats_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

        # États et leurs couleurs
        etat_colors = {
            'bon': 'C8E6C9',        # Vert clair
            'moyen': 'FFF9C4',      # Jaune clair
            'mauvais': 'FFCCBC',    # Orange clair
            'critique': 'FFCDD2',   # Rouge clair
        }

        # Filtrer par rôle (ADMIN, CLIENT, SUPERVISEUR)
        user = request.user
        structure_filter = None
        superviseur_filter = None
        if user.is_authenticated:
            roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

            if 'ADMIN' in roles:
                pass  # ADMIN voit tout
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                superviseur_filter = user.superviseur_profile

        # Pré-cacher les IDs de sites pour éviter de répéter la jointure dans chaque type
        role_site_ids = None
        if structure_filter:
            role_site_ids = set(
                Site.objects.filter(structure_client=structure_filter).values_list('id', flat=True)
            )
        elif superviseur_filter:
            role_site_ids = set(
                Site.objects.filter(superviseur=superviseur_filter).values_list('id', flat=True)
            )

        # Pré-extraire les filtres de la requête (une seule fois)
        filter_site_id = request.query_params.get('site')
        filter_etat = request.query_params.get('etat')
        filter_famille = request.query_params.get('famille')
        filter_search = request.query_params.get('search')

        # Pour chaque type, créer un onglet
        for type_name in types_list:
            model_class = self.MODEL_MAPPING[type_name]

            # Construire le queryset avec filtres
            queryset = model_class.objects.select_related('site', 'sous_site').all()

            # Filtrer par rôle (via IDs pré-cachés)
            if role_site_ids is not None:
                queryset = queryset.filter(site_id__in=role_site_ids)

            # Appliquer les filtres
            if filter_site_id:
                queryset = queryset.filter(site_id=filter_site_id)

            if filter_etat:
                queryset = queryset.filter(etat=filter_etat)

            if filter_famille and hasattr(model_class, 'famille'):
                queryset = queryset.filter(famille__icontains=filter_famille)

            if filter_search:
                if hasattr(model_class, 'nom'):
                    queryset = queryset.filter(nom__icontains=filter_search)
                elif hasattr(model_class, 'marque'):
                    queryset = queryset.filter(marque__icontains=filter_search)

            # Évaluer le queryset une seule fois (évite .exists() + .count() + itération séparés)
            objects_list = list(queryset)
            if not objects_list:
                continue

            # Créer l'onglet
            ws = wb.create_sheet(title=type_name[:31])

            # Récupérer les champs à exporter
            fields = self.FIELD_MAPPINGS.get(type_name, ['nom', 'site__nom_site', 'etat'])

            # Pré-charger les dernières interventions en une seule requête (évite N+1)
            intervention_map = {}
            if 'derniere_intervention' in fields:
                from api_planification.models import Tache
                objet_ids = [obj.objet_ptr_id for obj in objects_list]
                from django.db.models import Max
                interventions = (
                    Tache.objects
                    .filter(objets__id__in=objet_ids, statut='TERMINEE')
                    .values('objets__id')
                    .annotate(derniere=Max('date_fin_reelle'))
                )
                intervention_map = {row['objets__id']: row['derniere'] for row in interventions}

            # Écrire l'en-tête
            headers = [self.FIELD_LABELS.get(field, field) for field in fields]
            ws.append(headers)

            # Appliquer le style d'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            # Écrire les données - itérer directement sur les objets
            for obj in objects_list:
                row_values = []
                etat_value = None

                for field in fields:
                    value = None

                    # Gérer les champs calculés spéciaux
                    if field == 'superficie_calculee':
                        # Calculer la surface depuis la géométrie (pour polygones)
                        if type_name in self.POLYGON_TYPES and obj.geometry:
                            try:
                                # Utiliser la méthode area de GEOS (retourne m² pour géométries géographiques)
                                # Transformer en projection métrique pour calcul précis
                                from django.contrib.gis.geos import GEOSGeometry
                                geom = obj.geometry
                                # Transformer en Web Mercator (EPSG:3857) pour calcul en mètres
                                geom_projected = geom.transform(3857, clone=True)
                                value = geom_projected.area  # Surface en m²
                            except Exception:
                                value = None

                    elif field == 'derniere_intervention':
                        # Priorité au champ existant, sinon lookup pré-chargé
                        if type_name in self.TYPES_WITH_INTERVENTION_DATE and hasattr(obj, 'last_intervention_date'):
                            value = obj.last_intervention_date
                        if not value:
                            value = intervention_map.get(obj.objet_ptr_id)

                    elif field == 'site__nom_site':
                        # Champ lié: site.nom_site
                        value = obj.site.nom_site if obj.site else None

                    elif field == 'sous_site__nom':
                        # Champ lié: sous_site.nom
                        value = obj.sous_site.nom if obj.sous_site else None

                    else:
                        # Champ standard - accès direct à l'attribut
                        value = getattr(obj, field, None)

                    # Formater les dates
                    if field in ['last_intervention_date', 'derniere_intervention'] and value:
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%d/%m/%Y')

                    # Formater les nombres
                    elif field in ['area_sqm', 'superficie_calculee', 'profondeur', 'diametre', 'puissance', 'debit', 'pression', 'volume', 'densite']:
                        if value is not None:
                            value = round(float(value), 2)

                    # Capturer l'état pour la couleur
                    if field == 'etat':
                        etat_value = value

                    # Laisser vide si pas de valeur (au lieu de '-')
                    row_values.append(value if value is not None else '')

                ws.append(row_values)

                # Appliquer la couleur de fond selon l'état
                if etat_value and etat_value.lower() in etat_colors:
                    fill_color = etat_colors[etat_value.lower()]
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            # Ajouter une ligne de statistiques
            ws.append([])  # Ligne vide
            stats_row = ws.max_row + 1
            ws.append(['STATISTIQUES', '', '', '', '', '', ''])

            # Compter les états en mémoire (évite des requêtes supplémentaires)
            etat_summary = {}
            for obj in objects_list:
                etat_val = getattr(obj, 'etat', None) or ''
                etat_summary[etat_val] = etat_summary.get(etat_val, 0) + 1

            ws.append(['Total:', len(objects_list)])
            for etat_key, count in etat_summary.items():
                ws.append([f'{etat_key.capitalize()}:', count])

            # Appliquer le style aux statistiques
            for row in ws.iter_rows(min_row=stats_row, max_row=ws.max_row):
                for cell in row:
                    cell.font = stats_font
                    cell.fill = stats_fill

            # Ajuster les largeurs de colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Activer les filtres automatiques
            ws.auto_filter.ref = ws.dimensions

            # Figer la première ligne
            ws.freeze_panes = 'A2'

            # Ajouter le logo GreenSIG en haut à droite
            try:
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
                if os.path.exists(logo_path):
                    img = ExcelImage(logo_path)
                    # Redimensionner le logo (largeur en pixels)
                    img.width = 120
                    img.height = 60
                    # Positionner en haut à droite (colonne la plus à droite possible)
                    # Calculer la position basée sur le nombre de colonnes
                    last_col = chr(ord('A') + len(headers) - 1)
                    ws.add_image(img, f'{last_col}1')
            except Exception as e:
                pass  # Continuer sans logo si erreur

        # Vérifier qu'au moins un onglet a été créé
        if len(wb.sheetnames) == 0:
            return Response({'error': 'Aucune donnée à exporter avec les filtres appliqués'}, status=404)

        # Sauvegarder le workbook
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Créer la réponse HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


# ==============================================================================
# VUE EXPORT PDF POUR INVENTAIRE
# ==============================================================================

class InventoryExportPDFView(APIView):
    """
    Vue pour l'export PDF professionnel de l'inventaire.

    Permission: ADMIN, SUPERVISEUR et CLIENT (données filtrées par structure pour CLIENT)

    Génère un document PDF avec:
    - En-tête avec logo et titre
    - Tableau formaté des données
    - Statistiques par type
    - Pied de page avec numérotation

    Paramètres de requête identiques à InventoryExportExcelView:
    - types: liste de types séparés par virgules
    - site: ID du site (optionnel)
    - etat: état des objets
    - famille: famille botanique (optionnel)
    - search: recherche textuelle (optionnel)
    """
    permission_classes = [permissions.IsAuthenticated, CanExportData]

    MODEL_MAPPING = {
        'Arbre': Arbre,
        'Palmier': Palmier,
        'Gazon': Gazon,
        'Arbuste': Arbuste,
        'Vivace': Vivace,
        'Cactus': Cactus,
        'Graminee': Graminee,
        'Puit': Puit,
        'Pompe': Pompe,
        'Vanne': Vanne,
        'Clapet': Clapet,
        'Canalisation': Canalisation,
        'Aspersion': Aspersion,
        'Goutte': Goutte,
        'Ballon': Ballon,
    }

    FIELD_MAPPINGS = {
        'Arbre': ['nom', 'famille', 'taille', 'site__nom_site', 'etat'],
        'Palmier': ['nom', 'famille', 'taille', 'site__nom_site', 'etat'],
        'Gazon': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Arbuste': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Vivace': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Cactus': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Graminee': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Puit': ['nom', 'profondeur', 'diametre', 'site__nom_site', 'etat'],
        'Pompe': ['nom', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Vanne': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Clapet': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Canalisation': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Aspersion': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Goutte': ['type', 'diametre', 'site__nom_site', 'etat'],
        'Ballon': ['marque', 'volume', 'site__nom_site', 'etat'],
    }

    # Types polygones nécessitant le calcul de superficie
    POLYGON_TYPES = ['Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee']

    FIELD_LABELS = {
        'nom': 'Nom',
        'marque': 'Marque',
        'famille': 'Famille',
        'taille': 'Taille',
        'densite': 'Densité',
        'area_sqm': 'Surface (m²)',
        'superficie_calculee': 'Surface (m²)',
        'profondeur': 'Prof. (m)',
        'diametre': 'Diam. (cm)',
        'type': 'Type',
        'volume': 'Vol. (L)',
        'site__nom_site': 'Site',
        'etat': 'État',
    }

    def get(self, request, *args, **kwargs):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        import os
        from django.conf import settings

        # Récupérer les types à exporter
        types_param = request.query_params.get('types', '')
        if types_param:
            types_list = [t.strip() for t in types_param.split(',') if t.strip()]
        else:
            types_list = list(self.MODEL_MAPPING.keys())

        # Valider les types
        invalid_types = [t for t in types_list if t not in self.MODEL_MAPPING]
        if invalid_types:
            return Response({'error': f'Types invalides: {", ".join(invalid_types)}'}, status=400)

        # Créer le buffer
        buffer = BytesIO()

        # Créer le document PDF (paysage pour plus de colonnes)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Styles
        styles = getSampleStyleSheet()

        # Style titre principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Style sous-titre
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Style pour les sections
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=10,
            spaceBefore=15,
            fontName='Helvetica-Bold'
        )

        # Contenu du document
        story = []

        # Logo GreenSIG en haut à droite avec titre
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=4*cm, height=2*cm)

                # Tableau avec titre à gauche et logo à droite
                header_data = [
                    [Paragraph("INVENTAIRE - GreenSIG", title_style), logo]
                ]
                header_table = Table(header_data, colWidths=[20*cm, 5*cm])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(header_table)
            else:
                # Pas de logo, titre seul
                story.append(Paragraph("INVENTAIRE - GreenSIG", title_style))
        except Exception:
            # En cas d'erreur, titre seul
            story.append(Paragraph("INVENTAIRE - GreenSIG", title_style))

        story.append(Paragraph(f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
        story.append(Spacer(1, 0.5*cm))

        # Couleurs par état
        etat_colors = {
            'bon': colors.HexColor('#C8E6C9'),
            'moyen': colors.HexColor('#FFF9C4'),
            'mauvais': colors.HexColor('#FFCCBC'),
            'critique': colors.HexColor('#FFCDD2'),
        }

        # Filtrer par rôle (ADMIN, CLIENT, SUPERVISEUR)
        user = request.user
        structure_filter = None
        superviseur_filter = None
        if user.is_authenticated:
            roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

            if 'ADMIN' in roles:
                pass  # ADMIN voit tout
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                superviseur_filter = user.superviseur_profile

        # Pré-cacher les IDs de sites pour éviter de répéter la jointure dans chaque type
        role_site_ids = None
        if structure_filter:
            role_site_ids = set(
                Site.objects.filter(structure_client=structure_filter).values_list('id', flat=True)
            )
        elif superviseur_filter:
            role_site_ids = set(
                Site.objects.filter(superviseur=superviseur_filter).values_list('id', flat=True)
            )

        # Pré-extraire les filtres de la requête (une seule fois)
        filter_site_id = request.query_params.get('site')
        filter_etat = request.query_params.get('etat')
        filter_famille = request.query_params.get('famille')
        filter_search = request.query_params.get('search')

        # Pour chaque type
        total_objects = 0
        for idx, type_name in enumerate(types_list):
            model_class = self.MODEL_MAPPING[type_name]

            # Construire le queryset avec filtres
            queryset = model_class.objects.select_related('site', 'sous_site').all()

            # Filtrer par rôle (via IDs pré-cachés)
            if role_site_ids is not None:
                queryset = queryset.filter(site_id__in=role_site_ids)

            # Appliquer les filtres
            if filter_site_id:
                queryset = queryset.filter(site_id=filter_site_id)

            if filter_etat:
                queryset = queryset.filter(etat=filter_etat)

            if filter_famille and hasattr(model_class, 'famille'):
                queryset = queryset.filter(famille__icontains=filter_famille)

            if filter_search:
                if hasattr(model_class, 'nom'):
                    queryset = queryset.filter(nom__icontains=filter_search)
                elif hasattr(model_class, 'marque'):
                    queryset = queryset.filter(marque__icontains=filter_search)

            # Compter une seule fois (évite .exists() + .count() séparés)
            count = queryset.count()
            if count == 0:
                continue

            total_objects += count

            # Titre de section
            story.append(Paragraph(f"{type_name} ({count} élément{'s' if count > 1 else ''})", section_style))

            # Récupérer les champs
            fields = self.FIELD_MAPPINGS.get(type_name, ['nom', 'site__nom_site', 'etat'])
            headers = [self.FIELD_LABELS.get(field, field) for field in fields]

            # Pour les types polygones, annoter avec le calcul de superficie
            if type_name in self.POLYGON_TYPES and 'superficie_calculee' in fields:
                from django.db.models.functions import Coalesce
                from django.db.models import Value, FloatField
                from django.contrib.gis.db.models.functions import Area, Transform

                # Calculer la surface en m² via projection UTM
                queryset = queryset.annotate(
                    superficie_calculee=Coalesce(
                        Area(Transform('geometry', 32629)),  # UTM zone 29N pour le Maroc
                        Value(0.0),
                        output_field=FloatField()
                    )
                )

            # Données - Afficher TOUS les éléments
            data_rows = list(queryset.values(*fields))

            # Style pour les cellules avec retour à la ligne
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8,
                leading=10,  # Espacement entre les lignes
                wordWrap='CJK',  # Permet le retour à la ligne sur n'importe quel caractère
            )

            # Style pour l'en-tête (texte blanc sur fond vert)
            header_cell_style = ParagraphStyle(
                'HeaderCellStyle',
                parent=styles['Normal'],
                fontSize=9,
                leading=11,
                textColor=colors.whitesmoke,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER,
            )

            # Construire le tableau avec des Paragraph pour le retour à la ligne
            table_data = [[Paragraph(str(h), header_cell_style) for h in headers]]

            for row_data in data_rows:
                row = []
                for field in fields:
                    value = row_data.get(field)

                    # Formater les valeurs
                    if value is None:
                        value = '-'
                    elif field in ['area_sqm', 'superficie_calculee', 'profondeur', 'diametre', 'volume']:
                        value = f"{round(float(value), 1)}" if value else '-'
                    else:
                        value = str(value)

                    # Utiliser Paragraph pour permettre le retour à la ligne automatique
                    row.append(Paragraph(value, cell_style))

                table_data.append(row)

            # Calculer les statistiques en mémoire depuis data_rows (évite une requête supplémentaire)
            etat_summary = {}
            for row_data in data_rows:
                etat_val = row_data.get('etat', '') or ''
                etat_summary[etat_val] = etat_summary.get(etat_val, 0) + 1
            stats_text = "Répartition: " + " | ".join([f"{k.capitalize()}: {v}" for k, v in etat_summary.items()])

            # Ajouter une ligne de statistiques à la fin du tableau
            table_data.append([Paragraph(stats_text, cell_style)] + [Paragraph('', cell_style) for _ in range(len(fields) - 1)])
            stats_row_index = len(table_data) - 1

            # Créer le tableau avec des largeurs de colonnes adaptées
            # Première colonne (nom/marque) plus large pour les textes longs
            num_cols = len(fields)
            if num_cols >= 5:
                # Format typique: Nom(5cm), Famille(3cm), Taille/Surface(2.5cm), Site(4cm), État(2.5cm)
                col_widths = [5*cm] + [3*cm] * (num_cols - 3) + [4*cm, 2.5*cm]
            else:
                col_widths = [4*cm] * num_cols

            table = Table(table_data, colWidths=col_widths, repeatRows=1)  # repeatRows=1 pour répéter l'en-tête sur chaque page

            # Style du tableau
            table_style = TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

                # Corps
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2E7D32')),

                # Alternance de couleurs (exclure la dernière ligne de stats)
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),

                # Style de la ligne de statistiques (dernière ligne)
                ('SPAN', (0, stats_row_index), (-1, stats_row_index)),
                ('BACKGROUND', (0, stats_row_index), (-1, stats_row_index), colors.HexColor('#E8F5E9')),
                ('ALIGN', (0, stats_row_index), (-1, stats_row_index), 'CENTER'),
                ('FONTNAME', (0, stats_row_index), (-1, stats_row_index), 'Helvetica-Oblique'),
                ('TEXTCOLOR', (0, stats_row_index), (-1, stats_row_index), colors.HexColor('#2E7D32')),
            ])

            # Appliquer les couleurs par état si la colonne existe
            if 'etat' in fields:
                etat_col = fields.index('etat')
                for i, row_data in enumerate(data_rows, start=1):
                    etat_val = row_data.get('etat', '').lower()
                    if etat_val in etat_colors:
                        table_style.add('BACKGROUND', (etat_col, i), (etat_col, i), etat_colors[etat_val])

            table.setStyle(table_style)
            story.append(table)

            # Saut de page entre les types (sauf pour le dernier)
            if idx < len(types_list) - 1:
                story.append(PageBreak())
            else:
                story.append(Spacer(1, 0.5*cm))

        # Vérifier qu'il y a des données
        if total_objects == 0:
            return Response({'error': 'Aucune donnée à exporter avec les filtres appliqués'}, status=404)

        # Résumé final
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph("RÉSUMÉ GLOBAL", section_style))
        summary_table_data = [
            ['Total d\'objets exportés:', str(total_objects)],
            ['Nombre de types:', str(len([t for t in types_list if self.MODEL_MAPPING[t].objects.filter(**self._build_filters(request)).exists()]))],
            ['Date d\'export:', datetime.now().strftime('%d/%m/%Y à %H:%M')],
        ]
        summary_table = Table(summary_table_data, colWidths=[6*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)

        # Générer le PDF
        doc.build(story)

        # Retourner la réponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _build_filters(self, request):
        """Construit le dictionnaire de filtres à partir des paramètres de requête"""
        filters = {}

        site_id = request.query_params.get('site')
        if site_id:
            filters['site_id'] = site_id

        etat = request.query_params.get('etat')
        if etat:
            filters['etat'] = etat

        return filters
