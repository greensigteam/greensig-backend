# api/views.py
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q, Count
from django.contrib.gis.geos import GEOSGeometry
import json

from .models import (
    Site, SousSite, Objet, Arbre, Gazon, Palmier, Arbuste, Vivace, Cactus, Graminee,
    Puit, Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte, Ballon
)
from .serializers import (
    SiteSerializer, SousSiteSerializer, ArbreSerializer, GazonSerializer,
    PalmierSerializer, ArbusteSerializer, VivaceSerializer, CactusSerializer,
    GramineeSerializer, PuitSerializer, PompeSerializer, VanneSerializer,
    ClapetSerializer, CanalisationSerializer, AspersionSerializer,
    GoutteSerializer, BallonSerializer
)
from .filters import (
    SiteFilter, SousSiteFilter, ArbreFilter, GazonFilter, PalmierFilter,
    ArbusteFilter, VivaceFilter, CactusFilter, GramineeFilter, PuitFilter,
    PompeFilter, VanneFilter, ClapetFilter, CanalisationFilter,
    AspersionFilter, GoutteFilter, BallonFilter
)
from .site_statistics_view import SiteStatisticsView


# ==============================================================================
# MIXIN POUR LE FILTRAGE DES OBJETS GIS PAR PERMISSIONS
# ==============================================================================

class GISObjectPermissionMixin:
    """
    Mixin pour filtrer automatiquement les objets GIS selon les permissions utilisateur.

    Tous les objets GIS ont un champ 'site' qui permet de filtrer selon:
    - ADMIN: voit tous les objets
    - CLIENT: voit uniquement les objets de ses sites
    - SUPERVISEUR: voit uniquement les objets des sites qui lui sont affectés
    """
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if not user or not user.is_authenticated:
            return queryset.none()

        roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

        # ADMIN voit tout
        if 'ADMIN' in roles:
            return queryset

        # CLIENT voit uniquement les objets des sites de sa structure
        if 'CLIENT' in roles and hasattr(user, 'client_profile'):
            structure = user.client_profile.structure
            if structure:
                return queryset.filter(site__structure_client=structure)
            return queryset.none()

        # SUPERVISEUR voit uniquement les objets des sites qui lui sont affectés
        if 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
            return queryset.filter(site__superviseur=user.superviseur_profile)

        # Par défaut, aucun accès
        return queryset.none()


# ==============================================================================
# VUES POUR LA HIÉRARCHIE SPATIALE
# ==============================================================================

class SiteListCreateView(generics.ListCreateAPIView):
    serializer_class = SiteSerializer
    filterset_class = SiteFilter

    def get_queryset(self):
        """
        Filtrage automatique basé sur les permissions utilisateur:
        - ADMIN: voit tous les sites
        - CLIENT: voit uniquement ses sites
        - SUPERVISEUR: voit uniquement les sites liés aux tâches de ses équipes
        """
        queryset = Site.objects.all().order_by('id')
        user = self.request.user

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            # ADMIN voit tout
            if 'ADMIN' in roles:
                return queryset

            # CLIENT voit uniquement les sites de sa structure
            if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure = user.client_profile.structure
                if structure:
                    return queryset.filter(structure_client=structure)
                return queryset.none()

            # SUPERVISEUR voit uniquement les sites qui lui sont affectés directement
            if 'SUPERVISEUR' in roles:
                if hasattr(user, 'superviseur_profile'):
                    return queryset.filter(superviseur=user.superviseur_profile)
                else:
                    # Superviseur sans profil = aucun site visible
                    return queryset.none()

        # Par défaut, aucun accès pour les utilisateurs sans rôle reconnu
        return queryset.none()

    def _get_superviseur_site_ids(self, user):
        """
        Récupère les IDs des sites contenant les objets liés aux tâches du superviseur.
        Seuls les sites avec des objets dans les tâches sont retournés.

        OPTIMISÉ: Une seule requête SQL au lieu de N+1.
        """
        from api_planification.models import Tache

        try:
            superviseur = user.superviseur_profile
            # Équipes gérées par ce superviseur
            equipes_gerees = superviseur.equipes_gerees.filter(actif=True)
            equipes_gerees_ids = list(equipes_gerees.values_list('id', flat=True))

            if not equipes_gerees_ids:
                return []

            # Tâches assignées à ces équipes (non supprimées)
            taches_ids = Tache.objects.filter(
                deleted_at__isnull=True
            ).filter(
                Q(equipes__id__in=equipes_gerees_ids) | Q(id_equipe__in=equipes_gerees_ids)
            ).values_list('id', flat=True).distinct()

            # OPTIMISÉ: Récupérer les site_ids en une seule requête via la relation inverse
            # Au lieu de boucler sur chaque tâche puis chaque objet (N+1)
            site_ids = list(Objet.objects.filter(
                taches__id__in=taches_ids,
                site_id__isnull=False
            ).values_list('site_id', flat=True).distinct())

            return site_ids
        except Exception:
            return []


class SiteDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SiteSerializer

    def get_queryset(self):
        """
        Filtrage automatique basé sur les permissions utilisateur:
        - ADMIN: voit tous les sites
        - CLIENT: voit uniquement ses sites
        - SUPERVISEUR: voit uniquement les sites qui lui sont affectés
        """
        queryset = Site.objects.all()
        user = self.request.user

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            # ADMIN voit tout
            if 'ADMIN' in roles:
                return queryset

            # CLIENT voit uniquement les sites de sa structure
            if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure = user.client_profile.structure
                if structure:
                    return queryset.filter(structure_client=structure)
                return queryset.none()

            # SUPERVISEUR voit uniquement les sites qui lui sont affectés
            if 'SUPERVISEUR' in roles:
                if hasattr(user, 'superviseur_profile'):
                    return queryset.filter(superviseur=user.superviseur_profile)
                else:
                    return queryset.none()

        return queryset.none()


class SousSiteListCreateView(generics.ListCreateAPIView):
    queryset = SousSite.objects.all().order_by('id')
    serializer_class = SousSiteSerializer
    filterset_class = SousSiteFilter

    def get_queryset(self):
        """
        Filtrage automatique basé sur les permissions utilisateur:
        - ADMIN: voit tous les sous-sites
        - CLIENT: voit uniquement les sous-sites de ses sites
        - SUPERVISEUR: voit uniquement les sous-sites des sites qui lui sont affectés
        """
        queryset = super().get_queryset()
        user = self.request.user

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            # ADMIN voit tout
            if 'ADMIN' in roles:
                return queryset

            # CLIENT voit uniquement les sous-sites des sites de sa structure
            if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure = user.client_profile.structure
                if structure:
                    return queryset.filter(site__structure_client=structure)
                return queryset.none()

            # SUPERVISEUR voit uniquement les sous-sites des sites qui lui sont affectés
            if 'SUPERVISEUR' in roles:
                if hasattr(user, 'superviseur_profile'):
                    return queryset.filter(site__superviseur=user.superviseur_profile)
                else:
                    return queryset.none()

        return queryset.none()


class SousSiteDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SousSite.objects.all()
    serializer_class = SousSiteSerializer

    def get_queryset(self):
        """
        Filtrage automatique basé sur les permissions utilisateur:
        - ADMIN: voit tous les sous-sites
        - CLIENT: voit uniquement les sous-sites de ses sites
        - SUPERVISEUR: voit uniquement les sous-sites des sites qui lui sont affectés
        """
        queryset = super().get_queryset()
        user = self.request.user

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            # ADMIN voit tout
            if 'ADMIN' in roles:
                return queryset

            # CLIENT voit uniquement les sous-sites des sites de sa structure
            if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure = user.client_profile.structure
                if structure:
                    return queryset.filter(site__structure_client=structure)
                return queryset.none()

            # SUPERVISEUR voit uniquement les sous-sites des sites qui lui sont affectés
            if 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                return queryset.filter(site__superviseur=user.superviseur_profile)

        return queryset


class DetectSiteView(APIView):
    """
    Détecte le site contenant une géométrie donnée.

    POST /api/sites/detect/
    Body: { "geometry": { "type": "Point|Polygon|LineString", "coordinates": [...] } }

    Returns:
        - 200: { "site": { id, nom_site, code_site }, "sous_site": { id, nom } | null }
        - 404: { "error": "Aucun site ne contient cette géométrie" }
        - 400: { "error": "Géométrie invalide" }
    """

    def post(self, request):
        geometry_data = request.data.get('geometry')

        if not geometry_data:
            return Response(
                {'error': 'Le champ "geometry" est requis'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Convert GeoJSON to GEOS geometry
            geom = GEOSGeometry(json.dumps(geometry_data), srid=4326)
        except Exception as e:
            return Response(
                {'error': f'Géométrie invalide: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # For points, use intersects; for polygons/lines, check if contained or intersects
        # We use intersects to be more flexible (an object can touch the boundary)
        site = Site.objects.filter(geometrie_emprise__intersects=geom).first()

        if not site:
            return Response(
                {'error': 'Aucun site ne contient cette géométrie. Veuillez dessiner à l\'intérieur d\'un site existant.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Also try to detect sous-site (for points especially)
        sous_site = None
        if geom.geom_type == 'Point':
            # For points, find the nearest sous-site within a small tolerance
            sous_site_obj = SousSite.objects.filter(
                site=site,
                geometrie__distance_lte=(geom, 0.001)  # ~100m tolerance
            ).first()
            if sous_site_obj:
                sous_site = {
                    'id': sous_site_obj.id,
                    'nom': sous_site_obj.nom
                }

        return Response({
            'site': {
                'id': site.id,
                'nom_site': site.nom_site,
                'code_site': site.code_site
            },
            'sous_site': sous_site
        })


# ==============================================================================
# VUES POUR LES VÉGÉTAUX
# ==============================================================================

class ArbreListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Arbre.objects.all().order_by('id')
    serializer_class = ArbreSerializer
    filterset_class = ArbreFilter


class ArbreDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Arbre.objects.all()
    serializer_class = ArbreSerializer


class GazonListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Gazon.objects.all().order_by('id')
    serializer_class = GazonSerializer
    filterset_class = GazonFilter


class GazonDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Gazon.objects.all()
    serializer_class = GazonSerializer


class PalmierListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Palmier.objects.all().order_by('id')
    serializer_class = PalmierSerializer
    filterset_class = PalmierFilter


class PalmierDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Palmier.objects.all()
    serializer_class = PalmierSerializer


class ArbusteListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Arbuste.objects.all().order_by('id')
    serializer_class = ArbusteSerializer
    filterset_class = ArbusteFilter


class ArbusteDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Arbuste.objects.all()
    serializer_class = ArbusteSerializer


class VivaceListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Vivace.objects.all().order_by('id')
    serializer_class = VivaceSerializer
    filterset_class = VivaceFilter


class VivaceDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Vivace.objects.all()
    serializer_class = VivaceSerializer


class CactusListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Cactus.objects.all().order_by('id')
    serializer_class = CactusSerializer
    filterset_class = CactusFilter


class CactusDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Cactus.objects.all()
    serializer_class = CactusSerializer


class GramineeListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Graminee.objects.all().order_by('id')
    serializer_class = GramineeSerializer
    filterset_class = GramineeFilter


class GramineeDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Graminee.objects.all()
    serializer_class = GramineeSerializer


# ==============================================================================
# VUES POUR L'HYDRAULIQUE
# ==============================================================================

class PuitListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Puit.objects.all().order_by('id')
    serializer_class = PuitSerializer
    filterset_class = PuitFilter


class PuitDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Puit.objects.all()
    serializer_class = PuitSerializer


class PompeListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Pompe.objects.all().order_by('id')
    serializer_class = PompeSerializer
    filterset_class = PompeFilter


class PompeDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Pompe.objects.all()
    serializer_class = PompeSerializer


class VanneListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Vanne.objects.all().order_by('id')
    serializer_class = VanneSerializer
    filterset_class = VanneFilter


class VanneDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Vanne.objects.all()
    serializer_class = VanneSerializer


class ClapetListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Clapet.objects.all().order_by('id')
    serializer_class = ClapetSerializer
    filterset_class = ClapetFilter


class ClapetDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Clapet.objects.all()
    serializer_class = ClapetSerializer


class CanalisationListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Canalisation.objects.all().order_by('id')
    serializer_class = CanalisationSerializer
    filterset_class = CanalisationFilter


class CanalisationDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Canalisation.objects.all()
    serializer_class = CanalisationSerializer


class AspersionListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Aspersion.objects.all().order_by('id')
    serializer_class = AspersionSerializer
    filterset_class = AspersionFilter


class AspersionDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Aspersion.objects.all()
    serializer_class = AspersionSerializer


class GoutteListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Goutte.objects.all().order_by('id')
    serializer_class = GoutteSerializer
    filterset_class = GoutteFilter


class GoutteDetailView(GISObjectPermissionMixin, generics.RetrieveUpdateDestroyAPIView):
    queryset = Goutte.objects.all()
    serializer_class = GoutteSerializer


class BallonListCreateView(GISObjectPermissionMixin, generics.ListCreateAPIView):
    queryset = Ballon.objects.all().order_by('id')
    serializer_class = BallonSerializer
    filterset_class = BallonFilter


class TankDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ballon.objects.all()
    serializer_class = BallonSerializer

# ==============================================================================
# VUE POUR LA RECHERCHE
# ==============================================================================

class SearchView(APIView):
    """
    Vue pour la recherche multicritère sur TOUS les types d'objets.
    Accepte un paramètre de requête `q`.
    Recherche dans Sites, SousSites, et tous les 15 types d'objets (végétation + hydraulique).

    🔒 FILTRAGE PAR RÔLE:
    - ADMIN: voit tout
    - CLIENT: voit uniquement ses sites/objets
    - SUPERVISEUR: voit uniquement les sites liés à ses équipes
    """
    def get(self, request, *args, **kwargs):
        query = request.query_params.get('q', '').strip()

        if len(query) < 2:
            return Response([])

        results = []

        # 🔒 Filtrage par rôle utilisateur
        user = request.user
        structure_filter = None
        site_ids_filter = None

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            # CLIENT: uniquement les sites de sa structure
            if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure

            # SUPERVISEUR: uniquement les sites de ses équipes
            elif 'SUPERVISEUR' in roles and not ('ADMIN' in roles):
                site_ids_filter = self._get_superviseur_site_ids(user)

        # Recherche sur les Sites (par nom ou code)
        site_query = Q(nom_site__icontains=query) | Q(code_site__icontains=query)
        sites_queryset = Site.objects.filter(site_query)

        # Appliquer le filtrage par rôle
        if structure_filter:
            sites_queryset = sites_queryset.filter(structure_client=structure_filter)
        elif site_ids_filter is not None:
            sites_queryset = sites_queryset.filter(id__in=site_ids_filter)

        sites = sites_queryset.distinct()[:10]  # ✅ Limit to 10 unique sites
        for item in sites:
            location = item.centroid
            results.append({
                'id': f"site-{item.pk}",
                'name': f"{item.nom_site}",
                'type': 'Site',
                'location': {'type': 'Point', 'coordinates': [location.x, location.y]} if location else None,
            })

        # Recherche sur les Sous-Sites (par nom)
        sous_sites_queryset = SousSite.objects.filter(nom__icontains=query)

        # Appliquer le filtrage par rôle
        if structure_filter:
            sous_sites_queryset = sous_sites_queryset.filter(site__structure_client=structure_filter)
        elif site_ids_filter is not None:
            sous_sites_queryset = sous_sites_queryset.filter(site_id__in=site_ids_filter)

        sous_sites = sous_sites_queryset.distinct()[:5]  # ✅ Limit to 5 unique sous-sites
        for item in sous_sites:
            location = item.geometrie
            results.append({
                'id': f"soussite-{item.pk}",
                'name': f"{item.nom} ({item.site.nom_site})",
                'type': 'Sous-site',
                'location': {'type': 'Point', 'coordinates': [location.x, location.y]} if location else None,
            })

        # ✅ Recherche sur TOUS les types d'objets (végétation + hydraulique)
        # Mapping: (Model, type_name, id_prefix)
        object_models = [
            # Végétation (7 types)
            (Arbre, 'Arbre', 'arbre'),
            (Gazon, 'Gazon', 'gazon'),
            (Palmier, 'Palmier', 'palmier'),
            (Arbuste, 'Arbuste', 'arbuste'),
            (Vivace, 'Vivace', 'vivace'),
            (Cactus, 'Cactus', 'cactus'),
            (Graminee, 'Graminée', 'graminee'),
            # Hydraulique (8 types)
            (Puit, 'Puit', 'puit'),
            (Pompe, 'Pompe', 'pompe'),
            (Vanne, 'Vanne', 'vanne'),
            (Clapet, 'Clapet', 'clapet'),
            (Canalisation, 'Canalisation', 'canalisation'),
            (Aspersion, 'Aspersion', 'aspersion'),
            (Goutte, 'Goutte', 'goutte'),
            (Ballon, 'Ballon', 'ballon'),
        ]

        for Model, type_name, id_prefix in object_models:
            try:
                # Recherche par nom (ou marque pour certains types hydrauliques)
                query_filter = Q(nom__icontains=query)
                if hasattr(Model, 'marque'):
                    query_filter |= Q(marque__icontains=query)
                if hasattr(Model, 'famille'):
                    query_filter |= Q(famille__icontains=query)

                objects_queryset = Model.objects.filter(query_filter).select_related('site')

                # 🔒 Appliquer le filtrage par rôle
                if structure_filter:
                    objects_queryset = objects_queryset.filter(site__structure_client=structure_filter)
                elif site_ids_filter is not None:
                    objects_queryset = objects_queryset.filter(site_id__in=site_ids_filter)

                objects = objects_queryset[:5]  # Max 5 par type
                for obj in objects:
                    try:
                        location = obj.geometry if hasattr(obj, 'geometry') else None
                        site_name = obj.site.nom_site if hasattr(obj, 'site') and obj.site else 'Inconnu'

                        # Centroid pour polygones/lignes
                        if location and location.geom_type in ['Polygon', 'LineString', 'MultiPolygon', 'MultiLineString']:
                            location = location.centroid

                        # Nom de l'objet
                        obj_name = obj.nom if hasattr(obj, 'nom') and obj.nom else f"{type_name} #{obj.pk}"

                        results.append({
                            'id': f"{id_prefix}-{obj.pk}",
                            'name': f"{obj_name} ({site_name})",
                            'type': type_name,
                            'location': {'type': 'Point', 'coordinates': [location.x, location.y]} if location else None,
                        })
                    except Exception as e:
                        # Log error but continue with next object
                        print(f"Error processing {type_name} #{obj.pk}: {str(e)}")
                        continue

            except Exception as e:
                # Log error but continue with next model type
                print(f"Error searching {type_name}: {str(e)}")
                continue

        # Limiter le nombre total de résultats
        return Response(results[:30])

    def _get_superviseur_site_ids(self, user):
        """
        Récupère les IDs des sites contenant les objets liés aux tâches du superviseur.
        """
        from api_planification.models import Tache

        try:
            superviseur = user.superviseur_profile
            equipes_gerees = superviseur.equipes_gerees.filter(actif=True)
            equipes_gerees_ids = list(equipes_gerees.values_list('id', flat=True))

            if not equipes_gerees_ids:
                return []

            taches_ids = Tache.objects.filter(
                deleted_at__isnull=True
            ).filter(
                Q(equipes__id__in=equipes_gerees_ids) | Q(id_equipe__in=equipes_gerees_ids)
            ).values_list('id', flat=True).distinct()

            site_ids = list(Objet.objects.filter(
                taches__id__in=taches_ids,
                site_id__isnull=False
            ).values_list('site_id', flat=True).distinct())

            return site_ids
        except Exception:
            return []


# ==============================================================================
# VUE POUR L'EXPORT PDF
# ==============================================================================

class ExportPDFView(APIView):
    """
    Vue pour exporter la carte en PDF.
    Accepte les paramètres: title, mapImageBase64, visibleLayers, center, zoom
    """
    def post(self, request, *args, **kwargs):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib.utils import ImageReader
        from django.http import HttpResponse
        from datetime import datetime
        import base64
        import io

        # Récupérer les données du POST
        title = request.data.get('title', 'Export Carte GreenSIG')
        map_image_base64 = request.data.get('mapImageBase64', '')
        visible_layers = request.data.get('visibleLayers', {})
        center = request.data.get('center', [0, 0])
        zoom = request.data.get('zoom', 15)
        site_names = request.data.get('siteNames', [])  # Liste des noms de sites visibles

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        page_width, page_height = landscape(A4)
        pdf = canvas.Canvas(buffer, pagesize=landscape(A4))

        # Utilisateur
        user = request.user
        user_name = user.get_full_name() if hasattr(user, 'get_full_name') else f"{getattr(user, 'prenom', '')} {getattr(user, 'nom', user.username)}".strip()
        user_info = f"Exporté par: {user_name} ({user.email})" if user.email else f"Exporté par: {user_name}"
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(2*cm, page_height - 2*cm, user_info)

        # Site(s) - Afficher tous les sites
        if site_names:
            sites_text = f"Site(s): {', '.join(site_names)}"
            pdf.setFont("Helvetica", 10)

            # Si le texte est trop long, utiliser une police plus petite ou couper en lignes
            max_width = page_width - 10*cm  # Largeur disponible (en laissant de la marge pour la légende)
            text_width = pdf.stringWidth(sites_text, "Helvetica", 10)

            if text_width > max_width:
                # Réduire la taille de la police si nécessaire
                pdf.setFont("Helvetica", 8)
                text_width = pdf.stringWidth(sites_text, "Helvetica", 8)

                if text_width > max_width:
                    # Si toujours trop long, couper en plusieurs lignes
                    pdf.setFont("Helvetica", 8)
                    words = sites_text.split(', ')
                    lines = []
                    current_line = words[0]

                    for word in words[1:]:
                        test_line = current_line + ', ' + word
                        if pdf.stringWidth(test_line, "Helvetica", 8) <= max_width:
                            current_line = test_line
                        else:
                            lines.append(current_line)
                            current_line = word
                    lines.append(current_line)

                    y_pos = page_height - 2.6*cm
                    for line in lines:
                        pdf.drawString(2*cm, y_pos, line)
                        y_pos -= 0.4*cm
                    date_y = y_pos - 0.2*cm
                else:
                    pdf.drawString(2*cm, page_height - 2.6*cm, sites_text)
                    date_y = page_height - 3.2*cm
            else:
                pdf.drawString(2*cm, page_height - 2.6*cm, sites_text)
                date_y = page_height - 3.2*cm
        else:
            date_y = page_height - 2.6*cm

        # Date
        pdf.setFont("Helvetica", 10)
        date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        pdf.drawString(2*cm, date_y, f"Date d'export: {date_str}")

        # Logo GreenSIG (en haut à droite)
        try:
            import os
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo_width = 3*cm
                logo_height = 1.5*cm
                logo_x = page_width - logo_width - 2*cm
                logo_y = page_height - 2.5*cm
                pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            pass  # Si le logo n'est pas disponible, on continue sans

        # Image de la carte (si fournie)
        if map_image_base64:
            try:
                # Décoder l'image base64
                image_data = base64.b64decode(map_image_base64.split(',')[1] if ',' in map_image_base64 else map_image_base64)
                image = ImageReader(io.BytesIO(image_data))

                # Dessiner l'image (centré, 70% de la largeur)
                img_width = page_width * 0.7
                img_height = (page_height - 6*cm) * 0.7
                img_x = 2*cm
                img_y = page_height - 5*cm - img_height

                pdf.drawImage(image, img_x, img_y, width=img_width, height=img_height, preserveAspectRatio=True)
            except Exception as e:
                pdf.setFont("Helvetica", 10)
                pdf.drawString(2*cm, page_height - 5*cm, f"Erreur lors du chargement de l'image: {str(e)}")

        # Légende (à droite de l'image)
        legend_x = page_width - 6*cm
        legend_y = page_height - 3*cm

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(legend_x, legend_y, "Légende")
        legend_y -= 0.6*cm

        # Définition des catégories et éléments (Synchronisé avec constants.ts)
        legend_data = [
            {
                "title": "SITES",
                "items": [
                    {"name": "Sites", "color": (59, 130, 246)},      # #3b82f6
                ]
            },
            {
                "title": "VÉGÉTATION",
                "items": [
                    {"name": "Arbres", "color": (5, 150, 105)},      # #059669
                    {"name": "Gazons", "color": (132, 204, 22)},     # #84cc16
                    {"name": "Palmiers", "color": (249, 115, 22)},    # #f97316
                    {"name": "Arbustes", "color": (16, 185, 129)},    # #10b981
                    {"name": "Vivaces", "color": (236, 72, 153)},     # #ec4899
                    {"name": "Cactus", "color": (6, 182, 212)},      # #06b6d4
                    {"name": "Graminées", "color": (234, 179, 8)},     # #eab308
                ]
            },
            {
                "title": "HYDRAULIQUE",
                "items": [
                    {"name": "Puits", "color": (14, 165, 233)},      # #0ea5e9
                    {"name": "Pompes", "color": (6, 182, 212)},      # #06b6d4
                    {"name": "Vannes", "color": (20, 184, 166)},     # #14b8a6
                    {"name": "Clapets", "color": (8, 145, 178)},     # #0891b2
                    {"name": "Canalisations", "color": (2, 132, 199)}, # #0284c7
                    {"name": "Aspersions", "color": (56, 189, 248)}, # #38bdf8
                    {"name": "Gouttes", "color": (125, 211, 252)},   # #7dd3fc
                    {"name": "Ballons", "color": (3, 105, 161)},     # #0369a1
                ]
            },
            {
                "title": "RÉCLAMATIONS",
                "items": [
                    {"name": "En attente de lecture", "color": (239, 68, 68)},      # #ef4444
                    {"name": "Prise en compte", "color": (249, 115, 22)}, # #f97316
                    {"name": "En attente de réalisation", "color": (234, 179, 8)},       # #eab308
                    {"name": "Tâche terminée (côté admin.)", "color": (34, 197, 94)},        # #22c55e
                    {"name": "Clôturée (par le client)", "color": (16, 185, 129)},    # #10b981
                ]
            }
        ]

        # Rendre la légende complète
        for category in legend_data:
            # Titre de catégorie
            pdf.setFont("Helvetica-Bold", 9)
            pdf.setFillColorRGB(0.4, 0.4, 0.4)
            pdf.drawString(legend_x, legend_y, category["title"])
            legend_y -= 0.45*cm

            # Cas spécial pour les réclamations : afficher les symbologies
            if category["title"] == "RÉCLAMATIONS":
                for item in category["items"]:
                    color = item["color"]
                    
                    # Symbole Point: Cercle avec "!"
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.circle(legend_x + 0.25*cm, legend_y + 0.1*cm, 0.12*cm, fill=1)
                    pdf.setFillColorRGB(1, 1, 1)  # Blanc pour le "!"
                    pdf.setFont("Helvetica-Bold", 7)
                    pdf.drawString(legend_x + 0.22*cm, legend_y + 0.05*cm, "!")
                    
                    # Symbole Polygon: Rectangle avec bordure pointillée
                    pdf.setStrokeColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255, alpha=0.25)
                    pdf.setLineWidth(1.5)
                    pdf.setDash([2, 1])  # Ligne pointillée
                    pdf.rect(legend_x + 0.55*cm, legend_y + 0.02*cm, 0.2*cm, 0.16*cm, fill=1, stroke=1)
                    pdf.setDash([])  # Reset dash
                    
                    # Texte
                    pdf.setFillColorRGB(0, 0, 0)
                    pdf.setFont("Helvetica", 8)
                    pdf.drawString(legend_x + 0.9*cm, legend_y, item["name"])
                    legend_y -= 0.4*cm
            else:
                # Autres catégories: symbologie normale
                for item in category["items"]:
                    # Puce de couleur
                    color = item["color"]
                    pdf.setFillColorRGB(color[0]/255, color[1]/255, color[2]/255)
                    pdf.circle(legend_x + 0.2*cm, legend_y + 0.1*cm, 0.15*cm, fill=1)

                    # Texte
                    pdf.setFillColorRGB(0, 0, 0)
                    pdf.setFont("Helvetica", 8)
                    pdf.drawString(legend_x + 0.6*cm, legend_y, item["name"])
                    legend_y -= 0.4*cm
            
            legend_y -= 0.2*cm

        # Informations de la vue
        info_y = 2*cm
        pdf.setFont("Helvetica", 8)
        pdf.drawString(2*cm, info_y, f"Centre: [{center[0]:.6f}, {center[1]:.6f}] | Zoom: {zoom}")

        # Finaliser le PDF
        pdf.showPage()
        pdf.save()

        # Préparer la réponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"carte_greensig_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


# ==============================================================================
# VUE POUR LES STATISTIQUES
# ==============================================================================

class StatisticsView(APIView):
    """
    Vue pour retourner les statistiques contextuelles selon le rôle de l'utilisateur.
    - ADMIN: statistiques globales de tout le système
    - CLIENT: statistiques de ses sites uniquement
    - SUPERVISEUR: statistiques des sites qui lui sont affectés
    """

    def _get_filtered_querysets(self, request):
        """
        Retourne un dictionnaire de querysets filtrés selon le rôle de l'utilisateur.
        Chaque queryset est filtré automatiquement pour respecter les permissions.
        """
        user = request.user
        querysets = {}

        # Déterminer le filtre à appliquer
        site_filter = None
        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            if 'ADMIN' in roles:
                # ADMIN: pas de filtre, voit tout
                site_filter = Q()
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                # CLIENT: uniquement les sites de sa structure
                structure = user.client_profile.structure
                if structure:
                    site_filter = Q(site__structure_client=structure)
                else:
                    site_filter = Q(pk__in=[])  # Queryset vide
            elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                # SUPERVISEUR: uniquement les sites qui lui sont affectés
                site_filter = Q(site__superviseur=user.superviseur_profile)
            else:
                # Aucun accès
                site_filter = Q(pk__in=[])  # Queryset vide

        # Filtrer Sites et SousSites séparément (pas de champ 'site')
        if user.is_authenticated and 'ADMIN' in [ur.role.nom_role for ur in user.roles_utilisateur.all()]:
            querysets['Site'] = Site.objects.all()
            querysets['SousSite'] = SousSite.objects.all()
        elif user.is_authenticated and 'CLIENT' in [ur.role.nom_role for ur in user.roles_utilisateur.all()]:
            structure = user.client_profile.structure if hasattr(user, 'client_profile') else None
            if structure:
                querysets['Site'] = Site.objects.filter(structure_client=structure)
                querysets['SousSite'] = SousSite.objects.filter(site__structure_client=structure)
            else:
                querysets['Site'] = Site.objects.none()
                querysets['SousSite'] = SousSite.objects.none()
        elif user.is_authenticated and 'SUPERVISEUR' in [ur.role.nom_role for ur in user.roles_utilisateur.all()]:
            querysets['Site'] = Site.objects.filter(superviseur=user.superviseur_profile)
            querysets['SousSite'] = SousSite.objects.filter(site__superviseur=user.superviseur_profile)
        else:
            querysets['Site'] = Site.objects.none()
            querysets['SousSite'] = SousSite.objects.none()

        # Appliquer le filtre site à tous les objets GIS
        for model in [Arbre, Gazon, Palmier, Arbuste, Vivace, Cactus, Graminee,
                      Puit, Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte, Ballon]:
            if site_filter is not None:
                querysets[model.__name__] = model.objects.filter(site_filter)
            else:
                querysets[model.__name__] = model.objects.none()

        return querysets

    def get(self, request, *args, **kwargs):
        from django.db.models import Count, Avg, Sum, Max, Min, Q
        from django.apps import apps
        from django.utils import timezone

        # Obtenir les querysets filtrés selon le rôle
        qs = self._get_filtered_querysets(request)

        # Statistiques filtrées selon les permissions
        statistics = {
            # Statistiques de hiérarchie
            'hierarchy': {
                'total_sites': qs["Site"].count(),
                'total_sous_sites': qs["SousSite"].count(),
                'active_sites': qs["Site"].filter(actif=True).count(),
            },

            # Statistiques végétaux
            'vegetation': {
                'arbres': {
                    'total': qs["Arbre"].count(),
                    'by_taille': dict(qs["Arbre"].values('taille').annotate(count=Count('id')).values_list('taille', 'count')),
                    'top_families': list(qs["Arbre"].values('famille').annotate(count=Count('id')).order_by('-count')[:5].values('famille', 'count'))
                },
                'gazons': {
                    'total': qs["Gazon"].count(),
                    'total_area_sqm': qs["Gazon"].aggregate(Sum('area_sqm'))['area_sqm__sum'] or 0,
                },
                'palmiers': {
                    'total': qs["Palmier"].count(),
                    'by_taille': dict(qs["Palmier"].values('taille').annotate(count=Count('id')).values_list('taille', 'count')),
                },
                'arbustes': {
                    'total': qs["Arbuste"].count(),
                    'avg_densite': qs["Arbuste"].aggregate(Avg('densite'))['densite__avg'] or 0,
                },
                'vivaces': {
                    'total': qs["Vivace"].count(),
                },
                'cactus': {
                    'total': qs["Cactus"].count(),
                },
                'graminees': {
                    'total': qs["Graminee"].count(),
                }
            },

            # Statistiques hydraulique
            'hydraulique': {
                'puits': {
                    'total': qs["Puit"].count(),
                    'avg_profondeur': qs["Puit"].aggregate(Avg('profondeur'))['profondeur__avg'] or 0,
                    'max_profondeur': qs["Puit"].aggregate(Max('profondeur'))['profondeur__max'] or 0,
                },
                'pompes': {
                    'total': qs["Pompe"].count(),
                    'avg_puissance': qs["Pompe"].aggregate(Avg('puissance'))['puissance__avg'] or 0,
                    'avg_debit': qs["Pompe"].aggregate(Avg('debit'))['debit__avg'] or 0,
                },
                'vannes': {
                    'total': qs["Vanne"].count(),
                },
                'clapets': {
                    'total': qs["Clapet"].count(),
                },
                'canalisations': {
                    'total': qs["Canalisation"].count(),
                },
                'aspersions': {
                    'total': qs["Aspersion"].count(),
                },
                'gouttes': {
                    'total': qs["Goutte"].count(),
                },
                'ballons': {
                    'total': qs["Ballon"].count(),
                    'total_volume': qs["Ballon"].aggregate(Sum('volume'))['volume__sum'] or 0,
                }
            },

            # Statistiques globales
            'global': {
                'total_objets': (
                    qs["Arbre"].count() + qs["Gazon"].count() + qs["Palmier"].count() +
                    qs["Arbuste"].count() + qs["Vivace"].count() + qs["Cactus"].count() +
                    qs["Graminee"].count() + qs["Puit"].count() + qs["Pompe"].count() +
                    qs["Vanne"].count() + qs["Clapet"].count() + qs["Canalisation"].count() +
                    qs["Aspersion"].count() + qs["Goutte"].count() + qs["Ballon"].count()
                ),
                'total_vegetation': (
                    qs["Arbre"].count() + qs["Gazon"].count() + qs["Palmier"].count() +
                    qs["Arbuste"].count() + qs["Vivace"].count() + qs["Cactus"].count() +
                    qs["Graminee"].count()
                ),
                'total_hydraulique': (
                    qs["Puit"].count() + qs["Pompe"].count() + qs["Vanne"].count() +
                    qs["Clapet"].count() + qs["Canalisation"].count() + qs["Aspersion"].count() +
                    qs["Goutte"].count() + qs["Ballon"].count()
                )
            }
        }
        
        # --- LOGIQUE SPÉCIFIQUE POUR SUPERVISEUR ---
        user = request.user
        if user.is_authenticated:
            is_superviseur = user.roles_utilisateur.filter(role__nom_role='SUPERVISEUR').exists()
            if is_superviseur:
                try:
                    Tache = apps.get_model('api_planification', 'Tache')
                    Absence = apps.get_model('api_users', 'Absence')

                    # Récupérer le superviseur lié
                    superviseur = getattr(user, 'superviseur_profile', None)
                    if superviseur:
                        # Ses équipes gérées
                        mes_equipes = superviseur.equipes_gerees.filter(actif=True)
                        mes_equipes_ids = list(mes_equipes.values_list('id', flat=True))

                        # Tâches de ses équipes
                        mes_taches = Tache.objects.filter(
                            Q(equipes__id__in=mes_equipes_ids) | Q(id_equipe__in=mes_equipes_ids),
                            deleted_at__isnull=True
                        ).distinct()

                        # Absences dans ses équipes (membres)
                        # Récupérer tous les membres de ses équipes
                        membres_ids = set()
                        for eq in mes_equipes:
                            membres_ids.update(eq.operateurs.values_list('id', flat=True))

                        today = timezone.now().date()
                        absences_today = Absence.objects.filter(
                            operateur__id__in=membres_ids,
                            date_debut__lte=today,
                            date_fin__gte=today,
                            statut='VALIDEE'
                        ).count()

                        statistics['superviseur_stats'] = {
                            'taches_today': mes_taches.filter(date_debut_planifiee__date=today).count(),
                            'taches_en_cours': mes_taches.filter(statut='EN_COURS').count(),
                            'taches_a_faire': mes_taches.filter(statut='A_FAIRE').count(),
                            'taches_retard': mes_taches.filter(statut='EN_RETARD').count(), 
                            'absences_today': absences_today,
                            'equipes_count': len(mes_equipes_ids)
                        }
                except Exception as e:
                    print(f"Error calculating superviseur stats: {e}")
                    # Ne pas bloquer la réponse si erreur dans les stats spécifiques

        return Response(statistics)


# ==============================================================================
# VUES POUR L'EXPORT DE DONNÉES
# ==============================================================================

class ExportDataView(APIView):
    """
    Vue générique pour exporter les données en Excel, GeoJSON, KML ou Shapefile.
    Paramètres de requête:
    - model: nom du modèle (arbres, gazons, palmiers, etc.)
    - format: xlsx, geojson, kml, shp (défaut: xlsx)
    - ids: optionnel, liste d'IDs séparés par virgules pour export sélectif
    - filtres: optionnels (même syntaxe que les endpoints de liste)
    """

    MODEL_MAPPING = {
        'sites': Site,
        'sous-sites': SousSite,
        'arbres': Arbre,
        'gazons': Gazon,
        'palmiers': Palmier,
        'arbustes': Arbuste,
        'vivaces': Vivace,
        'cactus': Cactus,
        'graminees': Graminee,
        'puits': Puit,
        'pompes': Pompe,
        'vannes': Vanne,
        'clapets': Clapet,
        'canalisations': Canalisation,
        'aspersions': Aspersion,
        'gouttes': Goutte,
        'ballons': Ballon,
    }

    def get(self, request, model_name, *args, **kwargs):
        from openpyxl import Workbook
        from django.http import HttpResponse
        from datetime import datetime

        # Vérifier que le modèle existe
        if model_name not in self.MODEL_MAPPING:
            return Response({'error': f'Modèle invalide: {model_name}'}, status=400)

        model_class = self.MODEL_MAPPING[model_name]

        # Récupérer le format d'export
        export_format = request.query_params.get('format', 'xlsx').lower()
        valid_formats = ['xlsx', 'geojson', 'kml', 'shp', 'shapefile']
        if export_format not in valid_formats:
            return Response({'error': f'Format invalide. Utilisez: {", ".join(valid_formats)}'}, status=400)

        # Normaliser shapefile
        if export_format == 'shapefile':
            export_format = 'shp'

        # Appliquer les filtres
        queryset = model_class.objects.all()

        # Filtre par IDs si fourni
        ids_param = request.query_params.get('ids', '')
        if ids_param:
            try:
                ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
                queryset = queryset.filter(pk__in=ids)
            except ValueError:
                return Response({'error': 'ids parameter must be comma-separated integers'}, status=400)

        # Filtre par site si fourni
        site_id = request.query_params.get('site')
        if site_id:
            queryset = queryset.filter(site_id=site_id)

        # Vérifier qu'il y a des données
        if not queryset.exists():
            return Response({'error': 'Aucune donnée à exporter'}, status=404)

        # ==============================================================================
        # EXPORT GeoJSON
        # ==============================================================================
        if export_format == 'geojson':
            from .services.geo_io import export_to_geojson

            geojson_data = export_to_geojson(queryset)

            response = HttpResponse(
                json.dumps(geojson_data, ensure_ascii=False, indent=2),
                content_type='application/geo+json; charset=utf-8'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.geojson"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # ==============================================================================
        # EXPORT KML
        # ==============================================================================
        if export_format == 'kml':
            from .services.geo_io import export_to_kml

            kml_content = export_to_kml(queryset)

            response = HttpResponse(
                kml_content,
                content_type='application/vnd.google-earth.kml+xml; charset=utf-8'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.kml"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # ==============================================================================
        # EXPORT Shapefile (ZIP)
        # ==============================================================================
        if export_format == 'shp':
            from .services.geo_io import export_to_shapefile

            try:
                zip_content = export_to_shapefile(queryset, model_name)

                response = HttpResponse(
                    zip_content,
                    content_type='application/zip'
                )
                filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError as e:
                return Response({'error': str(e)}, status=500)
            except Exception as e:
                return Response({'error': f'Shapefile export error: {str(e)}'}, status=500)

        # ==============================================================================
        # EXPORT XLSX
        # ==============================================================================
        # Récupérer tous les objets (pas de pagination pour l'export)
        objects = list(queryset.values())

        # Obtenir les noms de colonnes
        field_names = list(objects[0].keys())

        # Export Excel
        if export_format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = model_name[:31]  # Excel limite à 31 caractères

            # Écrire l'en-tête
            ws.append(field_names)

            # Écrire les données
            for obj in objects:
                ws.append([obj[field] for field in field_names])

            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Sauvegarder dans un buffer
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response


# ==============================================================================
# VUE EXPORT EXCEL AMÉLIORÉ POUR INVENTAIRE
# ==============================================================================

class InventoryExportExcelView(APIView):
    """
    Vue spécialisée pour l'export Excel professionnel de l'inventaire.
    Supporte:
    - Formatage professionnel (styles, couleurs, filtres Excel)
    - Export multi-types avec onglets séparés
    - Filtres passés depuis le frontend
    - Statistiques par type

    Paramètres de requête:
    - types: liste de types séparés par virgules (ex: 'Arbre,Palmier,Gazon')
    - site: ID du site (optionnel)
    - etat: état des objets (bon, moyen, mauvais, critique)
    - famille: famille botanique (optionnel)
    - search: recherche textuelle (optionnel)
    """

    MODEL_MAPPING = {
        'Arbre': Arbre,
        'Palmier': Palmier,
        'Gazon': Gazon,
        'Arbuste': Arbuste,
        'Vivace': Vivace,
        'Cactus': Cactus,
        'Graminee': Graminee,
        'Puit': Puit,
        'Pompe': Pompe,
        'Vanne': Vanne,
        'Clapet': Clapet,
        'Canalisation': Canalisation,
        'Aspersion': Aspersion,
        'Goutte': Goutte,
        'Ballon': Ballon,
    }

    # Mapping des champs à exporter par type (colonnes personnalisées)
    # Note: 'superficie_calculee' est une annotation calculée dynamiquement
    # Note: 'derniere_intervention' est calculée depuis les tâches si last_intervention_date est null
    FIELD_MAPPINGS = {
        'Arbre': ['nom', 'famille', 'taille', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Palmier': ['nom', 'famille', 'taille', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Gazon': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Arbuste': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Vivace': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Cactus': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Graminee': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Puit': ['nom', 'profondeur', 'diametre', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Pompe': ['nom', 'type', 'diametre', 'puissance', 'debit', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Vanne': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Clapet': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Canalisation': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Aspersion': ['marque', 'type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Goutte': ['type', 'diametre', 'materiau', 'pression', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
        'Ballon': ['marque', 'pression', 'volume', 'materiau', 'site__nom_site', 'sous_site__nom', 'etat', 'derniere_intervention'],
    }

    # Types qui ont une géométrie polygone (pour calcul de surface)
    POLYGON_TYPES = ['Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee']

    # Types qui ont le champ last_intervention_date
    TYPES_WITH_INTERVENTION_DATE = [
        'Arbre', 'Palmier', 'Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee',
        'Puit', 'Pompe'
    ]

    # Labels français pour les colonnes
    FIELD_LABELS = {
        'nom': 'Nom',
        'marque': 'Marque',
        'famille': 'Famille',
        'taille': 'Taille',
        'densite': 'Densité',
        'area_sqm': 'Surface (m²)',
        'superficie_calculee': 'Surface (m²)',
        'profondeur': 'Profondeur (m)',
        'diametre': 'Diamètre (cm)',
        'type': 'Type',
        'puissance': 'Puissance (kW)',
        'debit': 'Débit (m³/h)',
        'materiau': 'Matériau',
        'pression': 'Pression (bar)',
        'volume': 'Volume (L)',
        'site__nom_site': 'Site',
        'sous_site__nom': 'Sous-site',
        'etat': 'État',
        'last_intervention_date': 'Dernière intervention',
        'derniere_intervention': 'Dernière intervention',
    }

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as ExcelImage
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        import os
        from django.conf import settings

        # Récupérer les types à exporter
        types_param = request.query_params.get('types', '')
        if types_param:
            types_list = [t.strip() for t in types_param.split(',') if t.strip()]
        else:
            # Si aucun type spécifié, exporter tous
            types_list = list(self.MODEL_MAPPING.keys())

        # Valider les types
        invalid_types = [t for t in types_list if t not in self.MODEL_MAPPING]
        if invalid_types:
            return Response({'error': f'Types invalides: {", ".join(invalid_types)}'}, status=400)

        # Créer le workbook
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        # Styles réutilisables
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        stats_font = Font(bold=True, size=10)
        stats_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

        # États et leurs couleurs
        etat_colors = {
            'bon': 'C8E6C9',        # Vert clair
            'moyen': 'FFF9C4',      # Jaune clair
            'mauvais': 'FFCCBC',    # Orange clair
            'critique': 'FFCDD2',   # Rouge clair
        }

        # Filtrer par rôle (ADMIN, CLIENT, SUPERVISEUR)
        user = request.user
        structure_filter = None
        superviseur_filter = None
        if user.is_authenticated:
            roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

            if 'ADMIN' in roles:
                pass  # ADMIN voit tout
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                superviseur_filter = user.superviseur_profile

        # Pour chaque type, créer un onglet
        for type_name in types_list:
            model_class = self.MODEL_MAPPING[type_name]

            # Construire le queryset avec filtres
            queryset = model_class.objects.select_related('site', 'sous_site').all()

            # Filtrer par rôle
            if structure_filter:
                queryset = queryset.filter(site__structure_client=structure_filter)
            elif superviseur_filter:
                queryset = queryset.filter(site__superviseur=superviseur_filter)

            # Appliquer les filtres
            site_id = request.query_params.get('site')
            if site_id:
                queryset = queryset.filter(site_id=site_id)

            etat = request.query_params.get('etat')
            if etat:
                queryset = queryset.filter(etat=etat)

            famille = request.query_params.get('famille')
            if famille and hasattr(model_class, 'famille'):
                queryset = queryset.filter(famille__icontains=famille)

            search = request.query_params.get('search')
            if search:
                # Recherche dans nom ou marque selon le type
                if hasattr(model_class, 'nom'):
                    queryset = queryset.filter(nom__icontains=search)
                elif hasattr(model_class, 'marque'):
                    queryset = queryset.filter(marque__icontains=search)

            # Vérifier s'il y a des données
            if not queryset.exists():
                continue

            # Créer l'onglet
            ws = wb.create_sheet(title=type_name[:31])

            # Récupérer les champs à exporter
            fields = self.FIELD_MAPPINGS.get(type_name, ['nom', 'site__nom_site', 'etat'])

            # Écrire l'en-tête
            headers = [self.FIELD_LABELS.get(field, field) for field in fields]
            ws.append(headers)

            # Appliquer le style d'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            # Écrire les données - itérer directement sur les objets
            for obj in queryset:
                row_values = []
                etat_value = None

                for field in fields:
                    value = None

                    # Gérer les champs calculés spéciaux
                    if field == 'superficie_calculee':
                        # Calculer la surface depuis la géométrie (pour polygones)
                        if type_name in self.POLYGON_TYPES and obj.geometry:
                            try:
                                # Utiliser la méthode area de GEOS (retourne m² pour géométries géographiques)
                                # Transformer en projection métrique pour calcul précis
                                from django.contrib.gis.geos import GEOSGeometry
                                geom = obj.geometry
                                # Transformer en Web Mercator (EPSG:3857) pour calcul en mètres
                                geom_projected = geom.transform(3857, clone=True)
                                value = geom_projected.area  # Surface en m²
                            except Exception:
                                value = None

                    elif field == 'derniere_intervention':
                        # Priorité au champ existant, sinon chercher dans les tâches
                        if type_name in self.TYPES_WITH_INTERVENTION_DATE and hasattr(obj, 'last_intervention_date'):
                            value = obj.last_intervention_date
                        if not value:
                            # Chercher la dernière tâche terminée liée à cet objet
                            from api_planification.models import Tache
                            derniere_tache = Tache.objects.filter(
                                objets__id=obj.objet_ptr_id,
                                statut='TERMINEE',
                                deleted_at__isnull=True
                            ).order_by('-date_fin_reelle').values('date_fin_reelle').first()
                            if derniere_tache:
                                value = derniere_tache['date_fin_reelle']

                    elif field == 'site__nom_site':
                        # Champ lié: site.nom_site
                        value = obj.site.nom_site if obj.site else None

                    elif field == 'sous_site__nom':
                        # Champ lié: sous_site.nom
                        value = obj.sous_site.nom if obj.sous_site else None

                    else:
                        # Champ standard - accès direct à l'attribut
                        value = getattr(obj, field, None)

                    # Formater les dates
                    if field in ['last_intervention_date', 'derniere_intervention'] and value:
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%d/%m/%Y')

                    # Formater les nombres
                    elif field in ['area_sqm', 'superficie_calculee', 'profondeur', 'diametre', 'puissance', 'debit', 'pression', 'volume', 'densite']:
                        if value is not None:
                            value = round(float(value), 2)

                    # Capturer l'état pour la couleur
                    if field == 'etat':
                        etat_value = value

                    # Laisser vide si pas de valeur (au lieu de '-')
                    row_values.append(value if value is not None else '')

                ws.append(row_values)

                # Appliquer la couleur de fond selon l'état
                if etat_value and etat_value.lower() in etat_colors:
                    fill_color = etat_colors[etat_value.lower()]
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            # Ajouter une ligne de statistiques
            ws.append([])  # Ligne vide
            stats_row = ws.max_row + 1
            ws.append(['STATISTIQUES', '', '', '', '', '', ''])

            # Compter les états
            etat_counts = queryset.values('etat').annotate(count=Count('id'))
            etat_summary = {item['etat']: item['count'] for item in etat_counts}

            ws.append(['Total:', queryset.count()])
            for etat_key, count in etat_summary.items():
                ws.append([f'{etat_key.capitalize()}:', count])

            # Appliquer le style aux statistiques
            for row in ws.iter_rows(min_row=stats_row, max_row=ws.max_row):
                for cell in row:
                    cell.font = stats_font
                    cell.fill = stats_fill

            # Ajuster les largeurs de colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Activer les filtres automatiques
            ws.auto_filter.ref = ws.dimensions

            # Figer la première ligne
            ws.freeze_panes = 'A2'

            # Ajouter le logo GreenSIG en haut à droite
            try:
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
                if os.path.exists(logo_path):
                    img = ExcelImage(logo_path)
                    # Redimensionner le logo (largeur en pixels)
                    img.width = 120
                    img.height = 60
                    # Positionner en haut à droite (colonne la plus à droite possible)
                    # Calculer la position basée sur le nombre de colonnes
                    last_col = chr(ord('A') + len(headers) - 1)
                    ws.add_image(img, f'{last_col}1')
            except Exception as e:
                pass  # Continuer sans logo si erreur

        # Vérifier qu'au moins un onglet a été créé
        if len(wb.sheetnames) == 0:
            return Response({'error': 'Aucune donnée à exporter avec les filtres appliqués'}, status=404)

        # Sauvegarder le workbook
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Créer la réponse HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


# ==============================================================================
# VUE EXPORT PDF POUR INVENTAIRE
# ==============================================================================

class InventoryExportPDFView(APIView):
    """
    Vue pour l'export PDF professionnel de l'inventaire.
    Génère un document PDF avec:
    - En-tête avec logo et titre
    - Tableau formaté des données
    - Statistiques par type
    - Pied de page avec numérotation

    Paramètres de requête identiques à InventoryExportExcelView:
    - types: liste de types séparés par virgules
    - site: ID du site (optionnel)
    - etat: état des objets
    - famille: famille botanique (optionnel)
    - search: recherche textuelle (optionnel)
    """

    MODEL_MAPPING = {
        'Arbre': Arbre,
        'Palmier': Palmier,
        'Gazon': Gazon,
        'Arbuste': Arbuste,
        'Vivace': Vivace,
        'Cactus': Cactus,
        'Graminee': Graminee,
        'Puit': Puit,
        'Pompe': Pompe,
        'Vanne': Vanne,
        'Clapet': Clapet,
        'Canalisation': Canalisation,
        'Aspersion': Aspersion,
        'Goutte': Goutte,
        'Ballon': Ballon,
    }

    FIELD_MAPPINGS = {
        'Arbre': ['nom', 'famille', 'taille', 'site__nom_site', 'etat'],
        'Palmier': ['nom', 'famille', 'taille', 'site__nom_site', 'etat'],
        'Gazon': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Arbuste': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Vivace': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Cactus': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Graminee': ['nom', 'famille', 'superficie_calculee', 'site__nom_site', 'etat'],
        'Puit': ['nom', 'profondeur', 'diametre', 'site__nom_site', 'etat'],
        'Pompe': ['nom', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Vanne': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Clapet': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Canalisation': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Aspersion': ['marque', 'type', 'diametre', 'site__nom_site', 'etat'],
        'Goutte': ['type', 'diametre', 'site__nom_site', 'etat'],
        'Ballon': ['marque', 'volume', 'site__nom_site', 'etat'],
    }

    # Types polygones nécessitant le calcul de superficie
    POLYGON_TYPES = ['Gazon', 'Arbuste', 'Vivace', 'Cactus', 'Graminee']

    FIELD_LABELS = {
        'nom': 'Nom',
        'marque': 'Marque',
        'famille': 'Famille',
        'taille': 'Taille',
        'densite': 'Densité',
        'area_sqm': 'Surface (m²)',
        'superficie_calculee': 'Surface (m²)',
        'profondeur': 'Prof. (m)',
        'diametre': 'Diam. (cm)',
        'type': 'Type',
        'volume': 'Vol. (L)',
        'site__nom_site': 'Site',
        'etat': 'État',
    }

    def get(self, request, *args, **kwargs):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        import os
        from django.conf import settings

        # Récupérer les types à exporter
        types_param = request.query_params.get('types', '')
        if types_param:
            types_list = [t.strip() for t in types_param.split(',') if t.strip()]
        else:
            types_list = list(self.MODEL_MAPPING.keys())

        # Valider les types
        invalid_types = [t for t in types_list if t not in self.MODEL_MAPPING]
        if invalid_types:
            return Response({'error': f'Types invalides: {", ".join(invalid_types)}'}, status=400)

        # Créer le buffer
        buffer = BytesIO()

        # Créer le document PDF (paysage pour plus de colonnes)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Styles
        styles = getSampleStyleSheet()

        # Style titre principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Style sous-titre
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Style pour les sections
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=10,
            spaceBefore=15,
            fontName='Helvetica-Bold'
        )

        # Contenu du document
        story = []

        # Logo GreenSIG en haut à droite avec titre
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=4*cm, height=2*cm)

                # Tableau avec titre à gauche et logo à droite
                header_data = [
                    [Paragraph("INVENTAIRE - GreenSIG", title_style), logo]
                ]
                header_table = Table(header_data, colWidths=[20*cm, 5*cm])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(header_table)
            else:
                # Pas de logo, titre seul
                story.append(Paragraph("INVENTAIRE - GreenSIG", title_style))
        except Exception:
            # En cas d'erreur, titre seul
            story.append(Paragraph("INVENTAIRE - GreenSIG", title_style))

        story.append(Paragraph(f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
        story.append(Spacer(1, 0.5*cm))

        # Couleurs par état
        etat_colors = {
            'bon': colors.HexColor('#C8E6C9'),
            'moyen': colors.HexColor('#FFF9C4'),
            'mauvais': colors.HexColor('#FFCCBC'),
            'critique': colors.HexColor('#FFCDD2'),
        }

        # Filtrer par rôle (ADMIN, CLIENT, SUPERVISEUR)
        user = request.user
        structure_filter = None
        superviseur_filter = None
        if user.is_authenticated:
            roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

            if 'ADMIN' in roles:
                pass  # ADMIN voit tout
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                superviseur_filter = user.superviseur_profile

        # Pour chaque type
        total_objects = 0
        for idx, type_name in enumerate(types_list):
            model_class = self.MODEL_MAPPING[type_name]

            # Construire le queryset avec filtres
            queryset = model_class.objects.select_related('site', 'sous_site').all()

            # Filtrer par rôle
            if structure_filter:
                queryset = queryset.filter(site__structure_client=structure_filter)
            elif superviseur_filter:
                queryset = queryset.filter(site__superviseur=superviseur_filter)

            # Appliquer les filtres (même logique que Excel)
            site_id = request.query_params.get('site')
            if site_id:
                queryset = queryset.filter(site_id=site_id)

            etat = request.query_params.get('etat')
            if etat:
                queryset = queryset.filter(etat=etat)

            famille = request.query_params.get('famille')
            if famille and hasattr(model_class, 'famille'):
                queryset = queryset.filter(famille__icontains=famille)

            search = request.query_params.get('search')
            if search:
                if hasattr(model_class, 'nom'):
                    queryset = queryset.filter(nom__icontains=search)
                elif hasattr(model_class, 'marque'):
                    queryset = queryset.filter(marque__icontains=search)

            # Vérifier s'il y a des données
            if not queryset.exists():
                continue

            count = queryset.count()
            total_objects += count

            # Titre de section
            story.append(Paragraph(f"{type_name} ({count} élément{'s' if count > 1 else ''})", section_style))

            # Récupérer les champs
            fields = self.FIELD_MAPPINGS.get(type_name, ['nom', 'site__nom_site', 'etat'])
            headers = [self.FIELD_LABELS.get(field, field) for field in fields]

            # Pour les types polygones, annoter avec le calcul de superficie
            if type_name in self.POLYGON_TYPES and 'superficie_calculee' in fields:
                from django.db.models.functions import Coalesce
                from django.db.models import Value, FloatField
                from django.contrib.gis.db.models.functions import Area, Transform

                # Calculer la surface en m² via projection UTM
                queryset = queryset.annotate(
                    superficie_calculee=Coalesce(
                        Area(Transform('geometry', 32629)),  # UTM zone 29N pour le Maroc
                        Value(0.0),
                        output_field=FloatField()
                    )
                )

            # Données - Afficher TOUS les éléments
            data_rows = list(queryset.values(*fields))

            # Construire le tableau
            table_data = [headers]

            for row_data in data_rows:
                row = []
                for field in fields:
                    value = row_data.get(field)

                    # Formater les valeurs
                    if value is None:
                        value = '-'
                    elif field in ['area_sqm', 'superficie_calculee', 'profondeur', 'diametre', 'volume']:
                        value = f"{round(float(value), 1)}" if value else '-'
                    else:
                        value = str(value)

                    row.append(value)

                table_data.append(row)

            # Calculer les statistiques pour les intégrer dans le tableau
            etat_counts = queryset.values('etat').annotate(count=Count('id'))
            etat_summary = {item['etat']: item['count'] for item in etat_counts}
            stats_text = "Répartition: " + " | ".join([f"{k.capitalize()}: {v}" for k, v in etat_summary.items()])

            # Ajouter une ligne de statistiques à la fin du tableau
            table_data.append([stats_text] + ['' for _ in range(len(fields) - 1)])
            stats_row_index = len(table_data) - 1

            # Créer le tableau
            col_widths = [3*cm] * len(fields)  # Largeur égale pour toutes les colonnes
            table = Table(table_data, colWidths=col_widths, repeatRows=1)  # repeatRows=1 pour répéter l'en-tête sur chaque page

            # Style du tableau
            table_style = TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

                # Corps
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2E7D32')),

                # Alternance de couleurs (exclure la dernière ligne de stats)
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),

                # Style de la ligne de statistiques (dernière ligne)
                ('SPAN', (0, stats_row_index), (-1, stats_row_index)),
                ('BACKGROUND', (0, stats_row_index), (-1, stats_row_index), colors.HexColor('#E8F5E9')),
                ('ALIGN', (0, stats_row_index), (-1, stats_row_index), 'CENTER'),
                ('FONTNAME', (0, stats_row_index), (-1, stats_row_index), 'Helvetica-Oblique'),
                ('TEXTCOLOR', (0, stats_row_index), (-1, stats_row_index), colors.HexColor('#2E7D32')),
            ])

            # Appliquer les couleurs par état si la colonne existe
            if 'etat' in fields:
                etat_col = fields.index('etat')
                for i, row_data in enumerate(data_rows, start=1):
                    etat_val = row_data.get('etat', '').lower()
                    if etat_val in etat_colors:
                        table_style.add('BACKGROUND', (etat_col, i), (etat_col, i), etat_colors[etat_val])

            table.setStyle(table_style)
            story.append(table)

            # Saut de page entre les types (sauf pour le dernier)
            if idx < len(types_list) - 1:
                story.append(PageBreak())
            else:
                story.append(Spacer(1, 0.5*cm))

        # Vérifier qu'il y a des données
        if total_objects == 0:
            return Response({'error': 'Aucune donnée à exporter avec les filtres appliqués'}, status=404)

        # Résumé final
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph("RÉSUMÉ GLOBAL", section_style))
        summary_table_data = [
            ['Total d\'objets exportés:', str(total_objects)],
            ['Nombre de types:', str(len([t for t in types_list if self.MODEL_MAPPING[t].objects.filter(**self._build_filters(request)).exists()]))],
            ['Date d\'export:', datetime.now().strftime('%d/%m/%Y à %H:%M')],
        ]
        summary_table = Table(summary_table_data, colWidths=[6*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)

        # Générer le PDF
        doc.build(story)

        # Retourner la réponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _build_filters(self, request):
        """Construit le dictionnaire de filtres à partir des paramètres de requête"""
        filters = {}

        site_id = request.query_params.get('site')
        if site_id:
            filters['site_id'] = site_id

        etat = request.query_params.get('etat')
        if etat:
            filters['etat'] = etat

        return filters


# ==============================================================================
# VUE INVENTAIRE UNIFIÉE (15 types combinés)
# ==============================================================================

class InventoryListView(APIView):
    """
    Vue unifiée qui retourne tous les objets (15 types combinés).
    Compatible avec le frontend GreenSIGV1 qui attend un endpoint unique /api/inventory/.

    Cette vue utilise le polymorphisme de la classe Objet pour agréger
    automatiquement tous les types d'objets (Arbre, Gazon, Puit, etc.).

    Endpoint: GET /api/inventory/

    Query params optionnels:
    - type: filtrer par type ('Arbre', 'Gazon', 'Puit', etc.)
    - site: filtrer par site ID ou liste d'IDs séparés par virgule
    - state: filtrer par état (bon, moyen, mauvais, critique) - liste séparée par virgule
    - search: recherche textuelle
    - page: numéro de page (pagination 50 items)
    
    Filtres par plages numériques:
    - surface_min, surface_max: plage de surface en m²
    - diameter_min, diameter_max: plage de diamètre en cm
    - depth_min, depth_max: plage de profondeur en m
    - density_min, density_max: plage de densité
    
    Filtres par date:
    - last_intervention_start, last_intervention_end: plage de dates (YYYY-MM-DD)
    - never_intervened: true/false - objets jamais intervenus
    - urgent_maintenance: true/false - objets nécessitant maintenance urgente (> 6 mois)
    
    Filtres spécifiques:
    - family: famille botanique (liste séparée par virgule)
    - size: taille (Petit, Moyen, Grand) - liste séparée par virgule
    - material: matériau (liste séparée par virgule)
    - equipment_type: type d'équipement (liste séparée par virgule)

    Returns:
        {
            "count": 450,
            "next": "url_page_suivante",
            "previous": null,
            "results": [...]
        }
    """

    def get(self, request):
        """
        Méthode GET optimisée: interroge directement les modèles enfants
        au lieu de passer par Objet, ce qui évite les itérations Python.
        """
        from rest_framework.pagination import PageNumberPagination
        from datetime import timedelta
        from django.utils import timezone
        from itertools import chain

        # Mapping des types vers les modèles et serializers
        MODEL_MAP = {
            'arbre': (Arbre, ArbreSerializer),
            'palmier': (Palmier, PalmierSerializer),
            'gazon': (Gazon, GazonSerializer),
            'arbuste': (Arbuste, ArbusteSerializer),
            'vivace': (Vivace, VivaceSerializer),
            'cactus': (Cactus, CactusSerializer),
            'graminee': (Graminee, GramineeSerializer),
            'puit': (Puit, PuitSerializer),
            'pompe': (Pompe, PompeSerializer),
            'vanne': (Vanne, VanneSerializer),
            'clapet': (Clapet, ClapetSerializer),
            'canalisation': (Canalisation, CanalisationSerializer),
            'aspersion': (Aspersion, AspersionSerializer),
            'goutte': (Goutte, GoutteSerializer),
            'ballon': (Ballon, BallonSerializer),
        }

        # Paramètres de filtrage
        type_filter = request.query_params.get('type', None)
        site_filter = request.query_params.get('site', None)
        etat_filter = request.query_params.get('etat', None)
        famille_filter = request.query_params.get('famille', None)
        search_query = request.query_params.get('search', '').strip()

        # Filtres de date
        never_intervened = request.query_params.get('never_intervened', '').lower() == 'true'
        urgent_maintenance = request.query_params.get('urgent_maintenance', '').lower() == 'true'
        last_intervention_start = request.query_params.get('last_intervention_start', None)

        # Déterminer quels types interroger
        if type_filter:
            target_types = [t.strip().lower() for t in type_filter.split(',')]
            # Normaliser 'graminée' -> 'graminee'
            target_types = ['graminee' if t == 'graminée' else t for t in target_types]
        else:
            target_types = list(MODEL_MAP.keys())

        # Filtrer par rôle (ADMIN, CLIENT, SUPERVISEUR)
        user = request.user
        structure_filter = None
        superviseur_filter = None
        if user.is_authenticated:
            roles = list(user.roles_utilisateur.values_list('role__nom_role', flat=True))

            # ADMIN voit tout - pas de filtre
            if 'ADMIN' in roles:
                pass
            # CLIENT voit uniquement les sites de sa structure
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            # SUPERVISEUR voit uniquement les sites qui lui sont affectés
            elif 'SUPERVISEUR' in roles:
                if hasattr(user, 'superviseur_profile'):
                    superviseur_filter = user.superviseur_profile
                else:
                    # Superviseur sans profil = aucun objet visible
                    return Response({
                        'count': 0,
                        'next': None,
                        'previous': None,
                        'results': []
                    })

        # Collecter les résultats de chaque type
        all_results = []

        for type_name in target_types:
            if type_name not in MODEL_MAP:
                continue

            model_class, serializer_class = MODEL_MAP[type_name]

            # Construire le queryset avec select_related pour éviter les N+1
            qs = model_class.objects.select_related('site', 'sous_site')

            # Appliquer les filtres de rôle
            if structure_filter:
                qs = qs.filter(site__structure_client=structure_filter)
            elif superviseur_filter:
                qs = qs.filter(site__superviseur=superviseur_filter)

            if site_filter:
                qs = qs.filter(site_id=site_filter)

            if etat_filter:
                qs = qs.filter(etat=etat_filter)

            # Filtre famille (pour les types qui ont ce champ)
            if famille_filter and hasattr(model_class, 'famille'):
                qs = qs.filter(famille__icontains=famille_filter)

            # Filtre recherche
            if search_query:
                q = Q(site__nom_site__icontains=search_query) | Q(site__code_site__icontains=search_query)
                if hasattr(model_class, 'nom'):
                    q |= Q(nom__icontains=search_query)
                if hasattr(model_class, 'famille'):
                    q |= Q(famille__icontains=search_query)
                if hasattr(model_class, 'marque'):
                    q |= Q(marque__icontains=search_query)
                if hasattr(model_class, 'type'):
                    q |= Q(type__icontains=search_query)
                qs = qs.filter(q)

            # Filtres de date d'intervention
            if hasattr(model_class, 'last_intervention_date'):
                if never_intervened:
                    qs = qs.filter(last_intervention_date__isnull=True)
                if urgent_maintenance:
                    six_months_ago = timezone.now().date() - timedelta(days=180)
                    qs = qs.filter(last_intervention_date__lt=six_months_ago)
                if last_intervention_start:
                    qs = qs.filter(last_intervention_date__gte=last_intervention_start)

            # Limiter le nombre d'objets récupérés pour éviter les timeout
            qs = qs.order_by('id')

            # Ajouter au résultat avec le type
            for obj in qs:
                all_results.append((obj, type_name.capitalize(), serializer_class))

        # Trier par ID pour un ordre cohérent
        all_results.sort(key=lambda x: x[0].id)

        # Pagination - d'abord sur les objets, puis sérialisation
        paginator = PageNumberPagination()
        paginator.page_size = 50

        page_size_param = request.query_params.get('page_size', None)
        if page_size_param:
            try:
                paginator.page_size = int(page_size_param)
            except ValueError:
                pass

        # Créer une structure paginable (liste d'objets)
        total_count = len(all_results)

        # Calculer la page manuellement pour éviter de tout sérialiser
        page_num = int(request.query_params.get('page', 1))
        start_idx = (page_num - 1) * paginator.page_size
        end_idx = start_idx + paginator.page_size

        page_items = all_results[start_idx:end_idx]

        # Sérialiser uniquement les éléments de la page
        serialized_results = []
        for obj, type_name, serializer_class in page_items:
            serializer = serializer_class(obj)
            data = serializer.data
            if 'properties' in data:
                data['properties']['object_type'] = type_name
            serialized_results.append(data)

        # Construire la réponse paginée
        base_url = request.build_absolute_uri().split('?')[0]
        query_params = request.query_params.copy()

        next_url = None
        if end_idx < total_count:
            query_params['page'] = page_num + 1
            next_url = f"{base_url}?{'&'.join(f'{k}={v}' for k, v in query_params.items())}"

        previous_url = None
        if page_num > 1:
            query_params['page'] = page_num - 1
            previous_url = f"{base_url}?{'&'.join(f'{k}={v}' for k, v in query_params.items())}"

        return Response({
            'count': total_count,
            'next': next_url,
            'previous': previous_url,
            'results': serialized_results
        })

    def apply_id_filter(self, queryset, request):
        """Filtre par ID d'objet (pour récupérer un objet spécifique)"""
        object_id = request.query_params.get('id', None)
        if object_id:
            try:
                queryset = queryset.filter(id=int(object_id))
            except ValueError:
                pass  # Ignore invalid ID
        return queryset

    def apply_type_filter(self, queryset, request):
        """Filtre par type d'objet"""
        type_filter = request.query_params.get('type', None)
        if type_filter:
            types = [t.lower().strip() for t in type_filter.split(',')]
            objets_ids = []
            for obj in queryset:
                for t in types:
                    if t == 'graminée': t = 'graminee'
                    if hasattr(obj, t):
                        objets_ids.append(obj.id)
                        break
            queryset = queryset.filter(id__in=objets_ids)
        return queryset

    def apply_site_filter(self, queryset, request):
        """Filtre par site (supporte liste d'IDs)"""
        site_filter = request.query_params.get('site', None)
        if site_filter:
            site_ids = [int(s.strip()) for s in site_filter.split(',') if s.strip().isdigit()]
            if site_ids:
                queryset = queryset.filter(site_id__in=site_ids)
        return queryset

    def apply_state_filter(self, queryset, request):
        """Filtre par état (supporte liste)"""
        state_filter = request.query_params.get('state', None)
        if state_filter:
            states = [s.strip() for s in state_filter.split(',')]
            queryset = queryset.filter(etat__in=states)
        return queryset

    def apply_search_filter(self, queryset, request):
        """Filtre par recherche textuelle"""
        search_query = request.query_params.get('search', '').strip()
        if search_query:
            q_search = Q()
            q_search |= Q(site__nom_site__icontains=search_query)
            q_search |= Q(site__code_site__icontains=search_query)
            q_search |= Q(sous_site__nom__icontains=search_query)

            for model_name in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee']:
                q_search |= Q(**{f"{model_name}__nom__icontains": search_query})
                q_search |= Q(**{f"{model_name}__famille__icontains": search_query})
            
            for model_name in ['puit', 'pompe']:
                q_search |= Q(**{f"{model_name}__nom__icontains": search_query})

            for model_name in ['vanne', 'clapet', 'canalisation', 'aspersion', 'ballon']:
                q_search |= Q(**{f"{model_name}__marque__icontains": search_query})
            
            q_search |= Q(goutte__type__icontains=search_query)
            queryset = queryset.filter(q_search)
        return queryset

    def apply_range_filters(self, queryset, request):
        """Filtre par plages numériques (surface, diamètre, profondeur, densité)"""
        # Surface (Gazon uniquement)
        surface_min = request.query_params.get('surface_min')
        surface_max = request.query_params.get('surface_max')
        if surface_min or surface_max:
            q = Q()
            if surface_min:
                q &= Q(gazon__area_sqm__gte=float(surface_min))
            if surface_max:
                q &= Q(gazon__area_sqm__lte=float(surface_max))
            queryset = queryset.filter(q)

        # Diamètre (Puit, Pompe, équipements hydrauliques)
        diameter_min = request.query_params.get('diameter_min')
        diameter_max = request.query_params.get('diameter_max')
        if diameter_min or diameter_max:
            q = Q()
            for model in ['puit', 'pompe', 'vanne', 'clapet', 'canalisation', 'aspersion', 'goutte']:
                if diameter_min:
                    q |= Q(**{f"{model}__diametre__gte": float(diameter_min)})
                if diameter_max:
                    q |= Q(**{f"{model}__diametre__lte": float(diameter_max)})
            queryset = queryset.filter(q)

        # Profondeur (Puit uniquement)
        depth_min = request.query_params.get('depth_min')
        depth_max = request.query_params.get('depth_max')
        if depth_min or depth_max:
            q = Q()
            if depth_min:
                q &= Q(puit__profondeur__gte=float(depth_min))
            if depth_max:
                q &= Q(puit__profondeur__lte=float(depth_max))
            queryset = queryset.filter(q)

        # Densité (Arbuste, Vivace, Cactus, Graminee)
        density_min = request.query_params.get('density_min')
        density_max = request.query_params.get('density_max')
        if density_min or density_max:
            q = Q()
            for model in ['arbuste', 'vivace', 'cactus', 'graminee']:
                if density_min:
                    q |= Q(**{f"{model}__densite__gte": float(density_min)})
                if density_max:
                    q |= Q(**{f"{model}__densite__lte": float(density_max)})
            queryset = queryset.filter(q)

        return queryset

    def apply_date_filters(self, queryset, request):
        """Filtre par dates d'intervention"""
        from datetime import timedelta
        from django.utils import timezone

        start_date = request.query_params.get('last_intervention_start')
        end_date = request.query_params.get('last_intervention_end')
        never_intervened = request.query_params.get('never_intervened', '').lower() == 'true'
        urgent = request.query_params.get('urgent_maintenance', '').lower() == 'true'

        # Jamais intervenu
        if never_intervened:
            q = Q()
            for model in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee', 'puit', 'pompe']:
                q |= Q(**{f"{model}__last_intervention_date__isnull": True})
            queryset = queryset.filter(q)

        # Maintenance urgente (> 6 mois)
        if urgent:
            six_months_ago = timezone.now().date() - timedelta(days=180)
            q = Q()
            for model in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee', 'puit', 'pompe']:
                q |= Q(**{f"{model}__last_intervention_date__lt": six_months_ago})
            queryset = queryset.filter(q)

        # Plage de dates personnalisée
        if start_date:
            q = Q()
            for model in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee', 'puit', 'pompe']:
                q |= Q(**{f"{model}__last_intervention_date__gte": start_date})
            queryset = queryset.filter(q)

        if end_date:
            q = Q()
            for model in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee', 'puit', 'pompe']:
                q |= Q(**{f"{model}__last_intervention_date__lte": end_date})
            queryset = queryset.filter(q)

        return queryset

    def apply_specific_filters(self, queryset, request):
        """Filtre par attributs spécifiques (famille, taille, matériau, type)"""
        # Famille (végétaux)
        family_filter = request.query_params.get('family')
        if family_filter:
            families = [f.strip() for f in family_filter.split(',')]
            q = Q()
            for model in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee']:
                q |= Q(**{f"{model}__famille__in": families})
            queryset = queryset.filter(q)

        # Taille (Arbre, Palmier)
        size_filter = request.query_params.get('size')
        if size_filter:
            sizes = [s.strip() for s in size_filter.split(',')]
            q = Q(arbre__taille__in=sizes) | Q(palmier__taille__in=sizes)
            queryset = queryset.filter(q)

        # Matériau (équipements hydrauliques)
        material_filter = request.query_params.get('material')
        if material_filter:
            materials = [m.strip() for m in material_filter.split(',')]
            q = Q()
            for model in ['vanne', 'clapet', 'canalisation', 'aspersion', 'goutte', 'ballon']:
                q |= Q(**{f"{model}__materiau__in": materials})
            queryset = queryset.filter(q)

        # Type d'équipement (Pompe, équipements hydrauliques)
        equipment_type_filter = request.query_params.get('equipment_type')
        if equipment_type_filter:
            types = [t.strip() for t in equipment_type_filter.split(',')]
            q = Q()
            for model in ['pompe', 'vanne', 'clapet', 'canalisation', 'aspersion', 'goutte']:
                q |= Q(**{f"{model}__type__in": types})
            queryset = queryset.filter(q)

        return queryset

    def _get_serializer_class(self, objet):
        """
        Retourne le serializer approprié selon le type de l'objet.

        Args:
            objet: Instance de Arbre, Gazon, Puit, etc.

        Returns:
            Serializer class correspondant au type
        """
        type_mapping = {
            'Arbre': ArbreSerializer,
            'Gazon': GazonSerializer,
            'Palmier': PalmierSerializer,
            'Arbuste': ArbusteSerializer,
            'Vivace': VivaceSerializer,
            'Cactus': CactusSerializer,
            'Graminee': GramineeSerializer,
            'Puit': PuitSerializer,
            'Pompe': PompeSerializer,
            'Vanne': VanneSerializer,
            'Clapet': ClapetSerializer,
            'Canalisation': CanalisationSerializer,
            'Aspersion': AspersionSerializer,
            'Goutte': GoutteSerializer,
            'Ballon': BallonSerializer,
        }
        return type_mapping.get(objet.__class__.__name__, ArbreSerializer)


# ==============================================================================
# ENDPOINT POUR LES OPTIONS DE FILTRAGE
# ==============================================================================

class FilterOptionsView(APIView):
    """
    Endpoint pour récupérer les options disponibles pour les filtres.
    
    Endpoint: GET /api/inventory/filter-options/
    
    Query params optionnels:
    - type: filtrer les options par type d'objet (ex: 'Arbre', 'Puit')
    
    Returns:
        {
            "sites": [{"id": 1, "name": "Jardin Majorelle"}, ...],
            "zones": ["Zone A", "Villa 1", ...],
            "families": ["Palmaceae", "Rosaceae", ...],
            "materials": ["PVC", "Acier", ...],
            "equipment_types": ["Centrifuge", "Immergée", ...],
            "sizes": ["Petit", "Moyen", "Grand"],
            "states": ["bon", "moyen", "mauvais", "critique"],
            "ranges": {
                "surface": [0, 5000],
                "diameter": [0, 200],
                "depth": [0, 150],
                "density": [0, 100]
            }
        }
    """
    
    def get(self, request):
        from django.db.models import Min, Max
        from django.core.cache import cache
        from django.utils.encoding import force_str
        
        object_type = request.query_params.get('type', '').strip()
        
        # Créer une clé de cache basée sur le type d'objet
        cache_key = f'filter_options_{object_type or "all"}'
        
        # Essayer de récupérer depuis le cache (5 minutes)
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Sites (toujours tous) - optimisé avec only()
        sites = Site.objects.filter(actif=True).only('id', 'nom_site').order_by('nom_site')
        sites_list = [{'id': s.id, 'name': s.nom_site} for s in sites]
        
        # Zones (sous-sites)
        zones = SousSite.objects.values_list('nom', flat=True).distinct().order_by('nom')
        zones_list = [z for z in zones if z]
        
        # États (depuis ETAT_CHOICES)
        from .models import ETAT_CHOICES
        states_list = [choice[0] for choice in ETAT_CHOICES]
        
        # Tailles (depuis TAILLE_CHOICES)
        from .models import TAILLE_CHOICES
        sizes_list = [choice[0] for choice in TAILLE_CHOICES]
        
        # Familles botaniques (végétaux)
        families = set()
        if not object_type or object_type.lower() in ['arbre', 'gazon', 'palmier', 'arbuste', 'vivace', 'cactus', 'graminee']:
            for model in [Arbre, Gazon, Palmier, Arbuste, Vivace, Cactus, Graminee]:
                model_families = model.objects.values_list('famille', flat=True).distinct()
                families.update([f for f in model_families if f])
        families_list = sorted(list(families))
        
        # Matériaux (équipements hydrauliques)
        materials = set()
        if not object_type or object_type.lower() in ['vanne', 'clapet', 'canalisation', 'aspersion', 'goutte', 'ballon']:
            for model in [Vanne, Clapet, Canalisation, Aspersion, Goutte, Ballon]:
                if hasattr(model, 'materiau'):
                    model_materials = model.objects.values_list('materiau', flat=True).distinct()
                    materials.update([m for m in model_materials if m])
        materials_list = sorted(list(materials))
        
        # Types d'équipements (hydraulique)
        equipment_types = set()
        if not object_type or object_type.lower() in ['pompe', 'vanne', 'clapet', 'canalisation', 'aspersion', 'goutte']:
            for model in [Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte]:
                if hasattr(model, 'type'):
                    model_types = model.objects.values_list('type', flat=True).distinct()
                    equipment_types.update([t for t in model_types if t])
        equipment_types_list = sorted(list(equipment_types))
        
        # Plages de valeurs
        ranges = {}
        
        # Surface (Gazon)
        if not object_type or object_type.lower() == 'gazon':
            surface_range = qs["Gazon"].aggregate(
                min=Min('area_sqm'),
                max=Max('area_sqm')
            )
            ranges['surface'] = [
                float(surface_range['min'] or 0),
                float(surface_range['max'] or 0)
            ]
        
        # Diamètre (Puit, Pompe, équipements hydrauliques)
        if not object_type or object_type.lower() in ['puit', 'pompe', 'vanne', 'clapet', 'canalisation', 'aspersion', 'goutte']:
            diameter_values = []
            for model in [Puit, Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte]:
                if hasattr(model, 'diametre'):
                    model_range = model.objects.aggregate(
                        min=Min('diametre'),
                        max=Max('diametre')
                    )
                    if model_range['min'] is not None:
                        diameter_values.append(float(model_range['min']))
                    if model_range['max'] is not None:
                        diameter_values.append(float(model_range['max']))
            
            if diameter_values:
                ranges['diameter'] = [min(diameter_values), max(diameter_values)]
            else:
                ranges['diameter'] = [0, 0]
        
        # Profondeur (Puit)
        if not object_type or object_type.lower() == 'puit':
            depth_range = qs["Puit"].aggregate(
                min=Min('profondeur'),
                max=Max('profondeur')
            )
            ranges['depth'] = [
                float(depth_range['min'] or 0),
                float(depth_range['max'] or 0)
            ]
        
        # Densité (Arbuste, Vivace, Cactus, Graminee)
        if not object_type or object_type.lower() in ['arbuste', 'vivace', 'cactus', 'graminee']:
            density_values = []
            for model in [Arbuste, Vivace, Cactus, Graminee]:
                if hasattr(model, 'densite'):
                    model_range = model.objects.aggregate(
                        min=Min('densite'),
                        max=Max('densite')
                    )
                    if model_range['min'] is not None:
                        density_values.append(float(model_range['min']))
                    if model_range['max'] is not None:
                        density_values.append(float(model_range['max']))
            
            if density_values:
                ranges['density'] = [min(density_values), max(density_values)]
            else:
                ranges['density'] = [0, 0]
        
        response_data = {
            'sites': sites_list,
            'zones': zones_list,
            'families': families_list,
            'materials': materials_list,
            'equipment_types': equipment_types_list,
            'sizes': sizes_list,
            'states': states_list,
            'ranges': ranges
        }
        
        # Mettre en cache pour 5 minutes (300 secondes)
        cache.set(cache_key, response_data, 300)
        
        return Response(response_data)


# ==============================================================================
# ENDPOINT UNIFIÉ POUR LA CARTE (avec Bounding Box)
# ==============================================================================

class MapObjectsView(APIView):
    """
    Endpoint unique et intelligent pour charger tous les objets de la carte.

    Permissions automatiques basées sur le rôle:
    - ADMIN: voit tout
    - CLIENT: voit uniquement ses sites et objets
    - SUPERVISEUR: voit uniquement les sites/objets liés aux tâches de ses équipes

    Endpoint: GET /api/map/

    Query params:
    - bbox: Bounding box au format "west,south,east,north" (ex: "-7.95,32.20,-7.90,32.25")
    - types: Liste des types à charger (ex: "sites,arbres,gazons")
    - zoom: Niveau de zoom (pour optimisations futures)

    Returns:
        {
            "type": "FeatureCollection",
            "features": [...],
            "count": 150,
            "bbox_used": true
        }
    """

    def get(self, request):
        from django.contrib.gis.geos import Polygon

        # Paramètres
        bbox_str = request.GET.get('bbox')
        types_str = request.GET.get('types', '')
        zoom = int(request.GET.get('zoom', 10))

        requested_types = [t.strip().lower() for t in types_str.split(',') if t.strip()]

        results = []

        # Déterminer les permissions basées sur le rôle
        user = request.user
        is_admin = False
        structure_filter = None
        superviseur_filter = None  # (site_ids, object_ids)

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            if 'ADMIN' in roles:
                is_admin = True
            elif 'CLIENT' in roles and hasattr(user, 'client_profile'):
                structure_filter = user.client_profile.structure
            elif 'SUPERVISEUR' in roles:
                superviseur_filter = self._get_superviseur_filters(user)

        # ==============================================================================
        # 1. CHARGER LES SITES (toujours tous car peu nombreux)
        # ==============================================================================
        if not requested_types or 'sites' in requested_types:
            sites = Site.objects.filter(actif=True).order_by('id')

            # Appliquer les filtres de permissions
            if not is_admin:
                if structure_filter:
                    sites = sites.filter(structure_client=structure_filter)
                elif superviseur_filter:
                    site_ids, _ = superviseur_filter
                    if site_ids:
                        sites = sites.filter(id__in=site_ids)
                    else:
                        sites = sites.none()

            for site in sites:
                # Utiliser le centroid pré-calculé (ou calculer depuis geometrie_emprise)
                centroid = site.centroid or site.geometrie_emprise.centroid

                serializer = SiteSerializer(site)
                feature = serializer.data

                # Ajouter des métadonnées pour le frontend
                feature['properties']['object_type'] = 'Site'
                feature['properties']['center'] = {
                    'lat': centroid.y,
                    'lng': centroid.x
                }

                results.append(feature)

        # ==============================================================================
        # 2. CHARGER VÉGÉTATION / HYDRAULIQUE (avec bbox si fourni)
        # ==============================================================================
        if bbox_str:
            try:
                west, south, east, north = map(float, bbox_str.split(','))
                bbox_polygon = Polygon.from_bbox((west, south, east, north))

                # Mapping type -> (model, serializer)
                type_mapping = {
                    'arbres': (Arbre, ArbreSerializer),
                    'gazons': (Gazon, GazonSerializer),
                    'palmiers': (Palmier, PalmierSerializer),
                    'arbustes': (Arbuste, ArbusteSerializer),
                    'vivaces': (Vivace, VivaceSerializer),
                    'cactus': (Cactus, CactusSerializer),
                    'graminees': (Graminee, GramineeSerializer),
                    'puits': (Puit, PuitSerializer),
                    'pompes': (Pompe, PompeSerializer),
                    'vannes': (Vanne, VanneSerializer),
                    'clapets': (Clapet, ClapetSerializer),
                    'canalisations': (Canalisation, CanalisationSerializer),
                    'aspersions': (Aspersion, AspersionSerializer),
                    'gouttes': (Goutte, GoutteSerializer),
                    'ballons': (Ballon, BallonSerializer),
                }

                # Déterminer quels types charger
                types_to_load = requested_types if requested_types else list(type_mapping.keys())

                # Types polygones qui nécessitent le calcul de superficie
                polygon_types = {'gazons', 'arbustes', 'vivaces', 'cactus', 'graminees'}

                # Charger chaque type avec filtrage bbox
                for type_name in types_to_load:
                    if type_name in type_mapping:
                        Model, Serializer = type_mapping[type_name]

                        # Query avec bbox filter
                        queryset = Model.objects.filter(
                            geometry__intersects=bbox_polygon
                        ).select_related('site', 'sous_site')

                        # OPTIMISÉ: Pré-calculer la superficie pour les types polygones
                        # Évite N+1 requêtes ST_Area dans le serializer
                        if type_name in polygon_types:
                            from django.db.models.expressions import RawSQL
                            queryset = queryset.annotate(
                                _superficie_annotee=RawSQL(
                                    "ST_Area(geometry::geography)",
                                    []
                                )
                            )

                        # Appliquer les filtres de permissions (sauf pour ADMIN)
                        if not is_admin:
                            if structure_filter:
                                queryset = queryset.filter(site__structure_client=structure_filter)
                            elif superviseur_filter:
                                _, object_ids = superviseur_filter
                                # SUPERVISEUR: ne voir QUE les objets directement liés aux tâches
                                if object_ids:
                                    queryset = queryset.filter(objet_ptr_id__in=object_ids)
                                else:
                                    queryset = queryset.none()

                        queryset = queryset.order_by('id')[:100]  # 100 par type max

                        # Serializer chaque objet
                        for obj in queryset:
                            serializer = Serializer(obj)
                            feature = serializer.data
                            feature['properties']['object_type'] = Model.__name__
                            results.append(feature)

            except (ValueError, AttributeError) as e:
                return Response({
                    'error': f'Invalid bbox format: {str(e)}'
                }, status=400)

        return Response({
            'type': 'FeatureCollection',
            'features': results,
            'count': len(results),
            'bbox_used': bbox_str is not None,
            'zoom': zoom
        })

    def _get_superviseur_filters(self, user):
        """
        Récupère les IDs des sites et objets affectés directement au superviseur.
        Returns: tuple (site_ids, object_ids)

        Le superviseur voit:
        - Les sites qui lui sont affectés directement (via site.superviseur)
        - Tous les objets de ces sites

        NOUVEAU SYSTÈME: Affectation directe superviseur → sites (plus simple et plus clair)
        """
        try:
            superviseur = user.superviseur_profile

            # Sites affectés directement au superviseur
            site_ids = list(
                Site.objects.filter(superviseur=superviseur).values_list('id', flat=True)
            )

            if not site_ids:
                return ([], [])

            # Tous les objets GIS de ces sites
            object_ids = list(
                Objet.objects.filter(site_id__in=site_ids).values_list('id', flat=True)
            )

            return (site_ids, object_ids)
        except Exception:
            return ([], [])


# ==============================================================================
# OPTIONS DE FILTRAGE POUR L'INVENTAIRE
# ==============================================================================

class InventoryFilterOptionsView(APIView):
    """
    Retourne les options de filtrage disponibles pour l'inventaire.

    Endpoint: GET /api/inventory/filter-options/

    Query params optionnels:
    - type: filtrer les options selon un type d'objet spécifique

    Returns:
        {
            "sites": [{"id": 1, "name": "Site A"}, ...],
            "zones": ["Zone 1", "Zone 2", ...],
            "families": ["Palmaceae", "Rosaceae", ...],
            "materials": ["PVC", "Acier", ...],
            "equipment_types": ["Centrifuge", "Submersible", ...],
            "sizes": ["Petit", "Moyen", "Grand"],
            "states": ["bon", "moyen", "mauvais", "critique"],
            "ranges": {
                "surface": [0, 10000],
                "diameter": [0, 500],
                "depth": [0, 100],
                "density": [0, 100]
            }
        }
    """

    def get(self, request):
        from django.db.models import Min, Max, Q

        type_filter = request.query_params.get('type', None)

        # Obtenir les querysets filtrés selon les permissions de l'utilisateur
        user = request.user
        site_filter = Q()  # Par défaut, pas de filtre (ADMIN)

        if user.is_authenticated:
            roles = [ur.role.nom_role for ur in user.roles_utilisateur.all()]

            if 'ADMIN' not in roles:
                if 'CLIENT' in roles and hasattr(user, 'client_profile'):
                    # CLIENT: uniquement les sites de sa structure
                    structure = user.client_profile.structure
                    if structure:
                        site_filter = Q(structure_client=structure)
                        object_filter = Q(site__structure_client=structure)
                    else:
                        site_filter = Q(pk__in=[])
                        object_filter = Q(pk__in=[])
                elif 'SUPERVISEUR' in roles and hasattr(user, 'superviseur_profile'):
                    # SUPERVISEUR: uniquement les sites qui lui sont affectés
                    site_filter = Q(superviseur=user.superviseur_profile)
                    object_filter = Q(site__superviseur=user.superviseur_profile)
                else:
                    # Aucun accès
                    site_filter = Q(pk__in=[])
                    object_filter = Q(pk__in=[])
            else:
                # ADMIN: pas de filtre
                object_filter = Q()
        else:
            # Non authentifié: aucun accès
            site_filter = Q(pk__in=[])
            object_filter = Q(pk__in=[])

        # ==============================================================================
        # SITES
        # ==============================================================================
        sites = Site.objects.filter(site_filter).filter(actif=True).values('id', 'nom_site').order_by('nom_site')
        sites_list = [{'id': s['id'], 'name': s['nom_site']} for s in sites]

        # ==============================================================================
        # ZONES (Sous-sites)
        # ==============================================================================
        zones = list(SousSite.objects.filter(object_filter).values_list('nom', flat=True).distinct().order_by('nom'))

        # ==============================================================================
        # FAMILLES (végétaux uniquement)
        # ==============================================================================
        families = set()
        if not type_filter or type_filter.lower() in ['arbre', 'arbres']:
            families.update(Arbre.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['gazon', 'gazons']:
            families.update(Gazon.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['palmier', 'palmiers']:
            families.update(Palmier.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['arbuste', 'arbustes']:
            families.update(Arbuste.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['vivace', 'vivaces']:
            families.update(Vivace.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['cactus']:
            families.update(Cactus.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['graminee', 'graminees']:
            families.update(Graminee.objects.filter(object_filter).exclude(famille__isnull=True).exclude(famille='').values_list('famille', flat=True).distinct())

        families_list = sorted(list(families))

        # ==============================================================================
        # MATÉRIAUX (hydraulique uniquement)
        # ==============================================================================
        materials = set()
        if not type_filter or type_filter.lower() in ['vanne', 'vannes']:
            materials.update(Vanne.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['clapet', 'clapets']:
            materials.update(Clapet.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['canalisation', 'canalisations']:
            materials.update(Canalisation.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['aspersion', 'aspersions']:
            materials.update(Aspersion.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['goutte', 'gouttes']:
            materials.update(Goutte.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['ballon', 'ballons']:
            materials.update(Ballon.objects.filter(object_filter).exclude(materiau__isnull=True).exclude(materiau='').values_list('materiau', flat=True).distinct())

        materials_list = sorted(list(materials))

        # ==============================================================================
        # TYPES D'ÉQUIPEMENT
        # ==============================================================================
        equipment_types = set()
        if not type_filter or type_filter.lower() in ['pompe', 'pompes']:
            equipment_types.update(Pompe.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['vanne', 'vannes']:
            equipment_types.update(Vanne.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['clapet', 'clapets']:
            equipment_types.update(Clapet.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['canalisation', 'canalisations']:
            equipment_types.update(Canalisation.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['aspersion', 'aspersions']:
            equipment_types.update(Aspersion.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())
        if not type_filter or type_filter.lower() in ['goutte', 'gouttes']:
            equipment_types.update(Goutte.objects.filter(object_filter).exclude(type__isnull=True).exclude(type='').values_list('type', flat=True).distinct())

        equipment_types_list = sorted(list(equipment_types))

        # ==============================================================================
        # TAILLES (statiques basées sur TAILLE_CHOICES)
        # ==============================================================================
        sizes = ['Petit', 'Moyen', 'Grand']

        # ==============================================================================
        # ÉTATS (statiques - à implémenter dans le modèle plus tard)
        # ==============================================================================
        states = ['bon', 'moyen', 'mauvais', 'critique']

        # ==============================================================================
        # PLAGES DE VALEURS
        # ==============================================================================
        ranges = {}

        # Surface (gazons)
        surface_range = Gazon.objects.filter(object_filter).aggregate(min_val=Min('area_sqm'), max_val=Max('area_sqm'))
        if surface_range['min_val'] is not None:
            ranges['surface'] = [
                float(surface_range['min_val'] or 0),
                float(surface_range['max_val'] or 10000)
            ]

        # Diamètre (puits, pompes, vannes, etc.)
        diameter_values = []
        for Model in [Puit, Pompe, Vanne, Clapet, Canalisation, Aspersion, Goutte]:
            agg = Model.objects.filter(object_filter).aggregate(min_val=Min('diametre'), max_val=Max('diametre'))
            if agg['min_val'] is not None:
                diameter_values.append(agg['min_val'])
            if agg['max_val'] is not None:
                diameter_values.append(agg['max_val'])
        if diameter_values:
            ranges['diameter'] = [float(min(diameter_values)), float(max(diameter_values))]

        # Profondeur (puits)
        depth_range = Puit.objects.filter(object_filter).aggregate(min_val=Min('profondeur'), max_val=Max('profondeur'))
        if depth_range['min_val'] is not None:
            ranges['depth'] = [
                float(depth_range['min_val'] or 0),
                float(depth_range['max_val'] or 100)
            ]

        # Densité (arbustes, vivaces, cactus, graminées)
        density_values = []
        for Model in [Arbuste, Vivace, Cactus, Graminee]:
            agg = Model.objects.filter(object_filter).aggregate(min_val=Min('densite'), max_val=Max('densite'))
            if agg['min_val'] is not None:
                density_values.append(agg['min_val'])
            if agg['max_val'] is not None:
                density_values.append(agg['max_val'])
        if density_values:
            ranges['density'] = [float(min(density_values)), float(max(density_values))]

        return Response({
            'sites': sites_list,
            'zones': zones,
            'families': families_list,
            'materials': materials_list,
            'equipment_types': equipment_types_list,
            'sizes': sizes,
            'states': states,
            'ranges': ranges
        })


# ==============================================================================
# IMPORT/EXPORT GEO SERVICES
# ==============================================================================

from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from .services.geo_io import (
    parse_geojson, parse_kml, parse_shapefile,
    convert_geometry, validate_geometry_for_type,
    export_to_geojson, export_to_kml, export_to_shapefile,
    apply_attribute_mapping, suggest_attribute_mapping,
    GEOMETRY_TYPE_MAPPING, OBJECT_FIELDS
)


class GeoImportPreviewView(APIView):
    """
    Preview imported geo data before validation.

    POST /api/import/preview/
    Content-Type: multipart/form-data

    Body:
        - file: The geo file (GeoJSON, KML, KMZ, or ZIP with Shapefile)
        - format: 'geojson' | 'kml' | 'shapefile' (auto-detected if not provided)

    Returns:
        {
            "features": [
                {
                    "index": 0,
                    "geometry": {...},
                    "geometry_type": "Point",
                    "properties": {...},
                    "compatible_types": ["Arbre", "Palmier", "Puit", ...]
                },
                ...
            ],
            "detected_types": {"Point": 10, "Polygon": 5},
            "attributes": ["name", "description", "type", ...],
            "suggested_mapping": {...},
            "errors": []
        }
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=400)

        file_format = request.data.get('format', '').lower()

        # Auto-detect format from filename
        filename = file_obj.name.lower()
        if not file_format or file_format == 'auto':
            if filename.endswith('.geojson') or filename.endswith('.json'):
                file_format = 'geojson'
            elif filename.endswith('.kml') or filename.endswith('.kmz'):
                file_format = 'kml'
            elif filename.endswith('.zip'):
                file_format = 'shapefile'
            else:
                return Response({'error': 'Could not detect file format. Please specify format parameter.'}, status=400)

        # Read file content
        file_content = file_obj.read()

        # Parse based on format
        if file_format == 'geojson':
            result = parse_geojson(file_content)
        elif file_format == 'kml':
            result = parse_kml(file_content)
        elif file_format == 'shapefile':
            result = parse_shapefile(file_content)
        else:
            return Response({'error': f'Unsupported format: {file_format}'}, status=400)

        # Transform response to match frontend expectations
        detected_types = result.get('detected_types', {})
        attributes = result.get('attributes', [])
        features = result.get('features', [])

        # Add suggested mapping if we have features
        suggested_mapping = {}
        if features and attributes:
            if detected_types:
                most_common_geom = max(detected_types.keys(), key=lambda k: detected_types[k])
                compatible = features[0].get('compatible_types', [])
                if compatible:
                    suggested_mapping = suggest_attribute_mapping(
                        attributes,
                        compatible[0]
                    )

        # Build response matching frontend ImportPreviewResponse interface
        response_data = {
            'format': file_format,
            'feature_count': len(features),
            'geometry_types': list(detected_types.keys()),
            'sample_properties': attributes,
            'features': features,
            'suggested_mapping': suggested_mapping,
            'errors': result.get('errors', []),
        }

        return Response(response_data)


class GeoImportValidateView(APIView):
    """
    Validate imported features before execution.

    POST /api/import/validate/

    Body:
        {
            "features": [...],  // From preview response
            "mapping": {        // Attribute mapping
                "source_attr": "target_field",
                ...
            },
            "target_type": "Arbre",  // Target object type
            "site_id": 1             // Target site ID
        }

    Returns:
        {
            "valid": true|false,
            "valid_count": 10,
            "errors": [
                {"index": 0, "error": "..."},
                ...
            ],
            "warnings": [
                {"index": 1, "warning": "Duplicate detected at ..."},
                ...
            ]
        }
    """

    def post(self, request):
        features = request.data.get('features', [])
        mapping = request.data.get('mapping', {})
        target_type = request.data.get('target_type')
        site_id = request.data.get('site_id')
        auto_detect_site = request.data.get('auto_detect_site', False)

        if not features:
            return Response({'error': 'No features provided'}, status=400)

        if not target_type:
            return Response({'error': 'target_type is required'}, status=400)

        if target_type not in GEOMETRY_TYPE_MAPPING:
            return Response({'error': f'Invalid target_type: {target_type}'}, status=400)

        # Site is not required for Site objects
        site = None
        all_sites = None  # For auto-detect mode
        if target_type != 'Site':
            if auto_detect_site:
                # Load all active sites for geometry-based detection
                all_sites = list(Site.objects.filter(actif=True))
                if not all_sites:
                    return Response({'error': 'No active sites found for auto-detection'}, status=400)
            elif not site_id:
                return Response({'error': 'site_id is required (or enable auto_detect_site)'}, status=400)
            else:
                # Verify site exists
                try:
                    site = Site.objects.get(pk=site_id)
                except Site.DoesNotExist:
                    return Response({'error': f'Site {site_id} not found'}, status=400)

        errors = []
        warnings = []
        valid_count = 0
        invalid_count = 0
        validated_features = []

        expected_geom_type = GEOMETRY_TYPE_MAPPING[target_type]

        for feature in features:
            idx = feature.get('index', 0)
            geometry = feature.get('geometry')
            properties = feature.get('properties', {})
            is_valid = True

            if not geometry:
                errors.append({'index': idx, 'message': 'Missing geometry', 'code': 'MISSING_GEOMETRY'})
                is_valid = False
                invalid_count += 1
                validated_features.append({
                    'index': idx,
                    'is_valid': False,
                    'geometry_type': 'Unknown',
                    'mapped_properties': {}
                })
                continue

            geom_type = geometry.get('type', 'Unknown')

            # Validate geometry type
            geom_valid, error_msg = validate_geometry_for_type(
                GEOSGeometry(json.dumps(geometry), srid=4326),
                target_type
            )

            if not geom_valid:
                errors.append({'index': idx, 'message': error_msg, 'code': 'INVALID_GEOMETRY'})
                is_valid = False
                invalid_count += 1
                validated_features.append({
                    'index': idx,
                    'is_valid': False,
                    'geometry_type': geom_type,
                    'mapped_properties': {}
                })
                continue

            # Check if geometry is within site boundary (only for non-Site objects)
            try:
                geom = convert_geometry(geometry, expected_geom_type)
                detected_site = site  # Use provided site by default

                if auto_detect_site and all_sites:
                    # Auto-detect: find the site that contains this geometry
                    detected_site = None
                    for candidate_site in all_sites:
                        if candidate_site.geometrie_emprise:
                            if candidate_site.geometrie_emprise.contains(geom) or candidate_site.geometrie_emprise.intersects(geom):
                                detected_site = candidate_site
                                break

                    if not detected_site:
                        errors.append({
                            'index': idx,
                            'message': 'Geometry is not within any site boundary',
                            'code': 'NO_SITE_FOUND'
                        })
                        is_valid = False
                        invalid_count += 1
                        validated_features.append({
                            'index': idx,
                            'is_valid': False,
                            'geometry_type': geom_type,
                            'mapped_properties': {},
                            'detected_site_id': None
                        })
                        continue

                elif detected_site and detected_site.geometrie_emprise:
                    # Manual mode: check if geometry is within selected site
                    if not detected_site.geometrie_emprise.contains(geom) and not detected_site.geometrie_emprise.intersects(geom):
                        warnings.append({
                            'index': idx,
                            'message': 'Geometry is outside site boundary',
                            'code': 'OUTSIDE_BOUNDARY'
                        })
            except Exception as e:
                errors.append({'index': idx, 'message': f'Geometry conversion error: {str(e)}', 'code': 'CONVERSION_ERROR'})
                is_valid = False
                invalid_count += 1
                validated_features.append({
                    'index': idx,
                    'is_valid': False,
                    'geometry_type': geom_type,
                    'mapped_properties': {}
                })
                continue

            # Check for duplicates (same location within tolerance)
            try:
                from django.contrib.gis.db.models.functions import Distance
                from django.contrib.gis.measure import D

                # Get model class
                model_class = self._get_model_class(target_type)
                if model_class:
                    if target_type == 'Site':
                        # For Sites, check by geometry overlap
                        nearby = model_class.objects.filter(
                            geometrie_emprise__intersects=geom
                        ).exists()
                    else:
                        # Use detected_site for duplicate check
                        check_site = detected_site if auto_detect_site else site
                        if check_site:
                            nearby = model_class.objects.filter(
                                site=check_site,
                                geometry__distance_lte=(geom, D(m=1))
                            ).exists()
                        else:
                            nearby = False
                    if nearby:
                        warnings.append({
                            'index': idx,
                            'message': 'Possible duplicate: object exists within 1m' if target_type != 'Site' else 'Possible duplicate: site with overlapping geometry exists',
                            'code': 'DUPLICATE'
                        })
            except Exception:
                pass  # Skip duplicate check on error

            # Apply mapping to get mapped properties preview
            mapped_props = apply_attribute_mapping(feature, mapping, target_type)

            valid_count += 1
            feature_result = {
                'index': idx,
                'is_valid': True,
                'geometry_type': geom_type,
                'mapped_properties': mapped_props
            }

            # Include detected site info if in auto-detect mode
            if auto_detect_site and detected_site:
                feature_result['detected_site_id'] = detected_site.pk
                feature_result['detected_site_name'] = detected_site.nom_site

            validated_features.append(feature_result)

        return Response({
            'valid_count': valid_count,
            'invalid_count': invalid_count,
            'errors': errors,
            'warnings': warnings,
            'features': validated_features
        })

    def _get_model_class(self, target_type):
        """Get Django model class from type name."""
        model_mapping = {
            'Site': Site,
            'Arbre': Arbre,
            'Gazon': Gazon,
            'Palmier': Palmier,
            'Arbuste': Arbuste,
            'Vivace': Vivace,
            'Cactus': Cactus,
            'Graminee': Graminee,
            'Puit': Puit,
            'Pompe': Pompe,
            'Vanne': Vanne,
            'Clapet': Clapet,
            'Ballon': Ballon,
            'Canalisation': Canalisation,
            'Aspersion': Aspersion,
            'Goutte': Goutte,
        }
        return model_mapping.get(target_type)


class GeoImportExecuteView(APIView):
    """
    Execute the import and create objects in database.

    POST /api/import/execute/

    Body:
        {
            "features": [...],
            "mapping": {...},
            "target_type": "Arbre",
            "site_id": 1,
            "sous_site_id": null  // Optional
        }

    Returns:
        {
            "created": [1, 2, 3, ...],  // IDs of created objects
            "errors": [
                {"index": 0, "error": "..."},
                ...
            ],
            "summary": {
                "total": 10,
                "created": 8,
                "failed": 2
            }
        }
    """

    def post(self, request):
        features = request.data.get('features', [])
        mapping = request.data.get('mapping', {})
        target_type = request.data.get('target_type')
        site_id = request.data.get('site_id')
        sous_site_id = request.data.get('sous_site_id')
        auto_detect_site = request.data.get('auto_detect_site', False)

        if not features:
            return Response({'error': 'No features provided'}, status=400)

        if not target_type or target_type not in GEOMETRY_TYPE_MAPPING:
            return Response({'error': 'Invalid target_type'}, status=400)

        # Site is not required for Site objects
        site = None
        sous_site = None
        all_sites = None  # For auto-detect mode
        if target_type != 'Site':
            if auto_detect_site:
                # Load all active sites for geometry-based detection
                all_sites = list(Site.objects.filter(actif=True))
                if not all_sites:
                    return Response({'error': 'No active sites found for auto-detection'}, status=400)
            elif not site_id:
                return Response({'error': 'site_id is required (or enable auto_detect_site)'}, status=400)
            else:
                # Get site and optionally sous_site
                try:
                    site = Site.objects.get(pk=site_id)
                except Site.DoesNotExist:
                    return Response({'error': f'Site {site_id} not found'}, status=400)

            if sous_site_id:
                try:
                    sous_site = SousSite.objects.get(pk=sous_site_id)
                except SousSite.DoesNotExist:
                    return Response({'error': f'SousSite {sous_site_id} not found'}, status=400)

        # Get model class
        model_class = self._get_model_class(target_type)
        if not model_class:
            return Response({'error': f'Unknown model for type: {target_type}'}, status=400)

        expected_geom_type = GEOMETRY_TYPE_MAPPING[target_type]

        created_ids = []
        errors = []

        for feature in features:
            idx = feature.get('index', 0)

            try:
                geometry = feature.get('geometry')
                if not geometry:
                    errors.append({'index': idx, 'error': 'Missing geometry'})
                    continue

                # Convert geometry
                geom = convert_geometry(geometry, expected_geom_type)

                # Apply attribute mapping
                attributes = apply_attribute_mapping(feature, mapping, target_type)

                # Handle Site objects differently
                if target_type == 'Site':
                    # Sites have different field names
                    attributes['geometrie_emprise'] = geom
                    attributes['centroid'] = geom.centroid
                    attributes['actif'] = True
                    # Set default name if not provided
                    if 'nom_site' not in attributes:
                        props = feature.get('properties', {})
                        attributes['nom_site'] = props.get('name') or props.get('nom') or f"Site Import {idx + 1}"
                    # Set default code if not provided
                    if 'code_site' not in attributes:
                        import uuid
                        attributes['code_site'] = f"SITE_{uuid.uuid4().hex[:8].upper()}"
                else:
                    # Add required fields for other objects
                    # Determine the site for this feature
                    target_site = site  # Use provided site by default

                    if auto_detect_site and all_sites:
                        # Auto-detect: find the site that contains this geometry
                        target_site = None
                        for candidate_site in all_sites:
                            if candidate_site.geometrie_emprise:
                                if candidate_site.geometrie_emprise.contains(geom) or candidate_site.geometrie_emprise.intersects(geom):
                                    target_site = candidate_site
                                    break

                        if not target_site:
                            errors.append({'index': idx, 'error': 'Geometry is not within any site boundary'})
                            continue

                    attributes['site'] = target_site
                    if sous_site:
                        attributes['sous_site'] = sous_site
                    attributes['geometry'] = geom

                    # Set default name if not provided
                    if 'nom' not in attributes and 'nom' in OBJECT_FIELDS.get(target_type, []):
                        attributes['nom'] = f"{target_type} Import {idx + 1}"

                # Create object
                obj = model_class.objects.create(**attributes)
                created_ids.append(obj.pk)

            except Exception as e:
                errors.append({'index': idx, 'error': str(e)})

        return Response({
            'created': created_ids,
            'errors': errors,
            'summary': {
                'total': len(features),
                'created': len(created_ids),
                'failed': len(errors)
            }
        })

    def _get_model_class(self, target_type):
        """Get Django model class from type name."""
        model_mapping = {
            'Site': Site,
            'Arbre': Arbre,
            'Gazon': Gazon,
            'Palmier': Palmier,
            'Arbuste': Arbuste,
            'Vivace': Vivace,
            'Cactus': Cactus,
            'Graminee': Graminee,
            'Puit': Puit,
            'Pompe': Pompe,
            'Vanne': Vanne,
            'Clapet': Clapet,
            'Ballon': Ballon,
            'Canalisation': Canalisation,
            'Aspersion': Aspersion,
            'Goutte': Goutte,
        }
        return model_mapping.get(target_type)


# ==============================================================================
# VUES POUR LES OPÉRATIONS GÉOMÉTRIQUES
# ==============================================================================

from .services.validation import (
    validate_geometry,
    detect_duplicates,
    check_within_site,
    simplify_geometry,
    split_polygon,
    merge_polygons,
    calculate_geometry_metrics,
    buffer_geometry,
)


class GeometrySimplifyView(APIView):
    """
    POST /api/geometry/simplify/
    Simplifie une géométrie en réduisant le nombre de sommets.

    Request body:
    {
        "geometry": { GeoJSON geometry },
        "tolerance": 0.0001,  // Optional, default 0.0001 degrees (~11m)
        "preserve_topology": true  // Optional, default true
    }

    Response:
    {
        "geometry": { simplified GeoJSON geometry },
        "stats": {
            "original_coords": 150,
            "simplified_coords": 45,
            "reduction_percent": 70.0,
            ...
        }
    }
    """

    def post(self, request):
        geometry_data = request.data.get('geometry')
        tolerance = request.data.get('tolerance', 0.0001)
        preserve_topology = request.data.get('preserve_topology', True)

        if not geometry_data:
            return Response({'error': 'geometry is required'}, status=400)

        try:
            # Convert GeoJSON to GEOS geometry
            geom = GEOSGeometry(json.dumps(geometry_data))

            # Simplify
            simplified, stats = simplify_geometry(
                geom,
                tolerance=tolerance,
                preserve_topology=preserve_topology
            )

            # Convert back to GeoJSON
            simplified_geojson = json.loads(simplified.geojson)

            return Response({
                'geometry': simplified_geojson,
                'stats': stats
            })

        except Exception as e:
            return Response({'error': str(e)}, status=400)


class GeometrySplitView(APIView):
    """
    POST /api/geometry/split/
    Divise un polygone avec une ligne de coupe.

    Request body:
    {
        "polygon": { GeoJSON Polygon },
        "split_line": { GeoJSON LineString }
    }

    Response:
    {
        "geometries": [ { GeoJSON Polygon }, ... ],
        "stats": {
            "success": true,
            "num_parts": 2,
            "areas": [0.001, 0.002]
        }
    }
    """

    def post(self, request):
        polygon_data = request.data.get('polygon')
        split_line_data = request.data.get('split_line')

        if not polygon_data:
            return Response({'error': 'polygon is required'}, status=400)
        if not split_line_data:
            return Response({'error': 'split_line is required'}, status=400)

        try:
            polygon = GEOSGeometry(json.dumps(polygon_data))
            split_line = GEOSGeometry(json.dumps(split_line_data))

            if polygon.geom_type not in ('Polygon', 'MultiPolygon'):
                return Response({'error': 'First geometry must be a Polygon'}, status=400)
            if split_line.geom_type != 'LineString':
                return Response({'error': 'Split line must be a LineString'}, status=400)

            result_polygons, stats = split_polygon(polygon, split_line)

            # Convert to GeoJSON
            geometries = [json.loads(p.geojson) for p in result_polygons]

            return Response({
                'geometries': geometries,
                'stats': stats
            })

        except Exception as e:
            return Response({'error': str(e)}, status=400)


class GeometryMergeView(APIView):
    """
    POST /api/geometry/merge/
    Fusionne plusieurs polygones en un seul.

    Request body:
    {
        "polygons": [ { GeoJSON Polygon }, { GeoJSON Polygon }, ... ]
    }

    Response:
    {
        "geometry": { GeoJSON Polygon or MultiPolygon },
        "stats": {
            "success": true,
            "input_count": 3,
            "output_type": "Polygon",
            ...
        }
    }
    """

    def post(self, request):
        polygons_data = request.data.get('polygons', [])

        if not polygons_data or len(polygons_data) < 2:
            return Response({'error': 'At least 2 polygons are required'}, status=400)

        try:
            polygons = []
            for i, poly_data in enumerate(polygons_data):
                poly = GEOSGeometry(json.dumps(poly_data))
                if poly.geom_type not in ('Polygon', 'MultiPolygon'):
                    return Response({
                        'error': f'Geometry at index {i} must be a Polygon'
                    }, status=400)
                polygons.append(poly)

            result, stats = merge_polygons(polygons)

            if result is None:
                return Response({
                    'geometry': None,
                    'stats': stats
                }, status=400)

            return Response({
                'geometry': json.loads(result.geojson),
                'stats': stats
            })

        except Exception as e:
            return Response({'error': str(e)}, status=400)


class GeometryValidateView(APIView):
    """
    POST /api/geometry/validate/
    Valide une géométrie et détecte les doublons potentiels.

    Request body:
    {
        "geometry": { GeoJSON geometry },
        "target_type": "Arbre",  // Optional, for duplicate detection
        "site_id": 1,  // Optional, for duplicate detection and boundary check
        "check_duplicates": true,  // Optional
        "check_within_site": true,  // Optional
        "duplicate_tolerance": 0.0001  // Optional, degrees
    }

    Response:
    {
        "validation": {
            "is_valid": true,
            "errors": [],
            "warnings": []
        },
        "duplicates": [...],  // If check_duplicates=true
        "within_site": {...}  // If check_within_site=true
    }
    """

    def post(self, request):
        geometry_data = request.data.get('geometry')
        target_type = request.data.get('target_type')
        site_id = request.data.get('site_id')
        check_dup = request.data.get('check_duplicates', False)
        check_site = request.data.get('check_within_site', False)
        dup_tolerance = request.data.get('duplicate_tolerance', 0.0001)

        if not geometry_data:
            return Response({'error': 'geometry is required'}, status=400)

        try:
            geom = GEOSGeometry(json.dumps(geometry_data))

            response_data = {}

            # Basic validation
            response_data['validation'] = validate_geometry(geom)

            # Duplicate detection
            if check_dup and target_type:
                model_class = self._get_model_class(target_type)
                if model_class:
                    response_data['duplicates'] = detect_duplicates(
                        geom,
                        model_class,
                        site_id=site_id,
                        tolerance=dup_tolerance
                    )
                else:
                    response_data['duplicates'] = []
                    response_data['validation']['warnings'].append({
                        'code': 'UNKNOWN_TYPE',
                        'message': f'Unknown target type: {target_type}'
                    })

            # Within site check
            if check_site and site_id:
                response_data['within_site'] = check_within_site(geom, site_id)

            return Response(response_data)

        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def _get_model_class(self, target_type):
        """Get Django model class from type name."""
        model_mapping = {
            'Site': Site,
            'Arbre': Arbre,
            'Gazon': Gazon,
            'Palmier': Palmier,
            'Arbuste': Arbuste,
            'Vivace': Vivace,
            'Cactus': Cactus,
            'Graminee': Graminee,
            'Puit': Puit,
            'Pompe': Pompe,
            'Vanne': Vanne,
            'Clapet': Clapet,
            'Ballon': Ballon,
            'Canalisation': Canalisation,
            'Aspersion': Aspersion,
            'Goutte': Goutte,
        }
        return model_mapping.get(target_type)


class GeometryCalculateView(APIView):
    """
    POST /api/geometry/calculate/
    Calcule les métriques d'une géométrie (aire, longueur, périmètre, etc.).

    Request body:
    {
        "geometry": { GeoJSON geometry }
    }

    Response:
    {
        "metrics": {
            "geometry_type": "Polygon",
            "area_m2": 1234.56,
            "area_hectares": 0.1234,
            "perimeter_m": 456.78,
            "centroid": { "lng": -7.5, "lat": 33.5 },
            "bbox": { "min_lng": ..., ... }
        }
    }
    """

    def post(self, request):
        geometry_data = request.data.get('geometry')

        if not geometry_data:
            return Response({'error': 'geometry is required'}, status=400)

        try:
            geom = GEOSGeometry(json.dumps(geometry_data))
            metrics = calculate_geometry_metrics(geom)

            return Response({'metrics': metrics})

        except Exception as e:
            return Response({'error': str(e)}, status=400)


class GeometryBufferView(APIView):
    """
    POST /api/geometry/buffer/
    Crée un buffer (zone tampon) autour d'une géométrie.

    Request body:
    {
        "geometry": { GeoJSON geometry },
        "distance": 10,  // Distance in meters
        "quad_segs": 8  // Optional, segments for curves (default 8)
    }

    Response:
    {
        "geometry": { GeoJSON Polygon },
        "stats": {
            "input_type": "Point",
            "output_type": "Polygon",
            "distance_meters": 10
        }
    }
    """

    def post(self, request):
        geometry_data = request.data.get('geometry')
        distance = request.data.get('distance')
        quad_segs = request.data.get('quad_segs', 8)

        if not geometry_data:
            return Response({'error': 'geometry is required'}, status=400)
        if distance is None:
            return Response({'error': 'distance (in meters) is required'}, status=400)

        try:
            geom = GEOSGeometry(json.dumps(geometry_data))
            buffered, stats = buffer_geometry(geom, distance, quad_segs)

            return Response({
                'geometry': json.loads(buffered.geojson),
                'stats': stats
            })

        except Exception as e:
            return Response({'error': str(e)}, status=400)
